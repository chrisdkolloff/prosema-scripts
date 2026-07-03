"""
Artikelnummer-Generator (PROSEMA / DURAL Masterliste).

Format:  MMM.SSS.NNNN
  MMM  = Hauptgruppe         (main group, 3 digits, read from column F)
  SSS  = Unterartikelgruppe  (sub group,  3 digits, read from column G)
  NNNN = laufende Nummer      (running number, unique WITHIN each MMM.SSS)

Key properties:
  * The running number resets per (Hauptgruppe, Unterartikelgruppe) pair.
  * Idempotent: rows that already carry a valid number keep it; only blank rows
    are filled, and new numbers continue past the highest existing number in
    that group. Re-running the script therefore never re-issues an identifier.
  * The whole scheme (widths, separator, start, step, columns) lives in one
    Scheme object, so extending or changing it is a one-line edit.
  * Malformed groups or capacity overflow raise instead of writing corrupt IDs.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string


@dataclass(frozen=True)
class Scheme:
    # --- field widths -----------------------------------------------------
    main_width: int = 3
    sub_width: int = 3
    running_width: int = 4          # "4 Dezimalstellen"; note mask said ZZZZZ (5)
    separator: str = "."
    # --- running-number policy -------------------------------------------
    start: int = 10                 # first number in a group
    step: int = 10                  # gap between numbers (room to insert later)
    # --- sheet layout -----------------------------------------------------
    main_group_col: str = "F"
    sub_group_col: str = "G"
    header_row: int = 1
    first_data_row: int = 2
    article_col_header: str = "PROSEMA Artikelnummer"
    data_check_columns: int = 7     # a row counts as data if A..this has content
    # --- fallbacks (used only when F/G are empty for a row) --------------
    default_main: str = "010"
    default_sub: str = "010"

    @property
    def max_running(self) -> int:
        return 10 ** self.running_width - 1

    def normalize_group(self, value, width: int, default: str, where: str) -> str:
        raw = "" if value is None else str(value).strip()
        if raw == "":
            raw = default
        if not raw.isdigit():
            raise ValueError(f"Ungültige Gruppe {raw!r} in {where} (nur Ziffern erlaubt).")
        if len(raw) > width:
            raise ValueError(f"Gruppe {raw!r} in {where} ist länger als {width} Stellen.")
        return f"{int(raw):0{width}d}"

    def format(self, main: str, sub: str, running: int) -> str:
        return f"{main}{self.separator}{sub}{self.separator}{running:0{self.running_width}d}"

    def pattern(self) -> re.Pattern:
        s = re.escape(self.separator)
        return re.compile(
            rf"^(\d{{{self.main_width}}}){s}(\d{{{self.sub_width}}}){s}(\d{{{self.running_width}}})$"
        )


def is_data_row(ws, row: int, scheme: Scheme) -> bool:
    return any(
        ws.cell(row=row, column=col).value not in (None, "")
        for col in range(1, scheme.data_check_columns + 1)
    )


def find_or_create_column(ws, header_name: str, header_row: int) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == header_name:
            return col
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header_name
    ws.cell(row=header_row, column=new_col).font = Font(bold=True)
    return new_col


def assign_article_numbers(
    input_file: str,
    output_file: str,
    scheme: Scheme = Scheme(),
    *,
    sheet_name: str | None = None,
    overwrite_existing: bool = False,
):
    """Returns (assigned_count, {group: (min, max)}). Does not mutate the input file."""
    wb = load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    main_idx = column_index_from_string(scheme.main_group_col)
    sub_idx = column_index_from_string(scheme.sub_group_col)
    article_idx = find_or_create_column(ws, scheme.article_col_header, scheme.header_row)
    pattern = scheme.pattern()

    counters: dict[tuple[str, str], int] = {}   # (main, sub) -> highest number used

    # Pass 1: register already-assigned numbers so we never collide or re-issue.
    if not overwrite_existing:
        for row in range(scheme.first_data_row, ws.max_row + 1):
            val = ws.cell(row=row, column=article_idx).value
            m = pattern.match(str(val).strip()) if val is not None else None
            if m:
                key = (m.group(1), m.group(2))
                counters[key] = max(counters.get(key, 0), int(m.group(3)))

    # Pass 2: fill blanks.
    assigned = 0
    for row in range(scheme.first_data_row, ws.max_row + 1):
        if not is_data_row(ws, row, scheme):
            continue
        existing = ws.cell(row=row, column=article_idx).value
        if not overwrite_existing and existing is not None and pattern.match(str(existing).strip()):
            continue

        main = scheme.normalize_group(ws.cell(row=row, column=main_idx).value,
                                      scheme.main_width, scheme.default_main, f"Zeile {row}, Spalte {scheme.main_group_col}")
        sub = scheme.normalize_group(ws.cell(row=row, column=sub_idx).value,
                                     scheme.sub_width, scheme.default_sub, f"Zeile {row}, Spalte {scheme.sub_group_col}")
        key = (main, sub)

        current = counters.get(key)
        nxt = scheme.start if current is None else current + scheme.step
        if nxt > scheme.max_running:
            raise OverflowError(
                f"Gruppe {main}.{sub} hat das Maximum {scheme.max_running:0{scheme.running_width}d} "
                f"in Zeile {row} überschritten. running_width erhöhen."
            )
        counters[key] = nxt
        ws.cell(row=row, column=article_idx).value = scheme.format(main, sub, nxt)
        assigned += 1

    wb.save(output_file)
    ranges = {f"{k[0]}.{k[1]}": v for k, v in sorted(counters.items())}
    return assigned, ranges


def main():
    INPUT = "input.xlsx"
    OUTPUT = "output_mit_artikelnummern.xlsx"
    if not Path(INPUT).exists():
        sys.exit(f"Eingabedatei nicht gefunden: {INPUT}")
    try:
        assigned, ranges = assign_article_numbers(INPUT, OUTPUT)
    except PermissionError:
        sys.exit(f"Konnte {OUTPUT} nicht speichern - ist die Datei in Excel geöffnet?")
    except (ValueError, OverflowError) as e:
        sys.exit(f"Abbruch: {e}")
    print(f"Fertig: {OUTPUT}  ({assigned} neue Nummern vergeben)")
    for grp, high in ranges.items():
        print(f"  {grp}: bis {high:04d}")


if __name__ == "__main__":
    main()
