"""
Interaktives Plotly-Dashboard für die Excel-Masterliste oder weclapp-Artikel-CSV.

Liest alle Zeilen und Spalten der Eingabedatei und erzeugt eine HTML-Seite mit
Kategorie-Filtern, Textsuche, Übersichtsdiagrammen und einer durchsuchbaren Tabelle.
"""

from __future__ import annotations

import csv
import json
import sys
import webbrowser
from pathlib import Path

from openpyxl import load_workbook

try:
    import plotly.graph_objects as go
except ImportError:  # pragma: no cover - handled at runtime
    go = None  # type: ignore[assignment]


FILTER_COLUMNS = (
    "Hauptgruppe",
    "Untergruppe",
    "Shopify Produkte",
    "Zielgruppe",
    "Name Lieferant",
    "Rabattkategorie_Lieferant",
    "Produktfamilie",
    "weclapp Aktiv",
    "weclapp Artikeltyp",
)

TABLE_COLUMNS = (
    "Prosema Artikelnummer",
    "Lieferanten-Art.-Nr.",
    "PROSEMA Kurztext",
    "Hauptgruppe",
    "Untergruppe",
    "Shopify Produkte",
    "Zielgruppe",
    "Name Lieferant",
    "Verkaufspreis Prosema CHF",
    "Grundmaterial",
    "Farbe",
)

SEARCH_COLUMNS = (
    "Prosema Artikelnummer",
    "Lieferanten-Art.-Nr.",
    "PROSEMA Kurztext",
    "PROSEMA Langtext",
    "Matchcode",
    "weclapp Artikel-ID",
    "weclapp Produkt-ID (Prosema)",
    "weclapp Varianten-ID (Prosema)",
)


def _project_root() -> Path:
    from scripts.paths import PROJECT_ROOT

    return PROJECT_ROOT


def _resolve_path(path: str | Path) -> Path:
    from scripts.paths import resolve_path

    return resolve_path(path)


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _ordered_master_headers(headers: list[str]) -> list[str]:
    from scripts.weclapp.master_columns import EXPORT_COLUMNS

    preferred = [column for column in EXPORT_COLUMNS if column in headers]
    extras = [header for header in headers if header not in preferred]
    return preferred + extras


def load_csv_data(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with open(path, encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        if not reader.fieldnames:
            raise ValueError(f"CSV enthält keine Spaltenüberschriften: {path}")
        headers = [header.strip() for header in reader.fieldnames]
        rows: list[dict[str, str]] = []
        for row in reader:
            record = {key: (row.get(key) or "").strip() for key in headers}
            if any(record.values()):
                rows.append(record)
    return headers, rows


def load_xlsx_data(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter)
        headers = ["" if cell is None else str(cell).strip() for cell in header_row]
        if not any(headers):
            raise ValueError(f"Masterdatei enthält keine Spaltenüberschriften: {path}")

        rows: list[dict[str, str]] = []
        for row in row_iter:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            record = {
                headers[idx]: _cell_text(row[idx]) if idx < len(row) else ""
                for idx in range(len(headers))
            }
            rows.append(record)
        return headers, rows
    finally:
        wb.close()


def load_master_data(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        raise FileNotFoundError(f"Eingabedatei nicht gefunden: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        from scripts.weclapp.master_columns import (
            EXPORT_COLUMNS,
            apply_master_column_renames,
            flat_weclapp_csv_to_master_row,
            is_master_format,
            is_raw_weclapp_export,
        )

        headers, rows = load_csv_data(path)
        if is_raw_weclapp_export(headers):
            rows = [flat_weclapp_csv_to_master_row(row) for row in rows]
            headers = list(EXPORT_COLUMNS)
        elif is_master_format(headers):
            headers = _ordered_master_headers(headers)
        return apply_master_column_renames(headers, rows)

    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        from scripts.weclapp.master_columns import apply_master_column_renames

        headers, rows = load_xlsx_data(path)
        return apply_master_column_renames(headers, rows)

    raise ValueError(
        f"Nicht unterstütztes Dateiformat: {path.suffix}. "
        "Erwartet werden .csv oder .xlsx."
    )


def _count_by(rows: list[dict[str, str]], column: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        key = row.get(column, "") or "(leer)"
        counts[key] = counts.get(key, 0) + 1
    return counts


def build_treemap_figure(rows: list[dict[str, str]]) -> go.Figure:
    main_counts = _count_by(rows, "Hauptgruppe")
    sorted_main = sorted(main_counts.items(), key=lambda item: (-item[1], item[0]))

    treemap_ids = ["root"]
    treemap_labels = ["Masterliste"]
    treemap_parents = [""]
    treemap_values = [len(rows)]

    for main_name, main_count in sorted_main:
        main_id = f"main::{main_name}"
        treemap_ids.append(main_id)
        treemap_labels.append(main_name)
        treemap_parents.append("root")
        treemap_values.append(main_count)

        sub_counts: dict[str, int] = {}
        for row in rows:
            if (row.get("Hauptgruppe", "") or "(leer)") != main_name:
                continue
            sub_name = row.get("Untergruppe", "") or "(leer)"
            sub_counts[sub_name] = sub_counts.get(sub_name, 0) + 1

        for sub_name, sub_count in sorted(
            sub_counts.items(), key=lambda item: (-item[1], item[0])
        ):
            treemap_ids.append(f"sub::{main_name}::{sub_name}")
            treemap_labels.append(sub_name)
            treemap_parents.append(main_id)
            treemap_values.append(sub_count)

    treemap_fig = go.Figure(
        go.Treemap(
            ids=treemap_ids,
            labels=treemap_labels,
            parents=treemap_parents,
            values=treemap_values,
            branchvalues="total",
            hovertemplate="<b>%{label}</b><br>%{value} Artikel<extra></extra>",
        )
    )
    treemap_fig.update_layout(
        title="Haupt- und Untergruppen",
        margin=dict(l=10, r=10, t=50, b=10),
        height=520,
    )
    return treemap_fig


def _figure_json(fig: go.Figure) -> str:
    return fig.to_json()


def _filter_options(rows: list[dict[str, str]], column: str) -> list[str]:
    values = sorted({row.get(column, "") or "(leer)" for row in rows}, key=str.casefold)
    return values


def _dashboard_columns(headers: list[str]) -> list[str]:
    from scripts.weclapp.master_columns import EXPORT_COLUMNS, EXPORT_DISPLAY_COLUMNS

    preferred = [column for column in EXPORT_DISPLAY_COLUMNS if column in headers]
    preferred.extend(
        column
        for column in EXPORT_COLUMNS
        if column in headers and column not in preferred
    )
    fallback = [
        column
        for column in (*TABLE_COLUMNS, *FILTER_COLUMNS, *SEARCH_COLUMNS)
        if column in headers and column not in preferred
    ]
    extras = [header for header in headers if header not in preferred and header not in fallback]
    return preferred + fallback + extras


def _slim_rows(rows: list[dict[str, str]], columns: list[str]) -> list[dict[str, str]]:
    return [{column: row.get(column, "") for column in columns} for row in rows]


def _dashboard_html(
    *,
    source_name: str,
    column_count: int,
    columns: list[str],
    rows: list[dict[str, str]],
    treemap_fig: go.Figure,
) -> str:
    payload = {
        "sourceName": source_name,
        "columnCount": column_count,
        "rows": rows,
        "filterColumns": [column for column in FILTER_COLUMNS if column in columns],
        "tableColumns": [column for column in TABLE_COLUMNS if column in columns],
        "availableColumns": columns,
        "searchColumns": [column for column in SEARCH_COLUMNS if column in columns],
        "filterOptions": {
            column: _filter_options(rows, column)
            for column in FILTER_COLUMNS
            if column in columns
        },
        "treemapFigure": json.loads(_figure_json(treemap_fig)),
    }
    data_json = json.dumps(payload, ensure_ascii=False)
    data_json = data_json.replace("</", "<\\/")

    return f"""<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Masterliste — Dashboard</title>
  <script src="https://cdn.plot.ly/plotly-3.6.0.min.js"></script>
  <style>
    :root {{
      color-scheme: light dark;
      --bg: #f4f6f8;
      --panel: #ffffff;
      --text: #1f2933;
      --muted: #52606d;
      --border: #d9e2ec;
      --accent: #2e7d32;
    }}
    @media (prefers-color-scheme: dark) {{
      :root {{
        --bg: #111827;
        --panel: #1f2937;
        --text: #f9fafb;
        --muted: #9ca3af;
        --border: #374151;
        --accent: #4ade80;
      }}
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    header {{
      padding: 20px 24px 8px;
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 1.5rem;
    }}
    .subtitle {{
      color: var(--muted);
      font-size: 0.95rem;
    }}
    .stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      padding: 0 24px 16px;
    }}
    .stat {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 12px 16px;
      min-width: 140px;
    }}
    .stat strong {{
      display: block;
      font-size: 1.35rem;
    }}
    .stat span {{
      color: var(--muted);
      font-size: 0.85rem;
    }}
    .filters {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 12px;
      padding: 0 24px 16px;
    }}
    .filters label {{
      display: flex;
      flex-direction: column;
      gap: 6px;
      font-size: 0.85rem;
      color: var(--muted);
    }}
    select, input[type="search"], button {{
      font: inherit;
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid var(--border);
      background: var(--panel);
      color: var(--text);
    }}
    .search-row {{
      display: flex;
      gap: 12px;
      align-items: end;
      padding: 0 24px 16px;
      flex-wrap: wrap;
    }}
    .search-row label {{
      flex: 1 1 320px;
      display: flex;
      flex-direction: column;
      gap: 6px;
      font-size: 0.85rem;
      color: var(--muted);
    }}
    button {{
      cursor: pointer;
      background: var(--accent);
      color: #fff;
      border: none;
      min-width: 120px;
    }}
    .charts {{
      padding: 0 24px 16px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 8px;
      overflow: hidden;
    }}
    .table-wrap {{
      margin: 0 24px 24px;
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      overflow: auto;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 0.88rem;
    }}
    thead th {{
      position: sticky;
      top: 0;
      background: var(--panel);
      border-bottom: 1px solid var(--border);
      text-align: left;
      padding: 10px 12px;
      white-space: nowrap;
    }}
    tbody td {{
      border-bottom: 1px solid var(--border);
      padding: 8px 12px;
      vertical-align: top;
    }}
    tbody tr:hover {{
      background: rgba(46, 125, 50, 0.08);
    }}
    .table-meta {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      padding: 12px 16px;
      color: var(--muted);
      font-size: 0.9rem;
      flex-wrap: wrap;
    }}
    .pager {{
      display: flex;
      gap: 8px;
      align-items: center;
    }}
    .pager button {{
      min-width: auto;
      padding: 6px 12px;
      background: var(--panel);
      color: var(--text);
      border: 1px solid var(--border);
    }}
    .pager button:disabled {{
      opacity: 0.45;
      cursor: default;
    }}
    .table-toolbar {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      padding: 12px 16px 0;
      flex-wrap: wrap;
    }}
    .table-toolbar-actions {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .table-toolbar button.secondary {{
      background: var(--panel);
      color: var(--text);
      border: 1px solid var(--border);
      min-width: auto;
    }}
    .toolbar-summary {{
      font-size: 0.9rem;
      color: var(--muted);
    }}
    .column-panel {{
      margin: 0 16px 12px;
      padding: 12px;
      border: 1px solid var(--border);
      border-radius: 10px;
      background: var(--bg);
    }}
    .column-panel-header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      margin-bottom: 10px;
      flex-wrap: wrap;
    }}
    .column-panel-title {{
      font-size: 0.9rem;
      color: var(--muted);
    }}
    .column-actions {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .column-actions button {{
      min-width: auto;
      padding: 6px 10px;
      font-size: 0.85rem;
      background: var(--panel);
      color: var(--text);
      border: 1px solid var(--border);
    }}
    .column-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(220px, 1fr));
      gap: 6px 12px;
      max-height: 220px;
      overflow: auto;
    }}
    .column-grid label {{
      display: flex;
      align-items: flex-start;
      gap: 8px;
      font-size: 0.85rem;
      color: var(--text);
      cursor: pointer;
    }}
    .column-grid input {{
      margin-top: 2px;
    }}
    .column-text {{
      display: inline-flex;
      align-items: baseline;
      gap: 6px;
      flex-wrap: wrap;
    }}
    .column-panel[hidden] {{
      display: none;
    }}
  </style>
</head>
<body>
  <header>
    <h1>Masterliste Dashboard</h1>
    <div class="subtitle" id="source-label"></div>
  </header>

  <section class="stats" id="stats"></section>

  <section class="filters" id="filter-bar"></section>

  <div class="search-row">
    <label>
      Textsuche
      <input id="search-input" type="search" placeholder="Artikelnummer, Kurztext, Beschreibung …" />
    </label>
    <button type="button" id="reset-filters">Filter zurücksetzen</button>
  </div>

  <section class="charts">
    <div class="panel"><div id="treemap-chart" style="height:520px;"></div></div>
  </section>

  <div class="table-wrap">
    <div class="table-toolbar">
      <div class="table-toolbar-actions">
        <button type="button" class="secondary" id="toggle-columns">Spalten auswählen</button>
        <button type="button" id="export-csv">Als CSV exportieren</button>
      </div>
      <div id="column-selection-summary" class="toolbar-summary"></div>
    </div>
    <div class="column-panel" id="column-panel" hidden>
      <div class="column-panel-header">
        <div class="column-panel-title">Angezeigte Spalten (werden auch exportiert)</div>
        <div class="column-actions">
          <button type="button" class="secondary" id="columns-default">Standard</button>
          <button type="button" class="secondary" id="columns-all">Alle</button>
          <button type="button" class="secondary" id="columns-none">Keine</button>
        </div>
      </div>
      <div class="column-grid" id="column-grid"></div>
    </div>
    <div class="table-meta">
      <div id="table-summary"></div>
      <div class="pager">
        <button type="button" id="prev-page">Zurück</button>
        <span id="page-label"></span>
        <button type="button" id="next-page">Weiter</button>
      </div>
    </div>
    <table>
      <thead id="table-head"></thead>
      <tbody id="table-body"></tbody>
    </table>
  </div>

  <script id="dashboard-data" type="application/json">{data_json}</script>
  <script>
    const DATA = JSON.parse(document.getElementById("dashboard-data").textContent);
    const PAGE_SIZE = 75;
    let filteredRows = DATA.rows.slice();
    let currentPage = 1;
    let visibleColumns = DATA.tableColumns.slice();
    const activeFilters = Object.fromEntries(DATA.filterColumns.map((col) => [col, ""]));

    function escapeHtml(value) {{
      return String(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function csvEscape(value) {{
      const text = String(value ?? "");
      if (/[;"\\n\\r]/.test(text)) {{
        return `"${{text.replaceAll('"', '""')}}"`;
      }}
      return text;
    }}

    function updateColumnSelectionSummary() {{
      const summary = document.getElementById("column-selection-summary");
      summary.textContent =
        `${{visibleColumns.length}} von ${{DATA.availableColumns.length}} Spalten`;
    }}

    function buildColumnPicker() {{
      const grid = document.getElementById("column-grid");
      grid.innerHTML = "";
      for (const column of DATA.availableColumns) {{
        const label = document.createElement("label");
        const checkbox = document.createElement("input");
        checkbox.type = "checkbox";
        checkbox.value = column;
        checkbox.checked = visibleColumns.includes(column);
        checkbox.addEventListener("change", () => {{
          let next = visibleColumns.slice();
          if (checkbox.checked) {{
            if (!next.includes(column)) next.push(column);
          }} else {{
            next = next.filter((col) => col !== column);
          }}
          if (next.length === 0) {{
            checkbox.checked = true;
            next = [column];
          }}
          visibleColumns = setSelectedColumns(next, DATA.availableColumns, DATA.availableColumns[0]);
          syncColumnCheckboxes();
          updateColumnSelectionSummary();
          renderTable();
        }});
        const text = document.createElement("span");
        text.className = "column-text";
        const name = document.createElement("span");
        name.textContent = column;
        text.appendChild(name);
        label.appendChild(checkbox);
        label.appendChild(text);
        grid.appendChild(label);
      }}
    }}

    function syncColumnCheckboxes() {{
      for (const input of document.querySelectorAll("#column-grid input[type=checkbox]")) {{
        input.checked = visibleColumns.includes(input.value);
      }}
    }}

    function setSelectedColumns(selectedColumns, allColumns, fallback) {{
      const next = selectedColumns.length ? selectedColumns.slice() : [fallback];
      next.sort((a, b) => allColumns.indexOf(a) - allColumns.indexOf(b));
      return next;
    }}

    function setVisibleColumns(columns) {{
      visibleColumns = setSelectedColumns(columns, DATA.availableColumns, DATA.availableColumns[0]);
      syncColumnCheckboxes();
      updateColumnSelectionSummary();
      renderTable();
    }}

    function matchesSearch(row, query) {{
      if (!query) return true;
      const needle = query.toLowerCase();
      return DATA.searchColumns.some((col) => (row[col] || "").toLowerCase().includes(needle));
    }}

    function applyFilters() {{
      const query = document.getElementById("search-input").value.trim();
      filteredRows = DATA.rows.filter((row) => {{
        if (!matchesSearch(row, query)) return false;
        return DATA.filterColumns.every((col) => {{
          const selected = activeFilters[col];
          if (!selected) return true;
          const value = row[col] || "(leer)";
          return value === selected;
        }});
      }});
      currentPage = 1;
      renderAll();
    }}

    function buildFilterBar() {{
      const container = document.getElementById("filter-bar");
      container.innerHTML = "";
      for (const col of DATA.filterColumns) {{
        const options = DATA.filterOptions[col] || [];
        const label = document.createElement("label");
        label.textContent = col;
        const select = document.createElement("select");
        select.dataset.column = col;
        const allOption = document.createElement("option");
        allOption.value = "";
        allOption.textContent = "Alle";
        select.appendChild(allOption);
        for (const value of options) {{
          const option = document.createElement("option");
          option.value = value;
          option.textContent = value;
          select.appendChild(option);
        }}
        select.addEventListener("change", () => {{
          activeFilters[col] = select.value;
          applyFilters();
        }});
        label.appendChild(select);
        container.appendChild(label);
      }}
    }}

    function renderStats() {{
      const stats = document.getElementById("stats");
      const mainGroups = new Set(filteredRows.map((row) => row.Hauptgruppe || "(leer)")).size;
      const subGroups = new Set(
        filteredRows.map((row) => `${{row.Hauptgruppe || ""}}|${{row.Untergruppe || ""}}`)
      ).size;
      const withArticleNo = filteredRows.filter((row) => row["Prosema Artikelnummer"]).length;
      stats.innerHTML = `
        <div class="stat"><strong>${{filteredRows.length}}</strong><span>Artikel (gefiltert)</span></div>
        <div class="stat"><strong>${{DATA.rows.length}}</strong><span>Artikel (gesamt)</span></div>
        <div class="stat"><strong>${{mainGroups}}</strong><span>Hauptgruppen</span></div>
        <div class="stat"><strong>${{subGroups}}</strong><span>Untergruppen</span></div>
        <div class="stat"><strong>${{withArticleNo}}</strong><span>mit Artikelnummer</span></div>
      `;
    }}

    function countBy(rows, column) {{
      const counts = new Map();
      for (const row of rows) {{
        const key = row[column] || "(leer)";
        counts.set(key, (counts.get(key) || 0) + 1);
      }}
      return [...counts.entries()].sort((a, b) => b[1] - a[1] || a[0].localeCompare(b[0], "de"));
    }}

    function renderCharts() {{
      const mainCounts = countBy(filteredRows, "Hauptgruppe");

      const treemapIds = ["root"];
      const treemapLabels = ["Auswahl"];
      const treemapParents = [""];
      const treemapValues = [filteredRows.length || 1];

      for (const [mainName, mainCount] of mainCounts) {{
        const mainId = `main::${{mainName}}`;
        treemapIds.push(mainId);
        treemapLabels.push(mainName);
        treemapParents.push("root");
        treemapValues.push(mainCount);

        const subCounts = new Map();
        for (const row of filteredRows) {{
          if ((row.Hauptgruppe || "(leer)") !== mainName) continue;
          const subName = row.Untergruppe || "(leer)";
          subCounts.set(subName, (subCounts.get(subName) || 0) + 1);
        }}
        for (const [subName, subCount] of [...subCounts.entries()].sort((a, b) => b[1] - a[1])) {{
          treemapIds.push(`sub::${{mainName}}::${{subName}}`);
          treemapLabels.push(subName);
          treemapParents.push(mainId);
          treemapValues.push(subCount);
        }}
      }}

      const treemapFigure = JSON.parse(JSON.stringify(DATA.treemapFigure));
      treemapFigure.data[0].ids = treemapIds;
      treemapFigure.data[0].labels = treemapLabels;
      treemapFigure.data[0].parents = treemapParents;
      treemapFigure.data[0].values = treemapValues;
      Plotly.react("treemap-chart", treemapFigure.data, treemapFigure.layout, {{ responsive: true }});
    }}

    function renderTable() {{
      const columns = visibleColumns;
      const head = document.getElementById("table-head");
      head.innerHTML = `<tr>${{columns.map((col) => `<th>${{escapeHtml(col)}}</th>`).join("")}}</tr>`;

      const totalPages = Math.max(1, Math.ceil(filteredRows.length / PAGE_SIZE));
      if (currentPage > totalPages) currentPage = totalPages;
      const start = (currentPage - 1) * PAGE_SIZE;
      const pageRows = filteredRows.slice(start, start + PAGE_SIZE);

      const body = document.getElementById("table-body");
      body.innerHTML = pageRows.map((row) => `
        <tr>${{columns.map((col) => `<td>${{escapeHtml(row[col] || "")}}</td>`).join("")}}</tr>
      `).join("");

      document.getElementById("table-summary").textContent =
        `${{filteredRows.length}} Zeilen · Seite ${{currentPage}} von ${{totalPages}}`;
      document.getElementById("page-label").textContent = `${{currentPage}} / ${{totalPages}}`;
      document.getElementById("prev-page").disabled = currentPage <= 1;
      document.getElementById("next-page").disabled = currentPage >= totalPages;
    }}

    function setFilter(column, value) {{
      if (!DATA.filterColumns.includes(column)) return;
      activeFilters[column] = value;
      const select = document.querySelector(`select[data-column="${{column}}"]`);
      if (select) select.value = value;
      applyFilters();
    }}

    function bindTreemapClicks() {{
      const chart = document.getElementById("treemap-chart");
      chart.on("plotly_click", (event) => {{
        const label = event.points?.[0]?.label;
        const id = event.points?.[0]?.id || "";
        if (!label || label === "Auswahl") {{
          DATA.filterColumns.forEach((col) => setFilter(col, ""));
          return;
        }}
        if (id.startsWith("main::")) {{
          setFilter("Hauptgruppe", label);
          setFilter("Untergruppe", "");
          return;
        }}
        if (id.startsWith("sub::")) {{
          const parts = id.split("::");
          setFilter("Hauptgruppe", parts[1] || "");
          setFilter("Untergruppe", parts[2] || "");
        }}
      }});
    }}

    function exportCsv() {{
      const columns = visibleColumns;
      const lines = [
        columns.map(csvEscape).join(";"),
        ...filteredRows.map((row) =>
          columns.map((col) => csvEscape(row[col] || "")).join(";")
        ),
      ];
      const blob = new Blob(["\ufeff" + lines.join("\\n")], {{
        type: "text/csv;charset=utf-8",
      }});
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      const stamp = new Date().toISOString().slice(0, 19).replace(/[:T]/g, "-");
      link.href = url;
      link.download = `masterliste-export-${{stamp}}.csv`;
      link.click();
      URL.revokeObjectURL(url);
    }}

    function renderAll() {{
      renderStats();
      renderCharts();
      renderTable();
    }}

    document.getElementById("source-label").textContent =
      `${{DATA.sourceName}} · ${{DATA.columnCount}} Spalten · ${{DATA.rows.length}} Zeilen`;
    document.getElementById("search-input").addEventListener("input", applyFilters);
    document.getElementById("toggle-columns").addEventListener("click", () => {{
      const panel = document.getElementById("column-panel");
      const open = panel.hasAttribute("hidden");
      panel.toggleAttribute("hidden", !open);
      document.getElementById("toggle-columns").textContent =
        open ? "Spalten ausblenden" : "Spalten auswählen";
    }});
    document.getElementById("columns-default").addEventListener("click", () => {{
      setVisibleColumns(DATA.tableColumns);
    }});
    document.getElementById("columns-all").addEventListener("click", () => {{
      setVisibleColumns(DATA.availableColumns);
    }});
    document.getElementById("columns-none").addEventListener("click", () => {{
      setVisibleColumns([DATA.availableColumns[0]]);
    }});
    document.getElementById("export-csv").addEventListener("click", exportCsv);
    document.getElementById("reset-filters").addEventListener("click", () => {{
      DATA.filterColumns.forEach((col) => {{
        activeFilters[col] = "";
        const select = document.querySelector(`select[data-column="${{col}}"]`);
        if (select) select.value = "";
      }});
      document.getElementById("search-input").value = "";
      applyFilters();
    }});
    document.getElementById("prev-page").addEventListener("click", () => {{
      currentPage -= 1;
      renderTable();
    }});
    document.getElementById("next-page").addEventListener("click", () => {{
      currentPage += 1;
      renderTable();
    }});

    buildFilterBar();
    buildColumnPicker();
    updateColumnSelectionSummary();
    Plotly.newPlot("treemap-chart", DATA.treemapFigure.data, DATA.treemapFigure.layout, {{ responsive: true }});
    bindTreemapClicks();
    renderAll();
  </script>
</body>
</html>
"""


def render_dashboard(
    headers: list[str],
    rows: list[dict[str, str]],
    output: Path,
    *,
    source_name: str,
) -> Path:
    if go is None:
        raise ImportError(
            "Das Paket 'plotly' ist nicht installiert. "
            "Bitte ausführen: pip install plotly"
        )

    treemap_fig = build_treemap_figure(rows)
    dashboard_columns = _dashboard_columns(headers)
    output.parent.mkdir(parents=True, exist_ok=True)
    page = _dashboard_html(
        source_name=source_name,
        column_count=len(headers),
        columns=dashboard_columns,
        rows=_slim_rows(rows, dashboard_columns),
        treemap_fig=treemap_fig,
    )
    output.write_text(page, encoding="utf-8")
    return output


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    input_path = _resolve_path(params["input"])
    output_path = _resolve_path(params["output"])

    headers, rows = load_master_data(input_path)
    is_weclapp = input_path.suffix.lower() == ".csv"
    source_label = f"{input_path.name} (weclapp)" if is_weclapp else input_path.name
    output_file = render_dashboard(
        headers,
        rows,
        output_path,
        source_name=source_label,
    )

    opened = webbrowser.open(output_file.resolve().as_uri())
    details = [
        f"Zeilen: {len(rows)}",
        f"Spalten: {len(headers)}",
        f"Filter: {', '.join(col for col in FILTER_COLUMNS if col in headers)}",
    ]
    if not opened:
        details.append("Hinweis: Browser konnte nicht automatisch geöffnet werden.")
    return RunResult(
        summary=f"Fertig: {output_file}",
        details=details,
        show_success_dialog=False,
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="master_dashboard",
        title="Masterliste-Dashboard",
        description=(
            "Interaktives Plotly-Dashboard aus der Excel-Masterliste oder dem "
            "weclapp-Artikel-CSV-Snapshot erzeugen. Filtern nach Kategorien, "
            "Spaltenauswahl, CSV-Export und paginierte Tabellenansicht."
        ),
        fields=(
            FieldSpec(
                "input",
                "Eingabedatei",
                FieldKind.FILE_IN,
                "output/export/weclapp_export.csv",
            ),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "output/reports/master_dashboard.html",
                output_name="master_dashboard.html",
            ),
        ),
        run=run_job,
    )


def main() -> None:
    _ensure_project_root()
    from gui.job_spec import args_to_params, build_argparser, coerce_params, validate_params

    parser = build_argparser(JOB_SPEC)
    args = parser.parse_args()
    params = coerce_params(JOB_SPEC, args_to_params(JOB_SPEC, args))
    try:
        validate_params(JOB_SPEC, params)
        result = run_job(params)
    except (FileNotFoundError, ValueError, ImportError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    print(result.summary)
    for line in result.details:
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
