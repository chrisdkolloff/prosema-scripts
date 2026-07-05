"""
Gruppenhierarchie aus gruppen.xlsx als interaktives Plotly-Diagramm erzeugen.

Liest die Tabellenblätter „Hauptgruppen“ und „Untergruppen“ und rendert eine
Sunburst-Ansicht: Hauptgruppen im Inneren, Untergruppen im äußeren Ring.
"""

from __future__ import annotations

import sys
import webbrowser
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

try:
    import plotly.graph_objects as go
except ImportError:  # pragma: no cover - handled at runtime
    go = None  # type: ignore[assignment]


@dataclass(frozen=True)
class Subgroup:
    main_code: str
    sub_code: str
    name: str


ROOT_ID = "root"


def _project_root() -> Path:
    from scripts.paths import PROJECT_ROOT

    return PROJECT_ROOT


def _resolve_path(path: str | Path) -> Path:
    from scripts.paths import resolve_path

    return resolve_path(path)


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def _format_code(value) -> str:
    raw = "" if value is None else str(value).strip()
    if raw == "":
        return ""
    if not raw.isdigit():
        raise ValueError(f"Ungültiger Gruppencode: {raw!r}")
    return f"{int(raw):03d}"


def _label(code: str, name: str) -> str:
    if name:
        return f"{code} {name}"
    return code


def _code_sort_key(code: str) -> int:
    return int(code)


def load_groups(path: Path) -> tuple[dict[str, str], list[Subgroup]]:
    if not path.exists():
        raise FileNotFoundError(f"Gruppenschlüssel nicht gefunden: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Hauptgruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Hauptgruppen".')
        if "Untergruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Untergruppen".')

        main_groups: dict[str, str] = {}
        for code_raw, name_raw in wb["Hauptgruppen"].iter_rows(
            min_row=2, max_col=2, values_only=True
        ):
            code = _format_code(code_raw)
            if code == "":
                continue
            name = "" if name_raw is None else str(name_raw).strip()
            main_groups[code] = name

        subgroups: list[Subgroup] = []
        for main_raw, sub_raw, name_raw, *_ in wb["Untergruppen"].iter_rows(
            min_row=2, max_col=3, values_only=True
        ):
            main_code = _format_code(main_raw)
            sub_code = _format_code(sub_raw)
            if main_code == "" or sub_code == "":
                continue
            name = "" if name_raw is None else str(name_raw).strip()
            if name == "":
                raise ValueError(
                    f"Leere Untergruppen-Bezeichnung für {main_code}.{sub_code}."
                )
            subgroups.append(Subgroup(main_code, sub_code, name))

        return main_groups, subgroups
    finally:
        wb.close()


def build_figure(
    main_groups: dict[str, str],
    subgroups: list[Subgroup],
):
    by_main: dict[str, list[Subgroup]] = {}
    for subgroup in subgroups:
        by_main.setdefault(subgroup.main_code, []).append(subgroup)

    ids = [ROOT_ID]
    labels = ["Produktgruppen"]
    parents = [""]
    values = [len(subgroups)]

    for main_code in sorted(by_main, key=_code_sort_key):
        main_id = f"main_{main_code}"
        main_name = main_groups.get(main_code, "")
        ids.append(main_id)
        labels.append(_label(main_code, main_name))
        parents.append(ROOT_ID)
        values.append(len(by_main[main_code]))

        for subgroup in sorted(by_main[main_code], key=lambda s: _code_sort_key(s.sub_code)):
            sub_id = f"sub_{main_code}_{subgroup.sub_code}"
            ids.append(sub_id)
            labels.append(_label(subgroup.sub_code, subgroup.name))
            parents.append(main_id)
            values.append(1)

    fig = go.Figure(
        go.Sunburst(
            ids=ids,
            labels=labels,
            parents=parents,
            values=values,
            branchvalues="total",
            sort=False,
            hovertext=labels,
            hovertemplate="<b>%{hovertext}</b><extra></extra>",
            hoverlabel=dict(namelength=-1),
        )
    )
    fig.update_layout(
        title="Haupt- und Untergruppen",
        width=1400,
        height=1400,
        margin=dict(t=50, l=10, r=10, b=10),
        uniformtext=dict(minsize=10, mode="hide"),
    )
    return fig


def _output_path(output: Path, fmt: str) -> Path:
    suffix = f".{fmt}"
    if output.suffix.lower() == suffix:
        return output
    return output.with_suffix(suffix)


def render_diagram(
    main_groups: dict[str, str],
    subgroups: list[Subgroup],
    output: Path,
    fmt: str,
) -> Path:
    if go is None:
        raise ImportError(
            "Das Paket 'plotly' ist nicht installiert. "
            "Bitte ausführen: pip install plotly"
        )

    fig = build_figure(main_groups, subgroups)
    output_file = _output_path(output, fmt)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    if fmt == "html":
        fig.write_html(str(output_file), include_plotlyjs="cdn")
        return output_file

    try:
        fig.write_image(str(output_file))
    except ValueError as exc:
        raise RuntimeError(
            f"Statisches Format {fmt!r} benötigt das Paket 'kaleido'. "
            "Bitte ausführen: pip install kaleido — oder HTML verwenden (-f html)."
        ) from exc

    return output_file


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    input_path = _resolve_path(params["input"])
    output_path = _resolve_path(params["output"])

    main_groups, subgroups = load_groups(input_path)
    output_file = render_diagram(main_groups, subgroups, output_path, "html")

    opened = webbrowser.open(output_file.resolve().as_uri())
    details = [
        f"Hauptgruppen: {len(main_groups)}",
        f"Untergruppen: {len(subgroups)}",
    ]
    if not opened:
        details.append("Hinweis: Browser konnte nicht automatisch geöffnet werden.")
    return RunResult(
        summary=f"Fertig: {output_file}",
        details=details,
        show_success_dialog=False,
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="gruppen_diagram",
        title="Gruppen-Diagramm",
        description=(
            "Interaktives Diagramm der Haupt- und Untergruppen aus dem Gruppenschlüssel "
            "erzeugen und im Browser öffnen."
        ),
        fields=(
            FieldSpec(
                "input",
                "Gruppenschlüssel",
                FieldKind.FILE_IN,
                "data/gruppen.xlsx",
            ),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "output/reports/gruppen_diagram.html",
                output_name="gruppen_diagram.html",
            ),
        ),
        run=run_job,
    )


def main() -> None:
    _ensure_project_root()
    from gui.job_spec import args_to_params, build_argparser, coerce_params, validate_params

    parser = build_argparser(JOB_SPEC)
    args = parser.parse_args()
    params = coerce_params(JOB_SPEC, args_to_params(JOB_SPEC, args))
    try:
        validate_params(JOB_SPEC, params)
        result = run_job(params)
    except (FileNotFoundError, ValueError, ImportError, RuntimeError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    print(result.summary)
    for line in result.details:
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
