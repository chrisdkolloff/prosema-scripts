"""
Artikelnummer-Generator (PROSEMA / DURAL Masterliste).

Format:  MMM.SSS.NNNN
  MMM  = Hauptgruppe         (main group, 3 digits, resolved from name via Gruppenschlüssel)
  SSS  = Unterartikelgruppe  (sub group,  3 digits, resolved from name via Gruppenschlüssel)
  NNNN = laufende Nummer      (running number, unique WITHIN each MMM.SSS)

Key properties:
  * The running number resets per (Hauptgruppe, Unterartikelgruppe) pair.
  * Idempotent: rows that already carry a valid number keep it; only blank rows
    are filled, and new numbers continue past the highest existing number in
    that group. Re-running the script therefore never re-issues an identifier.
  * The whole scheme (widths, separator, start, step, columns) lives in one
    Scheme object, so extending or changing it is a one-line edit.
  * Malformed groups or capacity overflow raise instead of writing corrupt IDs.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font


@dataclass(frozen=True)
class Scheme:
    # --- field widths -----------------------------------------------------
    main_width: int = 3
    sub_width: int = 3
    running_width: int = 4          # "4 Dezimalstellen"; note mask said ZZZZZ (5)
    separator: str = "."
    # --- running-number policy -------------------------------------------
    start: int = 10                 # first number in a group
    step: int = 10                  # gap between numbers (room to insert later)
    # --- sheet layout ----------------------------------------------------
    header_row: int = 1
    first_data_row: int = 2
    article_col_header: str = "Prosema Artikelnummer"
    main_group_header: str = "Hauptgruppe"
    sub_group_header: str = "Untergruppe"
    data_row_key_header: str = "Artikelnr."
    dictionary_file: str = "gruppenschluessel.xlsx"

    @property
    def max_running(self) -> int:
        return 10 ** self.running_width - 1

    def dictionary_path(self) -> Path:
        path = Path(self.dictionary_file)
        if path.is_absolute():
            return path
        return Path(__file__).resolve().parent / path

    def normalize_group(self, value, width: int, where: str) -> str:
        raw = "" if value is None else str(value).strip()
        if raw == "":
            raise ValueError(f"Leerer Gruppencode in {where}.")
        if not raw.isdigit():
            raise ValueError(f"Ungültige Gruppe {raw!r} in {where} (nur Ziffern erlaubt).")
        if len(raw) > width:
            raise ValueError(f"Gruppe {raw!r} in {where} ist länger als {width} Stellen.")
        return f"{int(raw):0{width}d}"

    def format(self, main: str, sub: str, running: int) -> str:
        return f"{main}{self.separator}{sub}{self.separator}{running:0{self.running_width}d}"

    def pattern(self) -> re.Pattern:
        s = re.escape(self.separator)
        return re.compile(
            rf"^(\d{{{self.main_width}}}){s}(\d{{{self.sub_width}}}){s}(\d{{{self.running_width}}})$"
        )


@dataclass
class RowResolutionError:
    row: int
    main_name: str | None = None
    sub_name: str | None = None
    main_unknown: bool = False
    sub_unknown: bool = False


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def find_column(ws, header_name: str, header_row: int) -> int:
    target = normalize_header(header_name)
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=header_row, column=col).value) == target:
            return col
    present = [
        str(ws.cell(row=header_row, column=col).value).strip()
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value not in (None, "")
    ]
    raise ValueError(
        f"Spalte {header_name!r} nicht gefunden. "
        f"Vorhandene Spaltenüberschriften: {', '.join(present) or '(keine)'}"
    )


def find_or_create_column(ws, header_name: str, header_row: int) -> int:
    target = normalize_header(header_name)
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=header_row, column=col).value) == target:
            return col
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header_name
    ws.cell(row=header_row, column=new_col).font = Font(bold=True)
    return new_col


def normalize_name(value) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def load_group_dictionary(path: Path) -> tuple[dict[str, str], dict[tuple[str, str], str]]:
    if not path.exists():
        raise FileNotFoundError(
            f"Gruppenschlüssel nicht gefunden: {path}\n"
            "Die Gruppencodes können ohne diese Datei nicht aufgelöst werden."
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Hauptgruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Hauptgruppen".')
        if "Untergruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Untergruppen".')

        ws_main = wb["Hauptgruppen"]
        code_col = find_column(ws_main, "Code", 1)
        name_col = find_column(ws_main, "Bezeichnung", 1)

        main_name_to_code: dict[str, str] = {}
        for row in range(2, ws_main.max_row + 1):
            code_raw = ws_main.cell(row=row, column=code_col).value
            name_raw = ws_main.cell(row=row, column=name_col).value
            if code_raw in (None, "") or name_raw in (None, ""):
                continue
            code = str(code_raw).strip()
            key = normalize_name(name_raw)
            if key in main_name_to_code:
                raise ValueError(
                    f"Doppelte Hauptgruppen-Bezeichnung im Gruppenschlüssel: {name_raw!r} "
                    f"(Codes {main_name_to_code[key]!r} und {code!r})."
                )
            main_name_to_code[key] = code

        ws_sub = wb["Untergruppen"]
        main_code_col = find_column(ws_sub, "Hauptgruppe", 1)
        sub_code_col = find_column(ws_sub, "Untergruppe", 1)
        sub_name_col = find_column(ws_sub, "Bezeichnung", 1)

        sub_name_to_code: dict[tuple[str, str], str] = {}
        for row in range(2, ws_sub.max_row + 1):
            main_code_raw = ws_sub.cell(row=row, column=main_code_col).value
            sub_code_raw = ws_sub.cell(row=row, column=sub_code_col).value
            sub_name_raw = ws_sub.cell(row=row, column=sub_name_col).value
            if main_code_raw in (None, "") or sub_code_raw in (None, "") or sub_name_raw in (None, ""):
                continue
            main_code = str(main_code_raw).strip()
            sub_code = str(sub_code_raw).strip()
            key = (main_code, normalize_name(sub_name_raw))
            if key in sub_name_to_code:
                raise ValueError(
                    f"Doppelte Untergruppen-Bezeichnung im Gruppenschlüssel: "
                    f"Hauptgruppe {main_code!r}, Bezeichnung {sub_name_raw!r} "
                    f"(Codes {sub_name_to_code[key]!r} und {sub_code!r})."
                )
            sub_name_to_code[key] = sub_code

        return main_name_to_code, sub_name_to_code
    finally:
        wb.close()


def is_data_row(ws, row: int, data_key_col: int) -> bool:
    val = ws.cell(row=row, column=data_key_col).value
    return val is not None and str(val).strip() != ""


def resolve_group_codes(
    ws,
    row: int,
    main_col: int,
    sub_col: int,
    main_name_to_code: dict[str, str],
    sub_name_to_code: dict[tuple[str, str], str],
    scheme: Scheme,
) -> tuple[str | None, str | None, RowResolutionError | None]:
    main_raw = ws.cell(row=row, column=main_col).value
    sub_raw = ws.cell(row=row, column=sub_col).value
    main_display = "" if main_raw is None else str(main_raw).strip()
    sub_display = "" if sub_raw is None else str(sub_raw).strip()

    err = RowResolutionError(row=row, main_name=main_display or None, sub_name=sub_display or None)
    main_code: str | None = None

    if main_display == "":
        err.main_unknown = True
    else:
        main_code = main_name_to_code.get(normalize_name(main_display))
        if main_code is None:
            err.main_unknown = True

    if sub_display == "":
        err.sub_unknown = True
    elif main_code is not None:
        sub_code = sub_name_to_code.get((main_code, normalize_name(sub_display)))
        if sub_code is None:
            err.sub_unknown = True
        else:
            try:
                main = scheme.normalize_group(main_code, scheme.main_width, f"Zeile {row}, {scheme.main_group_header}")
                sub = scheme.normalize_group(sub_code, scheme.sub_width, f"Zeile {row}, {scheme.sub_group_header}")
                return main, sub, None
            except ValueError as e:
                raise ValueError(f"Zeile {row}: {e}") from e
    else:
        err.sub_unknown = True

    return None, None, err


def format_resolution_errors(errors: list[RowResolutionError]) -> str:
    lines = ["Gruppennamen konnten nicht aufgelöst werden:"]
    for err in errors:
        parts = [f"  Zeile {err.row}:"]
        if err.main_unknown:
            parts.append(f" unbekannte Hauptgruppe {err.main_name!r}")
        elif err.sub_unknown:
            parts.append(f" unbekannte Untergruppe {err.sub_name!r}")
        lines.append("".join(parts))
    lines.append("Bitte korrigieren Sie die Namen in input.xlsx oder ergänzen Sie den Gruppenschlüssel.")
    return "\n".join(lines)


def assign_article_numbers(
    input_file: str,
    output_file: str,
    scheme: Scheme = Scheme(),
    *,
    sheet_name: str | None = None,
    overwrite_existing: bool = False,
    strict: bool = True,
):
    """Returns (assigned_count, {group: (min, max)}). Does not mutate the input file."""
    main_name_to_code, sub_name_to_code = load_group_dictionary(scheme.dictionary_path())

    wb = load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    main_col = find_column(ws, scheme.main_group_header, scheme.header_row)
    sub_col = find_column(ws, scheme.sub_group_header, scheme.header_row)
    data_key_col = find_column(ws, scheme.data_row_key_header, scheme.header_row)
    article_col = find_or_create_column(ws, scheme.article_col_header, scheme.header_row)
    pattern = scheme.pattern()

    counters: dict[tuple[str, str], int] = {}   # (main, sub) -> highest number used
    resolution_errors: list[RowResolutionError] = []
    resolved_codes: dict[int, tuple[str, str]] = {}

    # Pass 1: register already-assigned numbers so we never collide or re-issue.
    if not overwrite_existing:
        for row in range(scheme.first_data_row, ws.max_row + 1):
            val = ws.cell(row=row, column=article_col).value
            m = pattern.match(str(val).strip()) if val is not None else None
            if m:
                key = (m.group(1), m.group(2))
                counters[key] = max(counters.get(key, 0), int(m.group(3)))

    # Resolve group names for all data rows.
    for row in range(scheme.first_data_row, ws.max_row + 1):
        if not is_data_row(ws, row, data_key_col):
            continue
        main, sub, err = resolve_group_codes(
            ws, row, main_col, sub_col, main_name_to_code, sub_name_to_code, scheme
        )
        if err is not None:
            resolution_errors.append(err)
        else:
            resolved_codes[row] = (main, sub)

    if resolution_errors:
        print(format_resolution_errors(resolution_errors))
        if strict:
            raise ValueError(
                f"Abbruch: {len(resolution_errors)} Zeile(n) mit unbekannten Gruppennamen "
                "(strict=True, Datei wurde nicht gespeichert)."
            )

    # Pass 2: fill blanks.
    assigned = 0
    for row in range(scheme.first_data_row, ws.max_row + 1):
        if not is_data_row(ws, row, data_key_col):
            continue
        if row not in resolved_codes:
            continue

        existing = ws.cell(row=row, column=article_col).value
        if not overwrite_existing and existing is not None and pattern.match(str(existing).strip()):
            continue

        main, sub = resolved_codes[row]
        key = (main, sub)

        current = counters.get(key)
        nxt = scheme.start if current is None else current + scheme.step
        if nxt > scheme.max_running:
            raise OverflowError(
                f"Gruppe {main}.{sub} hat das Maximum {scheme.max_running:0{scheme.running_width}d} "
                f"in Zeile {row} überschritten. running_width erhöhen."
            )
        counters[key] = nxt
        ws.cell(row=row, column=article_col).value = scheme.format(main, sub, nxt)
        assigned += 1

    wb.save(output_file)
    ranges = {f"{k[0]}.{k[1]}": v for k, v in sorted(counters.items())}
    return assigned, ranges


def main():
    INPUT = "input.xlsx"
    OUTPUT = "output_mit_artikelnummern.xlsx"
    if not Path(INPUT).exists():
        sys.exit(f"Eingabedatei nicht gefunden: {INPUT}")
    try:
        assigned, ranges = assign_article_numbers(INPUT, OUTPUT)
    except FileNotFoundError as e:
        sys.exit(str(e))
    except PermissionError:
        sys.exit(f"Konnte {OUTPUT} nicht speichern - ist die Datei in Excel geöffnet?")
    except (ValueError, OverflowError) as e:
        sys.exit(f"Abbruch: {e}")
    print(f"Fertig: {OUTPUT}  ({assigned} neue Nummern vergeben)")
    for grp, high in ranges.items():
        print(f"  {grp}: bis {high:04d}")


if __name__ == "__main__":
    main()
