"""
Legacy-Kategorienamen in einer Excel-Masterliste auf neue Bezeichnungen umstellen.

Die Spalten „Hauptgruppe“ und „Untergruppe“ enthalten oft noch alte Namen aus dem
System „Original Kategorie“ / „Original Untergruppe“. Dieses Skript mappt sie auf
die neuen Bezeichnungen (Gruppenschlüssel-kompatibel) und speichert die Datei.

Artikelnummern werden hier nicht vergeben — dafür scripts/processing/artikelnummern.py verwenden.
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Ersetzungstabellen (authoritativ — Werte nicht ändern)
# ---------------------------------------------------------------------------

HAUPTGRUPPE_MAP: dict[str, str] = {
    "ZUBEHÖR": "Zubehör",
    "FLIESENPROFILE": "Profile",
    "DEHNFUGENPROFILE": "Dehnfugenprofile",
    "TREPPENSTUFENPROFILE": "Treppenprofile",
    "BALKON - TERRASSE": "Balkon und Terrasse",
    "MATTENSYSTEME": "Mattensysteme",
    "DUSCHSYSTEME": "Duschsysteme",
    "BODENBELAGSPROFILE": "Bodenbelagsprofile",
    "SOCKELLEISTEN": "Sockelleisten",
    "LED": "LED-Systeme",
    "TAKTIL": "Taktile Leitsysteme",
    "EINGANGSMATTEN": "Eingangsmatten",
    "TRITTSCHALLDÄMMUNG": "Trittschalldämmung",
}

UNTERGRUPPE_MAP: dict[str, str] = {
    "NIVELLIERSYSTEM": "Nivelliersystem",
    "VERLEGEMATERIAL": "Verlegematerial",
    "WERKZEUG": "Werkzeug",
    "ZUBEHÖR": "Zubehör allgemein",
    "FP ECKSCHUTZ": "Eckschutzprofile",
    "FP WAND, BODEN": "Wand- und Bodenprofile",
    "FP BODEN, BODEN": "Bodenprofile",
    "FP DEKORATION": "Dekorprofile",
    "FP ARBEITSPLATTEN": "Arbeitsplattenprofile",
    "DÜNNBETT": "Dünnbettprofile",
    "DICKBETT": "Dickbettprofile",
    "GEBÄUDETRENNFUGEN": "Gebäudetrennfugen",
    "ABDECKPROFILE": "Abdeckprofile",
    "FLIESEN": "Treppenprofile für Fliesen",
    "FÜR EINLAGEN": "Treppenprofile für Einlagen",
    "NACHTRÄGL. EINBAU": "Treppenprofile für nachträglichen Einbau",
    "TREPPENSTUFEN UA": "Treppenstufenprofile universell",
    "BALKONWINKELPROFILE": "Balkonwinkelprofile",
    "ENTWÄSSERUNG B-T": "Entwässerung Balkon und Terrasse",
    "DRAINAGE B-T": "Drainage Balkon und Terrasse",
    "VERLEGEM. BALKON-TER": "Verlegematerial Balkon und Terrasse",
    "VERLEGEM. BALKON-TERRASSE": "Verlegematerial Balkon und Terrasse",
    "ABDICHTMATTEN": "Abdichtmatten",
    "HEIZMATTEN": "Heizmatten",
    "ENTKOPPLUNGSMATTEN": "Entkopplungsmatten",
    "DRAINAGEMATTEN": "Drainagematten",
    "DÄMMMATTEN": "Dämmmatten",
    "ARMIERUNG": "Armierung",
    "LINIENENTWÄSSERUNG": "Linienentwässerung",
    "PUNKTENTWÄSSERUNG": "Punktentwässerung",
    "DUSCHABLAGEN": "Duschablagen",
    "ZUBEHÖR DUSCHSYSTEME": "Zubehör Duschsysteme",
    "BBP ABSCHLUSS": "Abschlussprofile Bodenbeläge",
    "BBP ÜBERGANG": "Übergangsprofile Bodenbeläge",
    "BBP ANPASSUNG": "Anpassungsprofile Bodenbeläge",
    "SL EINTEILIG": "Einteilige Sockelleisten",
    "PROFILE": "LED-Profile",
    "ZUBEHÖR LED": "LED-Zubehör",
    "NOPPEN EINZELN": "Einzelnoppen",
    "NOPPEN FOLIEN": "Noppenfolien",
    "NOPPEN MATTEN": "Noppenmatten",
    "RIPPEN EINZELN": "Einzelrippen",
    "FOLIEN": "Rippenfolien",
    "MARKIERUNGSBÄNDER": "Markierungsbänder",
    "ZUBEHÖR TAKTIL": "Zubehör taktile Systeme",
    "MATTE INDIVID.GR.": "Eingangsmatten individuelle Grösse",
    "RAHMEN EINZELN": "Eingangsmatten Rahmen",
    "MATTE STANDARD GR.": "Eingangsmatten Standardgrösse",
    "PE AS": "Pe As",
}

# „FP ABSCHLUSS“ ist mehrdeutig — Auflösung über Produktfamilie:
FP_ABSCHLUSS_BY_FAMILY: dict[str, str] = {
    "DURONDELL": "Abschlussprofile rund",
    "DUROSOL": "Abschlussprofile Winkel",
    "SQUARELINE": "Abschlussprofile quadratisch",
    "DURAPLUS DIAMOND": "Abschlussprofile Diamond",
    "DURAPLUS": "Abschlussprofile Spezial",
    "DUROSOL 5MIL": "Abschlussprofile 5mil",
}

_WHITESPACE = re.compile(r"\s+")


def normalize_legacy_key(value) -> str:
    """Strip, interne Leerzeichen auf eins reduzieren, case-insensitiv vergleichen."""
    if value is None:
        return ""
    return _WHITESPACE.sub(" ", str(value).strip()).casefold()


def _build_normalized_map(source: dict[str, str]) -> dict[str, str]:
    return {normalize_legacy_key(k): v for k, v in source.items()}


_HAUPTGRUPPE_LOOKUP = _build_normalized_map(HAUPTGRUPPE_MAP)
_UNTERGRUPPE_LOOKUP = _build_normalized_map(UNTERGRUPPE_MAP)
_FP_ABSCHLUSS_LOOKUP = _build_normalized_map(FP_ABSCHLUSS_BY_FAMILY)

_NEW_MAIN_KEYS = {normalize_legacy_key(v) for v in HAUPTGRUPPE_MAP.values()}
_NEW_SUB_KEYS = {normalize_legacy_key(v) for v in UNTERGRUPPE_MAP.values()} | {
    normalize_legacy_key(v) for v in FP_ABSCHLUSS_BY_FAMILY.values()
}

_FP_ABSCHLUSS_KEY = normalize_legacy_key("FP ABSCHLUSS")


@dataclass(frozen=True)
class LegacyScheme:
    header_row: int = 1
    first_data_row: int = 2
    main_group_header: str = "Hauptgruppe"
    sub_group_header: str = "Untergruppe"
    product_family_header: str = "Produktfamilie"
    data_row_key_header: str = "Artikelnr."


@dataclass
class LegacyNormalizationError:
    row: int
    main_value: str | None = None
    sub_value: str | None = None
    main_unresolved: bool = False
    sub_unresolved: bool = False
    note: str | None = None


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def find_column(ws, header_name: str, header_row: int) -> int:
    target = normalize_header(header_name)
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=header_row, column=col).value) == target:
            return col
    present = [
        str(ws.cell(row=header_row, column=col).value).strip()
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value not in (None, "")
    ]
    raise ValueError(
        f"Spalte {header_name!r} nicht gefunden. "
        f"Vorhandene Spaltenüberschriften: {', '.join(present) or '(keine)'}"
    )


def is_data_row(ws, row: int, data_key_col: int) -> bool:
    val = ws.cell(row=row, column=data_key_col).value
    return val is not None and str(val).strip() != ""


def _display_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _resolve_main(raw) -> tuple[str | None, LegacyNormalizationError | None]:
    display = _display_value(raw)
    if display is None:
        return None, LegacyNormalizationError(row=0, main_unresolved=True, main_value=None)

    key = normalize_legacy_key(raw)
    if key in _HAUPTGRUPPE_LOOKUP:
        return _HAUPTGRUPPE_LOOKUP[key], None
    if key in _NEW_MAIN_KEYS:
        return display, None
    return None, LegacyNormalizationError(row=0, main_unresolved=True, main_value=display)


def _resolve_sub(raw, family_raw) -> tuple[str | None, LegacyNormalizationError | None]:
    display = _display_value(raw)
    if display is None:
        return None, LegacyNormalizationError(row=0, sub_unresolved=True, sub_value=None)

    key = normalize_legacy_key(raw)
    if key == _FP_ABSCHLUSS_KEY:
        family_display = _display_value(family_raw)
        family_key = normalize_legacy_key(family_raw)
        if family_key in _FP_ABSCHLUSS_LOOKUP:
            return _FP_ABSCHLUSS_LOOKUP[family_key], None
        return None, LegacyNormalizationError(
            row=0,
            sub_unresolved=True,
            sub_value=display,
            note=(
                "FP ABSCHLUSS erfordert eine bekannte Produktfamilie "
                f"(aktuell: {family_display!r})"
            ),
        )

    if key in _UNTERGRUPPE_LOOKUP:
        return _UNTERGRUPPE_LOOKUP[key], None
    if key in _NEW_SUB_KEYS:
        return display, None
    return None, LegacyNormalizationError(row=0, sub_unresolved=True, sub_value=display)


def legacy_to_new_names(
    ws,
    row: int,
    scheme: LegacyScheme,
    *,
    main_col: int,
    sub_col: int,
    family_col: int,
) -> tuple[str | None, str | None, LegacyNormalizationError | None]:
    """Alte Haupt-/Untergruppen-Namen einer Zeile in neue Bezeichnungen übersetzen."""
    main_raw = ws.cell(row=row, column=main_col).value
    sub_raw = ws.cell(row=row, column=sub_col).value
    family_raw = ws.cell(row=row, column=family_col).value

    main_new, main_err = _resolve_main(main_raw)
    sub_new, sub_err = _resolve_sub(sub_raw, family_raw)

    if main_err is not None and sub_err is not None:
        return None, None, LegacyNormalizationError(
            row=row,
            main_unresolved=True,
            main_value=main_err.main_value,
            sub_unresolved=True,
            sub_value=sub_err.sub_value,
            note=sub_err.note,
        )
    if main_err is not None:
        return None, None, LegacyNormalizationError(
            row=row,
            main_unresolved=True,
            main_value=main_err.main_value,
        )
    if sub_err is not None:
        return None, None, LegacyNormalizationError(
            row=row,
            sub_unresolved=True,
            sub_value=sub_err.sub_value,
            note=sub_err.note,
        )
    return main_new, sub_new, None


def format_legacy_errors(errors: list[LegacyNormalizationError]) -> str:
    lines = ["Legacy-Gruppennamen konnten nicht aufgelöst werden:"]

    unresolved_mains = sorted({e.main_value for e in errors if e.main_unresolved and e.main_value})
    unresolved_subs = sorted({e.sub_value for e in errors if e.sub_unresolved and e.sub_value})

    if unresolved_mains:
        lines.append(f"\nUnbekannte Hauptgruppen ({len(unresolved_mains)}):")
        lines.extend(f"  - {name}" for name in unresolved_mains)
    if unresolved_subs:
        lines.append(f"\nUnbekannte Untergruppen ({len(unresolved_subs)}):")
        lines.extend(f"  - {name}" for name in unresolved_subs)

    lines.append(f"\nDetails ({len(errors)} Zeile(n)):")
    for err in errors:
        parts = [f"  Zeile {err.row}:"]
        if err.main_unresolved:
            parts.append(f" unbekannte Hauptgruppe {err.main_value!r}")
        if err.sub_unresolved:
            parts.append(f" unbekannte Untergruppe {err.sub_value!r}")
        if err.note:
            parts.append(f" ({err.note})")
        lines.append("".join(parts))

    lines.append(
        "\nBitte korrigieren Sie die Werte in input.xlsx oder ergänzen Sie die Ersetzungstabellen."
    )
    return "\n".join(lines)


def replace_legacy_names(
    input_file: str,
    output_file: str,
    scheme: LegacyScheme = LegacyScheme(),
    *,
    sheet_name: str | None = None,
    strict: bool = True,
) -> tuple[int, int]:
    """
    Legacy-Namen ersetzen und speichern.

    Returns (resolved_row_count, changed_cell_count).
    """
    wb = load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    main_col = find_column(ws, scheme.main_group_header, scheme.header_row)
    sub_col = find_column(ws, scheme.sub_group_header, scheme.header_row)
    family_col = find_column(ws, scheme.product_family_header, scheme.header_row)
    data_key_col = find_column(ws, scheme.data_row_key_header, scheme.header_row)

    errors: list[LegacyNormalizationError] = []
    resolved: dict[int, tuple[str, str]] = {}

    for row in range(scheme.first_data_row, ws.max_row + 1):
        if not is_data_row(ws, row, data_key_col):
            continue
        main_new, sub_new, err = legacy_to_new_names(
            ws, row, scheme, main_col=main_col, sub_col=sub_col, family_col=family_col
        )
        if err is not None:
            errors.append(err)
        else:
            resolved[row] = (main_new, sub_new)  # type: ignore[arg-type]

    if errors:
        print(format_legacy_errors(errors))
        if strict:
            raise ValueError(
                f"Abbruch: {len(errors)} Zeile(n) mit nicht auflösbaren Legacy-Namen "
                "(strict=True, Datei wurde nicht gespeichert)."
            )

    changed = 0
    for row, (main_new, sub_new) in resolved.items():
        old_main = ws.cell(row=row, column=main_col).value
        old_sub = ws.cell(row=row, column=sub_col).value
        if old_main != main_new:
            ws.cell(row=row, column=main_col).value = main_new
            changed += 1
        if old_sub != sub_new:
            ws.cell(row=row, column=sub_col).value = sub_new
            changed += 1

    wb.save(output_file)

    if errors and not strict:
        print(
            f"\nWarnung: {len(errors)} Zeile(n) unaufgelöst — "
            f"{len(resolved)} Zeile(n) wurden ersetzt und gespeichert."
        )

    return len(resolved), changed


def _build_verify_workbook(path: Path) -> None:
    """Kleine Testdatei für die Verifikation."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    headers = ["Artikelnr.", "Hauptgruppe", "Untergruppe", "Produktfamilie"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = header

    rows = [
        # Normal legacy pair
        ("T-001", "ZUBEHÖR", "NIVELLIERSYSTEM", None),
        # FP ABSCHLUSS + known family
        ("T-002", "FLIESENPROFILE", "FP ABSCHLUSS", "DUROSOL"),
        # FP ABSCHLUSS + unknown family
        ("T-003", "FLIESENPROFILE", "FP ABSCHLUSS", "UNKNOWN_FAMILY"),
        # Already new names (idempotent)
        ("T-004", "Zubehör", "Nivelliersystem", "NIVOFIX"),
        # Bogus legacy value
        ("T-005", "6", "NIVELLIERSYSTEM", None),
    ]
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(path)


def _print_sheet_groups(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    main_col = find_column(ws, "Hauptgruppe", 1)
    sub_col = find_column(ws, "Untergruppe", 1)
    key_col = find_column(ws, "Artikelnr.", 1)
    print(f"  {'Artikelnr.':<10} {'Hauptgruppe':<25} {'Untergruppe'}")
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=key_col).value
        if key is None:
            continue
        main = ws.cell(row=row, column=main_col).value
        sub = ws.cell(row=row, column=sub_col).value
        print(f"  {str(key):<10} {str(main):<25} {sub}")
    wb.close()


def _run_verify() -> None:
    base = Path(__file__).resolve().parent.parent / "test_legacy_names"
    base.mkdir(exist_ok=True)
    input_path = base / "input.xlsx"
    output_strict = base / "output_strict.xlsx"
    output_loose = base / "output_loose.xlsx"

    _build_verify_workbook(input_path)
    print("=== Testdatei erstellt ===")
    print(f"{input_path}\n")
    print("Eingabe:")
    _print_sheet_groups(input_path)

    print("\n=== strict=True (erwartet: Abbruch, keine Datei) ===")
    try:
        replace_legacy_names(str(input_path), str(output_strict), strict=True)
        print("FEHLER: strict=True hätte abbrechen müssen.")
    except ValueError as e:
        print(f"OK — Abbruch: {e}")
        if output_strict.exists():
            print("FEHLER: output_strict.xlsx wurde trotzdem geschrieben.")
        else:
            print("OK — keine Ausgabedatei gespeichert.")

    print("\n=== strict=False (erwartet: Teilersetzung + Warnung) ===")
    resolved, changed = replace_legacy_names(str(input_path), str(output_loose), strict=False)
    print(f"Ergebnis: {resolved} Zeile(n) aufgelöst, {changed} Zelle(n) geändert.")
    print("\nAusgabe:")
    _print_sheet_groups(output_loose)


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    from scripts.paths import ensure_parent_dir, resolve_path

    sheet = params["sheet_name"].strip()
    sheet_name = sheet or None
    input_path = resolve_path(params["input"])
    output_path = ensure_parent_dir(params["output"])
    try:
        resolved, changed = replace_legacy_names(
            str(input_path),
            str(output_path),
            strict=params["strict"],
            sheet_name=sheet_name,
        )
    except PermissionError as e:
        raise PermissionError(
            f"Konnte {params['output']} nicht speichern — ist die Datei in Excel geöffnet?"
        ) from e

    return RunResult(
        summary=(
            f"Fertig: {params['output']}  ({resolved} Zeile(n), {changed} Zelle(n) geändert)"
        ),
        details=[],
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="legacy_names",
        title="Legacy-Namen ersetzen",
        description="Alte Haupt- und Untergruppen-Namen auf neue Bezeichnungen umstellen.",
        fields=(
            FieldSpec("input", "Eingabedatei", FieldKind.FILE_IN, "input/input.xlsx"),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "output/processing/output_mit_neuen_namen.xlsx",
                output_name="output_mit_neuen_namen.xlsx",
            ),
            FieldSpec(
                "strict",
                "Bei Fehlern abbrechen (strict)",
                FieldKind.BOOL,
                True,
                help="Abbruch bei nicht auflösbaren Legacy-Namen",
            ),
            FieldSpec(
                "sheet_name",
                "Tabellenblatt",
                FieldKind.STR,
                "",
                help="Leer lassen für aktives Blatt",
                advanced=True,
            ),
        ),
        run=run_job,
    )


def main():
    if len(sys.argv) > 1 and sys.argv[1] == "--verify":
        _run_verify()
        return

    _ensure_project_root()
    from gui.job_spec import args_to_params, build_argparser, coerce_params, validate_params

    parser = build_argparser(JOB_SPEC)
    args = parser.parse_args()
    params = coerce_params(JOB_SPEC, args_to_params(JOB_SPEC, args))
    try:
        validate_params(JOB_SPEC, params)
        result = run_job(params)
    except FileNotFoundError as e:
        sys.exit(str(e))
    except PermissionError as e:
        sys.exit(str(e))
    except ValueError as e:
        sys.exit(f"Abbruch: {e}")
    print(result.summary)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
