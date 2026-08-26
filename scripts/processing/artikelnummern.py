"""
Artikelnummer-Generator (PROSEMA / DURAL Masterliste) — Excel-Adapter.

Reine Vergabelogik: ``core.numbering``. Gruppenschlüssel: ``core.groups``.
Dieses Modul liest/schreibt die Masterliste und bleibt der CLI / JOB_SPEC-Einstieg.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


_ensure_project_root()

from openpyxl import load_workbook
from openpyxl.styles import Font

from core.groups import load_group_dictionary
from core.numbering import (
    AssignmentResult,
    RowInput,
    Scheme,
    assign_numbers,
    format_resolution_errors,
)

FIXED_MASTER_COLUMNS = {
    "Währung": "EUR",
    "Verkaufsartikel-Währung": "EUR",
    "Vertriebsweg": "GROSS1",
}


@dataclass(frozen=True)
class ExcelLayout:
    """Spalten- und Zeilenlayout der Masterliste (kein Nummerierungsschema)."""

    header_row: int = 1
    first_data_row: int = 2
    article_col_header: str = "Prosema Artikelnummer"
    main_group_header: str = "Hauptgruppe"
    sub_group_header: str = "Untergruppe"
    data_row_key_header: str = "Artikelnr."
    dictionary_file: str = "data/gruppen.xlsx"


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def find_column(ws, header_name: str, header_row: int) -> int:
    target = normalize_header(header_name)
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=header_row, column=col).value) == target:
            return col
    present = [
        str(ws.cell(row=header_row, column=col).value).strip()
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=col).value not in (None, "")
    ]
    raise ValueError(
        f"Spalte {header_name!r} nicht gefunden. "
        f"Vorhandene Spaltenüberschriften: {', '.join(present) or '(keine)'}"
    )


def find_or_create_column(ws, header_name: str, header_row: int) -> int:
    target = normalize_header(header_name)
    for col in range(1, ws.max_column + 1):
        if normalize_header(ws.cell(row=header_row, column=col).value) == target:
            return col
    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = header_name
    ws.cell(row=header_row, column=new_col).font = Font(bold=True)
    return new_col


def is_data_row(ws, row: int, data_key_col: int) -> bool:
    val = ws.cell(row=row, column=data_key_col).value
    return val is not None and str(val).strip() != ""


def _read_row_inputs(ws, layout: ExcelLayout, col: dict[str, int]) -> list[RowInput]:
    rows: list[RowInput] = []
    for row in range(layout.first_data_row, ws.max_row + 1):
        rows.append(
            RowInput(
                row=row,
                main_name=ws.cell(row=row, column=col["main"]).value,
                sub_name=ws.cell(row=row, column=col["sub"]).value,
                existing_article_number=ws.cell(row=row, column=col["article"]).value,
                is_data_row=is_data_row(ws, row, col["data_key"]),
            )
        )
    return rows


def assign_article_numbers(
    input_file: str,
    output_file: str,
    scheme: Scheme = Scheme(),
    layout: ExcelLayout = ExcelLayout(),
    *,
    dictionary_path: Path | None = None,
    sheet_name: str | None = None,
    overwrite_existing: bool = False,
    strict: bool = True,
) -> tuple[int, dict[str, int]]:
    """Returns (assigned_count, {group: high_water}). Does not mutate the input file."""
    from scripts.paths import resolve_path

    dict_path = dictionary_path if dictionary_path is not None else resolve_path(layout.dictionary_file)
    groups = load_group_dictionary(dict_path)

    wb = load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    cols = {
        "main": find_column(ws, layout.main_group_header, layout.header_row),
        "sub": find_column(ws, layout.sub_group_header, layout.header_row),
        "data_key": find_column(ws, layout.data_row_key_header, layout.header_row),
        "article": find_or_create_column(ws, layout.article_col_header, layout.header_row),
    }
    fixed_cols = {
        name: find_or_create_column(ws, name, layout.header_row)
        for name in FIXED_MASTER_COLUMNS
    }

    row_inputs = _read_row_inputs(ws, layout, cols)
    result: AssignmentResult = assign_numbers(
        row_inputs,
        groups,
        scheme,
        overwrite_existing=overwrite_existing,
        strict=strict,
    )

    if result.errors and not strict:
        print(format_resolution_errors(result.errors))

    for row_in in row_inputs:
        if not row_in.is_data_row:
            continue
        for header, fixed_value in FIXED_MASTER_COLUMNS.items():
            ws.cell(row=row_in.row, column=fixed_cols[header]).value = fixed_value
        if row_in.row in result.assigned:
            ws.cell(row=row_in.row, column=cols["article"]).value = result.assigned[row_in.row]

    wb.save(output_file)
    return result.assigned_count, result.high_water


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    scheme = Scheme(
        start=params["start"],
        step=params["step"],
    )
    layout = ExcelLayout(dictionary_file=params["dictionary_file"])
    from scripts.paths import ensure_parent_dir, resolve_path

    input_path = resolve_path(params["input"])
    output_path = ensure_parent_dir(params["output"])
    try:
        assigned, ranges = assign_article_numbers(
            str(input_path),
            str(output_path),
            scheme=scheme,
            layout=layout,
            dictionary_path=resolve_path(params["dictionary_file"]),
            overwrite_existing=params["overwrite_existing"],
            strict=params["strict"],
        )
    except PermissionError as e:
        raise PermissionError(
            f"Konnte {params['output']} nicht speichern — ist die Datei in Excel geöffnet?"
        ) from e

    details = [f"  {grp}: bis {high:04d}" for grp, high in ranges.items()]
    return RunResult(
        summary=f"Fertig: {params['output']}  ({assigned} neue Nummern vergeben)",
        details=details,
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="artikelnummern",
        title="Artikelnummern erstellen",
        description="Fehlende Artikelnummern in einer Excel-Masterliste vergeben.",
        fields=(
            FieldSpec("input", "Eingabedatei", FieldKind.FILE_IN, "input/input.xlsx"),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "output/processing/output_mit_artikelnummern.xlsx",
                output_name="output_mit_artikelnummern.xlsx",
            ),
            FieldSpec(
                "strict",
                "Bei Fehlern abbrechen (strikt)",
                FieldKind.BOOL,
                True,
                help="Abbruch bei unbekannten Gruppennamen",
            ),
            FieldSpec(
                "overwrite_existing",
                "Bestehende Nummern überschreiben",
                FieldKind.BOOL,
                False,
                help="Vorhandene Artikelnummern neu vergeben",
            ),
            FieldSpec("start", "Startnummer", FieldKind.INT, 10, advanced=True),
            FieldSpec("step", "Schrittweite", FieldKind.INT, 10, advanced=True),
            FieldSpec(
                "dictionary_file",
                "Gruppenschlüssel",
                FieldKind.STR,
                "data/gruppen.xlsx",
                advanced=True,
            ),
        ),
        run=run_job,
    )


def main():
    _ensure_project_root()
    from gui.job_spec import args_to_params, build_argparser, coerce_params, validate_params

    parser = build_argparser(JOB_SPEC)
    args = parser.parse_args()
    params = coerce_params(JOB_SPEC, args_to_params(JOB_SPEC, args))
    try:
        validate_params(JOB_SPEC, params)
        result = run_job(params)
    except FileNotFoundError as e:
        sys.exit(str(e))
    except PermissionError as e:
        sys.exit(str(e))
    except (ValueError, OverflowError) as e:
        sys.exit(f"Abbruch: {e}")
    print(result.summary)
    for line in result.details:
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
