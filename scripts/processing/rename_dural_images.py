"""
Dural-Bilder nach Prosema-Artikelnummern umbenennen und in separate Ordner kopieren.

Liest die Zuordnung Dural Artikelnr. -> Prosema Artikelnummer aus input/input.xlsx.
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from scripts.paths import PROJECT_ROOT, resolve_path

DEFAULT_BASE = Path(
    "/Users/chris-mbp/Library/CloudStorage/Dropbox/PPROSEMA/Dural/Bilder Preisliste"
)
SOURCE_DIRS = {
    "Bilder PL farblich": "Bilder PL farblich umbenannt",
    "Bilder PL Strichzeichnung": "Bilder PL Strichzeichnung umbenannt",
}
DURAL_COLUMN = "Artikelnr."
PROSEMA_COLUMN = "Prosema Artikelnummer"
FILENAME_SUFFIX_RE = re.compile(r"^(.+)-(\d+)$")


@dataclass
class RenameStats:
    copied: int = 0
    unmatched_files: int = 0
    unmatched_article_numbers: set[str] = field(default_factory=set)
    errors: int = 0


class RunLog:
    """Append-only log written incrementally while the script runs."""

    def __init__(self, log_path: Path) -> None:
        self.path = log_path
        self._file = log_path.open("w", encoding="utf-8")

    def close(self) -> None:
        self._file.close()

    def _write(self, line: str = "") -> None:
        self._file.write(line + "\n")
        self._file.flush()

    def write_header(
        self,
        *,
        master_path: Path,
        base_dir: Path,
        mapping_count: int,
    ) -> None:
        self._write("Dural -> Prosema Bildumbenennung")
        self._write(f"Gestartet: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._write(f"Masterdatei: {master_path}")
        self._write(f"Basisordner: {base_dir}")
        self._write(f"Zuordnungen geladen: {mapping_count}")
        self._write("")

    def write_section(self, label: str) -> None:
        self._write(f"[{label}]")
        self._write(f"  Gestartet: {datetime.now().strftime('%H:%M:%S')}")

    def write_unmatched(self, filename: str, article_number: str) -> None:
        self._write(
            f"  Nicht zugeordnet: {filename} (Artikelnr. {article_number})"
        )

    def write_error(self, message: str) -> None:
        self._write(f"  FEHLER: {message}")

    def write_section_done(self, stats: RenameStats) -> None:
        self._write(
            f"  Abgeschlossen: {stats.copied} kopiert, "
            f"{stats.unmatched_files} nicht zugeordnet, "
            f"{stats.errors} Fehler"
        )
        self._write("")

    def write_footer(
        self,
        *,
        all_stats: dict[str, RenameStats],
        excel_without_image: list[str],
    ) -> None:
        copied = sum(s.copied for s in all_stats.values())
        unmatched_files = sum(s.unmatched_files for s in all_stats.values())
        unmatched_articles = {
            article for s in all_stats.values() for article in s.unmatched_article_numbers
        }
        errors = sum(s.errors for s in all_stats.values())

        self._write("Zusammenfassung")
        self._write(f"  Abgeschlossen: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self._write(f"  Kopiert/umbenannt: {copied}")
        self._write(f"  Nicht zuordenbar (Artikelnummern): {len(unmatched_articles)}")
        self._write(f"  Nicht zuordenbar (Dateien): {unmatched_files}")
        self._write(f"  Master-Artikel ohne Bild: {len(excel_without_image)}")
        self._write(f"  Fehler: {errors}")
        self._write("")

        if excel_without_image:
            self._write("Artikelnummern in Masterdatei ohne passendes Bild")
            for article in excel_without_image:
                self._write(f"  {article}")
            self._write("")


def _article_lookup_keys(value) -> set[str]:
    if value is None:
        return set()
    raw = str(value).strip()
    if not raw:
        return set()
    keys = {raw.upper()}
    if raw.isdigit():
        stripped = raw.lstrip("0") or "0"
        keys.add(stripped.upper())
    return keys


def load_mapping(master_path: Path) -> tuple[dict[str, str], set[str]]:
    wb = load_workbook(master_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        dural_idx = headers.index(DURAL_COLUMN)
        prosema_idx = headers.index(PROSEMA_COLUMN)
    except ValueError as exc:
        wb.close()
        raise ValueError(
            f"Erwartete Spalten {DURAL_COLUMN!r} und {PROSEMA_COLUMN!r} "
            f"in {master_path} nicht gefunden."
        ) from exc

    mapping: dict[str, str] = {}
    excel_articles: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        dural = row[dural_idx]
        prosema = row[prosema_idx]
        if dural is None or prosema is None:
            continue
        prosema_str = str(prosema).strip()
        if not prosema_str:
            continue
        dural_str = str(dural).strip()
        excel_articles.add(dural_str)
        for key in _article_lookup_keys(dural):
            mapping[key] = prosema_str
    wb.close()
    return mapping, excel_articles


def parse_filename(filename: str) -> tuple[str, str, str]:
    stem, ext = Path(filename).stem, Path(filename).suffix
    match = FILENAME_SUFFIX_RE.match(stem)
    if match:
        return match.group(1), f"-{match.group(2)}", ext
    return stem, "", ext


def resolve_prosema(article_number: str, mapping: dict[str, str]) -> str | None:
    for key in _article_lookup_keys(article_number):
        if key in mapping:
            return mapping[key]
    return None


def process_directory(
    source_dir: Path,
    target_dir: Path,
    mapping: dict[str, str],
    *,
    label: str,
    log: RunLog | None,
    dry_run: bool,
) -> RenameStats:
    stats = RenameStats()
    if not source_dir.is_dir():
        message = f"Quellordner nicht gefunden: {source_dir}"
        stats.errors += 1
        if log:
            log.write_error(message)
        return stats

    if log:
        log.write_section(label)

    files = sorted(
        f.name for f in source_dir.iterdir() if f.is_file() and not f.name.startswith(".")
    )
    grouped: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for filename in files:
        article, suffix, ext = parse_filename(filename)
        grouped[article.upper()].append((filename, suffix, ext))

    if not dry_run:
        target_dir.mkdir(parents=True, exist_ok=True)

    total = len(files)
    processed = 0
    for _article_key, entries in sorted(grouped.items(), key=lambda item: item[0]):
        article_number, _, _ = parse_filename(entries[0][0])
        prosema = resolve_prosema(article_number, mapping)
        if prosema is None:
            stats.unmatched_article_numbers.add(article_number)
            for filename, _, _ in entries:
                stats.unmatched_files += 1
                processed += 1
                if log:
                    log.write_unmatched(filename, article_number)
            if processed % 100 == 0 or processed == total:
                print(
                    f"  [{label}] {processed}/{total} Dateien verarbeitet …",
                    file=sys.stderr,
                )
            continue

        multiple = len(entries) > 1
        for index, (filename, suffix, ext) in enumerate(entries, start=1):
            if multiple:
                target_name = f"{prosema}_{index}{suffix}{ext}"
            else:
                target_name = f"{prosema}{suffix}{ext}"

            source_path = source_dir / filename
            target_path = target_dir / target_name
            if dry_run:
                stats.copied += 1
                processed += 1
                continue

            try:
                shutil.copy2(source_path, target_path)
                stats.copied += 1
                processed += 1
            except OSError as exc:
                stats.errors += 1
                processed += 1
                if log:
                    log.write_error(f"{filename}: {exc}")

            if processed % 100 == 0 or processed == total:
                print(
                    f"  [{label}] {processed}/{total} Dateien verarbeitet …",
                    file=sys.stderr,
                )

    if log:
        log.write_section_done(stats)
    return stats


def run(
    *,
    master_path: Path,
    base_dir: Path,
    dry_run: bool = False,
) -> Path:
    mapping, excel_articles = load_mapping(master_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = base_dir / f"umbenennung_log_{timestamp}.txt"

    log: RunLog | None = None
    if not dry_run:
        log = RunLog(log_path)
        log.write_header(
            master_path=master_path,
            base_dir=base_dir,
            mapping_count=len(mapping),
        )
        print(f"Log (wird fortlaufend geschrieben): {log_path}", file=sys.stderr)

    all_stats: dict[str, RenameStats] = {}
    found_articles: set[str] = set()

    for source_name, target_name in SOURCE_DIRS.items():
        source_dir = base_dir / source_name
        target_dir = base_dir / target_name
        stats = process_directory(
            source_dir,
            target_dir,
            mapping,
            label=source_name,
            log=log,
            dry_run=dry_run,
        )
        all_stats[source_name] = stats

        if source_dir.is_dir():
            for filename in source_dir.iterdir():
                if not filename.is_file() or filename.name.startswith("."):
                    continue
                article, _, _ = parse_filename(filename.name)
                if resolve_prosema(article, mapping) is not None:
                    found_articles.update(_article_lookup_keys(article))

    excel_without_image = sorted(
        article
        for article in excel_articles
        if not _article_lookup_keys(article) & found_articles
    )

    if log:
        log.write_footer(
            all_stats=all_stats,
            excel_without_image=excel_without_image,
        )
        log.close()

    return log_path


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Dural-Bilder anhand der Masterliste nach Prosema-Artikelnummern umbenennen."
    )
    parser.add_argument(
        "--master",
        type=Path,
        default=PROJECT_ROOT / "input" / "input.xlsx",
        help="Pfad zur Master-Excel (Standard: input/input.xlsx)",
    )
    parser.add_argument(
        "--base-dir",
        type=Path,
        default=DEFAULT_BASE,
        help="Basisordner mit Quell- und Ziel-Unterordnern",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Nur simulieren, keine Dateien kopieren und kein Log schreiben",
    )
    args = parser.parse_args(argv)

    master_path = resolve_path(args.master)
    base_dir = args.base_dir.expanduser().resolve()

    if not master_path.is_file():
        print(f"Masterdatei nicht gefunden: {master_path}", file=sys.stderr)
        return 1
    if not base_dir.is_dir():
        print(f"Basisordner nicht gefunden: {base_dir}", file=sys.stderr)
        return 1

    log_path = run(master_path=master_path, base_dir=base_dir, dry_run=args.dry_run)
    if args.dry_run:
        print("Dry-Run abgeschlossen (keine Dateien kopiert, kein Log geschrieben).")
    else:
        print("Fertig.")
        print("Zielordner:")
        for target_name in SOURCE_DIRS.values():
            print(f"  {base_dir / target_name}")
        print(f"Log: {log_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
