"""Seed the group registry from data/gruppen.xlsx.

Default is a dry run. Writing requires --commit.

The workbook has sheets "Hauptgruppen" (Code, Bezeichnung) and "Untergruppen"
(Hauptgruppe, Untergruppe, Bezeichnung). There is no alias sheet; aliases are
managed in the application, not by this script.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from app.db import SessionLocal
from app.groups_service import (
    CODE_RE,
    create_hauptgruppe,
    create_untergruppe,
)
from app.models import Hauptgruppe, Untergruppe
from scripts.paths import DATA_DIR, PROJECT_ROOT

SEED_ACTOR = {"oid": "seed", "name": "Seed-Skript"}
DEFAULT_WORKBOOK = DATA_DIR / "gruppen.xlsx"


@dataclass
class RejectedRow:
    sheet: str
    row: int
    message: str

    def format(self) -> str:
        return f"{self.sheet} Zeile {self.row}: {self.message}"


@dataclass
class SeedPlan:
    hauptgruppen: list[tuple[int, str, str]] = field(default_factory=list)
    untergruppen: list[tuple[int, str, str, str]] = field(default_factory=list)
    rejected: list[RejectedRow] = field(default_factory=list)
    haupt_inserted: int = 0
    haupt_present: int = 0
    unter_inserted: int = 0
    unter_present: int = 0


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text if text else None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _find_column(headers: tuple[Any, ...], header_name: str) -> int:
    target = _normalize_header(header_name)
    for idx, cell in enumerate(headers):
        if _normalize_header(cell) == target:
            return idx
    present = [str(cell).strip() for cell in headers if cell not in (None, "")]
    raise ValueError(
        f"Spalte {header_name!r} nicht gefunden. "
        f"Vorhandene Spaltenüberschriften: {', '.join(present) or '(keine)'}"
    )


def parse_workbook(path: Path) -> SeedPlan:
    plan = SeedPlan()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Hauptgruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Hauptgruppen".')
        if "Untergruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Untergruppen".')

        ws_main = wb["Hauptgruppen"]
        main_rows = list(ws_main.iter_rows(values_only=True))
        if not main_rows:
            raise ValueError('Tabellenblatt "Hauptgruppen" ist leer.')
        main_headers = tuple(main_rows[0])
        code_idx = _find_column(main_headers, "Code")
        name_idx = _find_column(main_headers, "Bezeichnung")

        seen_main: dict[str, int] = {}
        for offset, row in enumerate(main_rows[1:], start=2):
            code_raw = _cell_text(row[code_idx] if code_idx < len(row) else None)
            name_raw = _cell_text(row[name_idx] if name_idx < len(row) else None)
            if code_raw is None and name_raw is None:
                continue
            if code_raw is None or not CODE_RE.fullmatch(code_raw):
                plan.rejected.append(
                    RejectedRow(
                        "Hauptgruppen",
                        offset,
                        f"Code {code_raw!r} muss aus genau drei Ziffern bestehen",
                    )
                )
                continue
            if name_raw is None:
                plan.rejected.append(
                    RejectedRow("Hauptgruppen", offset, "Bezeichnung fehlt")
                )
                continue
            if code_raw in seen_main:
                plan.rejected.append(
                    RejectedRow(
                        "Hauptgruppen",
                        offset,
                        f"Code {code_raw!r} ist doppelt (bereits Zeile {seen_main[code_raw]})",
                    )
                )
                continue
            seen_main[code_raw] = offset
            plan.hauptgruppen.append((offset, code_raw, name_raw))

        ws_sub = wb["Untergruppen"]
        sub_rows = list(ws_sub.iter_rows(values_only=True))
        if not sub_rows:
            raise ValueError('Tabellenblatt "Untergruppen" ist leer.')
        sub_headers = tuple(sub_rows[0])
        main_code_idx = _find_column(sub_headers, "Hauptgruppe")
        sub_code_idx = _find_column(sub_headers, "Untergruppe")
        sub_name_idx = _find_column(sub_headers, "Bezeichnung")

        seen_sub: dict[tuple[str, str], int] = {}
        for offset, row in enumerate(sub_rows[1:], start=2):
            main_raw = _cell_text(row[main_code_idx] if main_code_idx < len(row) else None)
            sub_raw = _cell_text(row[sub_code_idx] if sub_code_idx < len(row) else None)
            name_raw = _cell_text(row[sub_name_idx] if sub_name_idx < len(row) else None)
            if main_raw is None and sub_raw is None and name_raw is None:
                continue
            if main_raw is None or not CODE_RE.fullmatch(main_raw):
                plan.rejected.append(
                    RejectedRow(
                        "Untergruppen",
                        offset,
                        f"Hauptgruppe {main_raw!r} muss aus genau drei Ziffern bestehen",
                    )
                )
                continue
            if sub_raw is None or not CODE_RE.fullmatch(sub_raw):
                plan.rejected.append(
                    RejectedRow(
                        "Untergruppen",
                        offset,
                        f"Code {sub_raw!r} muss aus genau drei Ziffern bestehen",
                    )
                )
                continue
            if name_raw is None:
                plan.rejected.append(
                    RejectedRow("Untergruppen", offset, "Bezeichnung fehlt")
                )
                continue
            if main_raw not in seen_main:
                plan.rejected.append(
                    RejectedRow(
                        "Untergruppen",
                        offset,
                        f"Hauptgruppe {main_raw!r} fehlt im Blatt Hauptgruppen",
                    )
                )
                continue
            key = (main_raw, sub_raw)
            if key in seen_sub:
                plan.rejected.append(
                    RejectedRow(
                        "Untergruppen",
                        offset,
                        f"Code {main_raw}.{sub_raw} ist doppelt "
                        f"(bereits Zeile {seen_sub[key]})",
                    )
                )
                continue
            seen_sub[key] = offset
            plan.untergruppen.append((offset, main_raw, sub_raw, name_raw))
    finally:
        wb.close()
    return plan


def apply_seed(db: Session, plan: SeedPlan) -> SeedPlan:
    """Insert missing groups. Caller commits. Does not write if plan.rejected."""
    if plan.rejected:
        return plan

    mains = list(db.scalars(select(Hauptgruppe)))
    existing_main = {row.code: row for row in mains}
    id_to_code = {row.id: row.code for row in mains}
    existing_sub: dict[tuple[str, str], Untergruppe] = {}
    for row in db.scalars(select(Untergruppe)):
        parent_code = id_to_code.get(row.hauptgruppe_id)
        if parent_code is not None:
            existing_sub[(parent_code, row.code)] = row

    for _row_num, code, name in plan.hauptgruppen:
        found = existing_main.get(code)
        if found is not None:
            plan.haupt_present += 1
            continue
        created = create_hauptgruppe(db, code=code, name=name, actor=SEED_ACTOR)
        existing_main[code] = created
        id_to_code[created.id] = code
        plan.haupt_inserted += 1

    for row_num, main_code, sub_code, name in plan.untergruppen:
        if (main_code, sub_code) in existing_sub:
            plan.unter_present += 1
            continue
        parent = existing_main.get(main_code)
        if parent is None:
            plan.rejected.append(
                RejectedRow(
                    "Untergruppen",
                    row_num,
                    f"Hauptgruppe {main_code!r} fehlt in der Datenbank",
                )
            )
            continue
        created = create_untergruppe(
            db, parent, code=sub_code, name=name, actor=SEED_ACTOR
        )
        existing_sub[(main_code, sub_code)] = created
        plan.unter_inserted += 1
    return plan


def format_report(plan: SeedPlan, *, committed: bool) -> str:
    lines = [
        f"Hauptgruppen eingefügt: {plan.haupt_inserted}",
        f"Hauptgruppen bereits vorhanden: {plan.haupt_present}",
        f"Untergruppen eingefügt: {plan.unter_inserted}",
        f"Untergruppen bereits vorhanden: {plan.unter_present}",
    ]
    if plan.rejected:
        lines.append("Abgelehnt:")
        lines.extend(f"  {item.format()}" for item in plan.rejected)
    else:
        lines.append("Abgelehnt: (keine)")
    if not committed:
        lines.append("Dry-run: nichts geschrieben. Zum Schreiben --commit übergeben.")
    return "\n".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Gruppenregister aus gruppen.xlsx befüllen.")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="Pfad zur Excel-Datei (Standard: data/gruppen.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Prüfen und Zusammenfassung drucken, ohne zu schreiben (Standard)",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Änderungen in die Datenbank schreiben",
    )
    args = parser.parse_args(argv)
    path = args.file if args.file.is_absolute() else PROJECT_ROOT / args.file
    if not path.exists():
        print(f"Datei nicht gefunden: {path}", file=sys.stderr)
        return 1

    try:
        plan = parse_workbook(path)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    db = SessionLocal()
    try:
        if plan.rejected:
            print(format_report(plan, committed=False))
            return 1
        apply_seed(db, plan)
        if plan.rejected:
            db.rollback()
            print(format_report(plan, committed=False))
            return 1
        if args.commit:
            db.commit()
            print(format_report(plan, committed=True))
            return 0
        db.rollback()
        print(format_report(plan, committed=False))
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
