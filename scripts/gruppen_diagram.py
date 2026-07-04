"""
Gruppenhierarchie aus gruppen.xlsx als Graphviz-Diagramm erzeugen.

Liest die Tabellenblätter „Hauptgruppen“ und „Untergruppen“ und rendert eine
Baumstruktur: jede Hauptgruppe mit ihren Untergruppen als Cluster.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

try:
    from graphviz import Digraph
    from graphviz.backend.execute import ExecutableNotFound
except ImportError:  # pragma: no cover - handled at runtime
    Digraph = None  # type: ignore[misc, assignment]
    ExecutableNotFound = OSError  # type: ignore[misc, assignment]


@dataclass(frozen=True)
class Subgroup:
    main_code: str
    sub_code: str
    name: str


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _resolve_path(path: str | Path) -> Path:
    p = Path(path)
    if not p.is_absolute():
        p = _project_root() / p
    return p


def _ensure_project_root() -> None:
    root = _project_root()
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def _format_code(value) -> str:
    raw = "" if value is None else str(value).strip()
    if raw == "":
        return ""
    if not raw.isdigit():
        raise ValueError(f"Ungültiger Gruppencode: {raw!r}")
    return f"{int(raw):03d}"


def load_groups(path: Path) -> tuple[dict[str, str], list[Subgroup]]:
    if not path.exists():
        raise FileNotFoundError(f"Gruppenschlüssel nicht gefunden: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Hauptgruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Hauptgruppen".')
        if "Untergruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Untergruppen".')

        main_groups: dict[str, str] = {}
        for code_raw, name_raw in wb["Hauptgruppen"].iter_rows(
            min_row=2, max_col=2, values_only=True
        ):
            code = _format_code(code_raw)
            if code == "":
                continue
            name = "" if name_raw is None else str(name_raw).strip()
            main_groups[code] = name

        subgroups: list[Subgroup] = []
        for main_raw, sub_raw, name_raw, *_ in wb["Untergruppen"].iter_rows(
            min_row=2, max_col=3, values_only=True
        ):
            main_code = _format_code(main_raw)
            sub_code = _format_code(sub_raw)
            if main_code == "" or sub_code == "":
                continue
            name = "" if name_raw is None else str(name_raw).strip()
            if name == "":
                raise ValueError(
                    f"Leere Untergruppen-Bezeichnung für {main_code}.{sub_code}."
                )
            subgroups.append(Subgroup(main_code, sub_code, name))

        return main_groups, subgroups
    finally:
        wb.close()


def _node_label(code: str, name: str) -> str:
    if name:
        return f"{code}\\n{name}"
    return code


def build_diagram(
    main_groups: dict[str, str],
    subgroups: list[Subgroup],
) -> Digraph:
    dot = Digraph(
        name="gruppen",
        graph_attr={
            "rankdir": "TB",
            "splines": "ortho",
            "nodesep": "0.35",
            "ranksep": "0.55",
            "fontsize": "11",
            "fontname": "Helvetica",
        },
        node_attr={
            "shape": "box",
            "style": "rounded,filled",
            "fillcolor": "#f8f9fa",
            "color": "#6c757d",
            "fontname": "Helvetica",
            "fontsize": "10",
        },
        edge_attr={
            "color": "#adb5bd",
        },
    )

    by_main: dict[str, list[Subgroup]] = {}
    for subgroup in subgroups:
        by_main.setdefault(subgroup.main_code, []).append(subgroup)

    for main_code in sorted(by_main):
        main_name = main_groups.get(main_code, "")
        cluster_name = f"cluster_{main_code}"
        with dot.subgraph(name=cluster_name) as cluster:
            cluster.attr(
                label=_node_label(main_code, main_name),
                style="rounded,filled",
                color="#dee2e6",
                fillcolor="#ffffff",
                fontsize="12",
                fontname="Helvetica-Bold",
            )
            main_id = f"main_{main_code}"
            cluster.node(
                main_id,
                label=_node_label(main_code, main_name),
                fillcolor="#e7f1ff",
                color="#0d6efd",
                penwidth="1.5",
            )
            for subgroup in sorted(by_main[main_code], key=lambda s: s.sub_code):
                sub_id = f"sub_{main_code}_{subgroup.sub_code}"
                cluster.node(
                    sub_id,
                    label=_node_label(subgroup.sub_code, subgroup.name),
                )
                cluster.edge(main_id, sub_id)

    return dot


def render_diagram(
    main_groups: dict[str, str],
    subgroups: list[Subgroup],
    output_stem: Path,
    fmt: str,
) -> Path:
    if Digraph is None:
        raise ImportError(
            "Das Paket 'graphviz' ist nicht installiert. "
            "Bitte ausführen: pip install graphviz"
        )

    dot = build_diagram(main_groups, subgroups)
    output_stem.parent.mkdir(parents=True, exist_ok=True)

    try:
        rendered = dot.render(
            filename=str(output_stem),
            format=fmt,
            cleanup=True,
        )
    except ExecutableNotFound as exc:
        raise RuntimeError(
            "Graphviz-Binary nicht gefunden. "
            "Bitte installieren, z. B. mit: brew install graphviz"
        ) from exc

    return Path(rendered)


def build_argparser() -> argparse.ArgumentParser:
    root = _project_root()
    parser = argparse.ArgumentParser(
        description="Haupt- und Untergruppen aus gruppen.xlsx als Diagramm erzeugen.",
    )
    parser.add_argument(
        "-i",
        "--input",
        default=str(root / "data" / "gruppen.xlsx"),
        help="Gruppenschlüssel (.xlsx, Standard: data/gruppen.xlsx)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=str(root / "data" / "gruppen_diagram"),
        help="Ausgabedatei ohne Endung (Standard: data/gruppen_diagram)",
    )
    parser.add_argument(
        "-f",
        "--format",
        default="png",
        choices=("png", "svg", "pdf"),
        help="Ausgabeformat (Standard: png)",
    )
    return parser


def main() -> None:
    _ensure_project_root()
    args = build_argparser().parse_args()

    input_path = _resolve_path(args.input)
    output_stem = _resolve_path(args.output)

    try:
        main_groups, subgroups = load_groups(input_path)
        output_file = render_diagram(main_groups, subgroups, output_stem, args.format)
    except (FileNotFoundError, ValueError, ImportError, RuntimeError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    print(f"Fertig: {output_file}")
    print(f"  Hauptgruppen:  {len(main_groups)}")
    print(f"  Untergruppen:  {len(subgroups)}")


if __name__ == "__main__":
    main()
