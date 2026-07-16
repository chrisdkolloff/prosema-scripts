"""
Weclapp-Bezugsquellen-Import aus Masterdatei und Produktgruppen-Rabatten erzeugen.

Liest die Spaltenüberschriften direkt aus der Weclapp-Importvorlage (CSV) und
füllt nur die gemappten Spalten; Rabatte werden anhand der Spalte
Rabattkategorie_Lieferant aus produktgruppen_rabatte.csv nachgeschlagen.
"""

from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

DISCOUNT_PRICE_TYPE = "DISCOUNT_PCT"

RABATT_CATEGORY_COLUMN = "Rabattkategorie_Lieferant"

MASTER_COLUMNS = {
    "ARTIKELNAME": "PROSEMA Kurztext",
    "Lieferantenartikelnummer": "Artikelnr.",
    "Lieferanten Firmenname": "Lieferanten Firmenname",
    "LIEFERANTENNUMMER": "Lieferantennummer",
    "Bruttokaufpreis": "Verkaufspreis €, BE",
    "Bruttopreis des zugehörigen Verkaufsartikels": "Verkaufspreis €, BE",
    "Verkaufsartikel-Nummer": "Prosema Artikelnummer",
    "Artikel-Mengeneinheit": "Basiseinheitencode",
}

FIXED_COLUMNS = {
    "Serienartikel": "ja",
    "Dropshipping möglich": "ja",
    "Primäre Bezugsquelle": "ja",
    "Währung": "EUR",
    "Verkaufsartikel-Währung": "EUR",
    "Vertriebsweg": "GROSS1",
    "Preis-Eintritt": "01.08.2026",
}

DISCOUNT_COLUMNS = (
    "Zu- und Abschläge Preisart 1",
    "Zu- und Abschläge Wert 1",
    "Zu- und Abschläge Preisart 2",
    "Zu- und Abschläge Wert 2",
)

FORCE_EMPTY_COLUMNS = {
    "Bruttopreis des zugehörigen Verkaufsartikels",
}


@dataclass
class GenerationStats:
    rows_read: int = 0
    rows_written: int = 0
    rows_without_discount: int = 0
    skipped_missing_article: int = 0
    warnings: list[str] = field(default_factory=list)

    def summary_lines(self) -> list[str]:
        lines = [
            f"Masterzeilen gelesen: {self.rows_read}",
            f"Zeilen geschrieben:   {self.rows_written}",
            f"  — davon ohne Rabatt:              {self.rows_without_discount}",
            f"Zeilen übersprungen:  {self.skipped_missing_article}",
            f"  — fehlende Prosema Artikelnummer: {self.skipped_missing_article}",
        ]
        if self.warnings:
            lines.append(f"Warnungen:            {len(self.warnings)}")
        return lines


def _detect_csv_format(path: Path) -> tuple[str, str, bool]:
    """Return (encoding, delimiter, use_bom) for a Weclapp template CSV."""
    raw = path.read_bytes()
    has_bom = raw.startswith(b"\xef\xbb\xbf")
    candidates = ("utf-8-sig", "utf-8", "cp1252", "latin-1") if has_bom else (
        "utf-8",
        "cp1252",
        "latin-1",
    )
    for encoding in candidates:
        try:
            raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        write_encoding = "utf-8-sig" if has_bom and encoding.startswith("utf-8") else encoding
        return write_encoding, ";", has_bom
    return ("utf-8-sig" if has_bom else "utf-8"), ";", has_bom


def read_template_headers(path: Path) -> tuple[list[str], str, str, bool]:
    encoding, delimiter, has_bom = _detect_csv_format(path)
    with open(path, encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        headers = next(reader)
    if not headers:
        raise ValueError(f"Importvorlage enthält keine Spaltenüberschriften: {path}")
    return headers, encoding, delimiter, has_bom


def parse_discount_percent(value: object) -> int:
    """Parse '50%', '0', '–' etc. into an integer percentage for Weclapp."""
    if value is None:
        return 0
    text = str(value).strip()
    if text in {"", "–", "-", "—", "n/a", "N/A"}:
        return 0
    text = text.replace("%", "").strip()
    text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"Ungültiger Rabattwert: {value!r}") from exc
    return int(number) if number.is_integer() else int(round(number))


def load_discount_table(path: Path) -> dict[str, tuple[int, int]]:
    with open(path, encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    discounts: dict[str, tuple[int, int]] = {}
    for row in rows:
        kategorie = str(row.get("Kategorie", "")).strip()
        if not kategorie:
            continue
        discounts[kategorie] = (
            parse_discount_percent(row.get("Grundrabatt")),
            parse_discount_percent(row.get("Kundenrabatt")),
        )
    return discounts


def _article_label(value: object) -> str:
    if value is None:
        return "(ohne Artikelnummer)"
    text = str(value).strip()
    return text or "(ohne Artikelnummer)"


def validate_discount_categories(
    master_rows: list[dict[str, object]],
    discounts: dict[str, tuple[int, int]],
) -> tuple[list[str], list[str]]:
    """Return warnings for missing categories and errors for unknown ones."""
    warnings: list[str] = []
    unknown: dict[str, list[str]] = {}

    for master_row in master_rows:
        article_raw = master_row["Prosema Artikelnummer"]
        if article_raw is None or str(article_raw).strip() == "":
            continue

        category_raw = master_row[RABATT_CATEGORY_COLUMN]
        category = "" if category_raw is None else str(category_raw).strip()
        if not category:
            warnings.append(
                f"{_article_label(article_raw)} — keine Rabattkategorie in "
                f"{RABATT_CATEGORY_COLUMN!r}"
            )
            continue

        if category not in discounts:
            unknown.setdefault(category, []).append(_article_label(article_raw))

    errors: list[str] = []
    for category in sorted(unknown):
        articles = ", ".join(unknown[category])
        errors.append(
            f"Rabattkategorie {category!r} nicht in produktgruppen_rabatte.csv "
            f"(Artikel: {articles})"
        )
    return warnings, errors


def format_output_value(column: str, value: object) -> str:
    if value is None:
        return ""
    if column in ("Bruttokaufpreis", "Bruttopreis des zugehörigen Verkaufsartikels") and isinstance(
        value, (int, float)
    ):
        number = float(value)
        formatted = f"{number:.2f}".rstrip("0").rstrip(".")
        return formatted.replace(".", ",")
    if column == "LIEFERANTENNUMMER" and isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def read_master_rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    header_row = next(row_iter)
    headers = ["" if cell is None else str(cell) for cell in header_row]
    required_columns = (*MASTER_COLUMNS.values(), RABATT_CATEGORY_COLUMN)
    missing = [src for src in required_columns if src not in headers]
    if missing:
        wb.close()
        raise ValueError(
            f"Masterdatei fehlen Spalten: {', '.join(missing)} ({path})"
        )
    col_index = {name: headers.index(name) for name in required_columns}
    rows: list[dict[str, object]] = []
    for row in row_iter:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rows.append({name: row[col_index[name]] for name in required_columns})
    wb.close()
    return rows


def generate_weclapp_import(
    master_path: Path,
    template_path: Path,
    rabatte_path: Path,
    output_path: Path,
) -> GenerationStats:
    headers, encoding, delimiter, _has_bom = read_template_headers(template_path)
    discounts = load_discount_table(rabatte_path)
    master_rows = read_master_rows(master_path)

    category_warnings, category_errors = validate_discount_categories(
        master_rows, discounts
    )
    if category_errors:
        message = "Unbekannte Rabattkategorien:\n" + "\n".join(
            f"  {line}" for line in category_errors
        )
        raise ValueError(message)

    stats = GenerationStats(
        rows_read=len(master_rows),
        warnings=category_warnings,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding=encoding, newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=headers,
            delimiter=delimiter,
            lineterminator="\n",
            extrasaction="ignore",
        )
        writer.writeheader()

        for master_row in master_rows:
            article_raw = master_row["Prosema Artikelnummer"]
            if article_raw is None or str(article_raw).strip() == "":
                stats.skipped_missing_article += 1
                continue

            category_raw = master_row[RABATT_CATEGORY_COLUMN]
            category = "" if category_raw is None else str(category_raw).strip()
            discount = discounts.get(category) if category else None
            out_row = {column: "" for column in headers}

            for out_col, master_col in MASTER_COLUMNS.items():
                if out_col not in headers:
                    raise ValueError(
                        f"Spalte {out_col!r} fehlt in der Importvorlage."
                    )
                if out_col in FORCE_EMPTY_COLUMNS:
                    continue
                out_row[out_col] = format_output_value(
                    out_col, master_row[master_col]
                )

            for out_col, fixed in FIXED_COLUMNS.items():
                if out_col not in headers:
                    raise ValueError(
                        f"Spalte {out_col!r} fehlt in der Importvorlage."
                    )
                out_row[out_col] = fixed

            if discount is not None:
                grundrabatt, kundenrabatt = discount
                preisart1, wert1, preisart2, wert2 = DISCOUNT_COLUMNS
                for col in DISCOUNT_COLUMNS:
                    if col not in headers:
                        raise ValueError(
                            f"Rabatt-Spalte {col!r} fehlt in der Importvorlage."
                        )
                out_row[preisart1] = DISCOUNT_PRICE_TYPE
                out_row[wert1] = str(grundrabatt)
                out_row[preisart2] = DISCOUNT_PRICE_TYPE
                out_row[wert2] = str(kundenrabatt)
            else:
                stats.rows_without_discount += 1

            writer.writerow(out_row)
            stats.rows_written += 1

    return stats


def _project_root() -> Path:
    from scripts.paths import PROJECT_ROOT

    return PROJECT_ROOT


def _resolve_path(path: str | Path) -> Path:
    from scripts.paths import resolve_path

    return resolve_path(path)


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    try:
        stats = generate_weclapp_import(
            _resolve_path(params["input"]),
            _resolve_path(params["template"]),
            _resolve_path(params["rabatte"]),
            _resolve_path(params["output"]),
        )
    except PermissionError as exc:
        raise PermissionError(
            f"Konnte {params['output']} nicht speichern — ist die Datei geöffnet?"
        ) from exc

    details = stats.summary_lines()
    if stats.warnings:
        details.append("Warnung: Artikel ohne Rabattkategorie:")
        details.extend(f"  {warning}" for warning in stats.warnings)

    return RunResult(
        summary=f"Fertig: {params['output']}  ({stats.rows_written} Zeilen)",
        details=details,
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    template_default = (
        "data/SupplySourcesWeclapp DemoImportfile_de (28.10.2024)(1).csv"
    )
    return JobSpec(
        id="weclapp_import",
        title="Weclapp-Import erzeugen",
        description=(
            "Bezugsquellen-Import für Weclapp aus der Masterdatei erzeugen. "
            "Rabatte werden anhand der Rabattkategorie_Lieferant nachgeschlagen."
        ),
        fields=(
            FieldSpec(
                "input",
                "Masterdatei",
                FieldKind.FILE_IN,
                "input/input.xlsx",
            ),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "output/export/weclapp_import.csv",
                output_name="weclapp_import.csv",
            ),
            FieldSpec(
                "template",
                "Weclapp-Importvorlage",
                FieldKind.FILE_IN,
                template_default,
                advanced=True,
            ),
            FieldSpec(
                "rabatte",
                "Produktgruppen-Rabatte",
                FieldKind.FILE_IN,
                "data/produktgruppen_rabatte.csv",
                advanced=True,
            ),
        ),
        run=run_job,
    )


def build_argparser() -> argparse.ArgumentParser:
    root = _project_root()
    parser = argparse.ArgumentParser(
        description="Weclapp-Bezugsquellen-Import aus Masterdatei erzeugen.",
    )
    parser.add_argument(
        "master",
        nargs="?",
        default=str(root / "input" / "input.xlsx"),
        help="Masterdatei (.xlsx)",
    )
    parser.add_argument(
        "template",
        nargs="?",
        default=str(
            root
            / "data"
            / "SupplySourcesWeclapp DemoImportfile_de (28.10.2024)(1).csv"
        ),
        help="Weclapp-Importvorlage (.csv)",
    )
    parser.add_argument(
        "rabatte",
        nargs="?",
        default=str(root / "data" / "produktgruppen_rabatte.csv"),
        help="Produktgruppen-Rabatte (.csv)",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(root / "output" / "export" / "weclapp_import.csv"),
        help="Ausgabe-CSV",
    )
    return parser


def main() -> None:
    _ensure_project_root()
    parser = build_argparser()
    args = parser.parse_args()

    master_path = _resolve_path(args.master)
    template_path = _resolve_path(args.template)
    rabatte_path = _resolve_path(args.rabatte)
    output_path = _resolve_path(args.output)

    for label, path in (
        ("Masterdatei", master_path),
        ("Importvorlage", template_path),
        ("Rabattdatei", rabatte_path),
    ):
        if not path.exists():
            sys.exit(f"{label} nicht gefunden: {path}")

    try:
        stats = generate_weclapp_import(
            master_path, template_path, rabatte_path, output_path
        )
    except (ValueError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    if stats.warnings:
        print("Warnung: Artikel ohne Rabattkategorie:", file=sys.stderr)
        for warning in stats.warnings:
            print(f"  {warning}", file=sys.stderr)

    print(f"Fertig: {output_path}")
    for line in stats.summary_lines():
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
