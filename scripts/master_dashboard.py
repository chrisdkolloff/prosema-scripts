"""
Interaktives Plotly-Dashboard für die Excel-Masterliste erzeugen.

Liest alle Zeilen und Spalten der Masterdatei und erzeugt eine HTML-Seite mit
Kategorie-Filtern, Textsuche, Übersichtsdiagrammen und einer durchsuchbaren Tabelle.
"""

from __future__ import annotations

import json
import sys
import webbrowser
from pathlib import Path

from openpyxl import load_workbook

try:
    import plotly.graph_objects as go
except ImportError:  # pragma: no cover - handled at runtime
    go = None  # type: ignore[assignment]


FILTER_COLUMNS = (
    "Hauptgruppe",
    "Untergruppe",
    "Kategorie",
    "Datenstatus",
    "Lieferanten Firmenname",
    "Kategorie2",
    "Produktfamilie",
)

TABLE_COLUMNS = (
    "Prosema Artikelnummer",
    "Artikelnr.",
    "PROSEMA Kurztext",
    "Hauptgruppe",
    "Untergruppe",
    "Kategorie",
    "Datenstatus",
    "Lieferanten Firmenname",
    "Verkaufspreis €, BE",
    "Grundmaterial",
    "Farbe",
)

SEARCH_COLUMNS = (
    "Prosema Artikelnummer",
    "Artikelnr.",
    "PROSEMA Kurztext",
    "Beschreibung",
    "Referenz (Matchcode)",
)


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _resolve_path(path: str | Path) -> Path:
    p = Path(path)
    if not p.is_absolute():
        p = _project_root() / p
    return p


def _ensure_project_root() -> None:
    root = _project_root()
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_master_data(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists():
        raise FileNotFoundError(f"Masterdatei nicht gefunden: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter)
        headers = ["" if cell is None else str(cell).strip() for cell in header_row]
        if not any(headers):
            raise ValueError(f"Masterdatei enthält keine Spaltenüberschriften: {path}")

        rows: list[dict[str, str]] = []
        for row in row_iter:
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            record = {
                headers[idx]: _cell_text(row[idx]) if idx < len(row) else ""
                for idx in range(len(headers))
            }
            rows.append(record)
        return headers, rows
    finally:
        wb.close()


def _count_by(rows: list[dict[str, str]], column: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        key = row.get(column, "") or "(leer)"
        counts[key] = counts.get(key, 0) + 1
    return counts


def build_overview_figures(rows: list[dict[str, str]]) -> tuple[go.Figure, go.Figure]:
    main_counts = _count_by(rows, "Hauptgruppe")
    sorted_main = sorted(main_counts.items(), key=lambda item: (-item[1], item[0]))

    bar_fig = go.Figure(
        go.Bar(
            x=[count for _, count in sorted_main],
            y=[name for name, _ in sorted_main],
            orientation="h",
            marker_color="#2e7d32",
            hovertemplate="<b>%{y}</b><br>%{x} Artikel<extra></extra>",
        )
    )
    bar_fig.update_layout(
        title="Artikel pro Hauptgruppe",
        margin=dict(l=180, r=20, t=50, b=30),
        height=max(320, 28 * len(sorted_main)),
        xaxis_title="Anzahl Artikel",
        yaxis=dict(autorange="reversed"),
    )

    treemap_ids = ["root"]
    treemap_labels = ["Masterliste"]
    treemap_parents = [""]
    treemap_values = [len(rows)]

    for main_name, main_count in sorted_main:
        main_id = f"main::{main_name}"
        treemap_ids.append(main_id)
        treemap_labels.append(main_name)
        treemap_parents.append("root")
        treemap_values.append(main_count)

        sub_counts: dict[str, int] = {}
        for row in rows:
            if (row.get("Hauptgruppe", "") or "(leer)") != main_name:
                continue
            sub_name = row.get("Untergruppe", "") or "(leer)"
            sub_counts[sub_name] = sub_counts.get(sub_name, 0) + 1

        for sub_name, sub_count in sorted(
            sub_counts.items(), key=lambda item: (-item[1], item[0])
        ):
            treemap_ids.append(f"sub::{main_name}::{sub_name}")
            treemap_labels.append(sub_name)
            treemap_parents.append(main_id)
            treemap_values.append(sub_count)

    treemap_fig = go.Figure(
        go.Treemap(
            ids=treemap_ids,
            labels=treemap_labels,
            parents=treemap_parents,
            values=treemap_values,
            branchvalues="total",
            hovertemplate="<b>%{label}</b><br>%{value} Artikel<extra></extra>",
        )
    )
    treemap_fig.update_layout(
        title="Haupt- und Untergruppen",
        margin=dict(l=10, r=10, t=50, b=10),
        height=520,
    )
    return bar_fig, treemap_fig


def _figure_json(fig: go.Figure) -> str:
    return fig.to_json()


def _filter_options(rows: list[dict[str, str]], column: str) -> list[str]:
    values = sorted({row.get(column, "") or "(leer)" for row in rows}, key=str.casefold)
    return values


def _dashboard_columns(headers: list[str]) -> list[str]:
    wanted = set(FILTER_COLUMNS) | set(TABLE_COLUMNS) | set(SEARCH_COLUMNS)
    return [header for header in headers if header in wanted]


def _slim_rows(rows: list[dict[str, str]], columns: list[str]) -> list[dict[str, str]]:
    return [{column: row.get(column, "") for column in columns} for row in rows]


def _dashboard_html(
    *,
    source_name: str,
    column_count: int,
    rows: list[dict[str, str]],
    bar_fig: go.Figure,
    treemap_fig: go.Figure,
) -> str:
    columns = list(rows[0].keys()) if rows else list(TABLE_COLUMNS)
    payload = {
        "sourceName": source_name,
        "columnCount": column_count,
        "rows": rows,
        "filterColumns": [column for column in FILTER_COLUMNS if column in columns],
        "tableColumns": [column for column in TABLE_COLUMNS if column in columns],
        "searchColumns": [column for column in SEARCH_COLUMNS if column in columns],
        "filterOptions": {
            column: _filter_options(rows, column)
            for column in FILTER_COLUMNS
            if column in columns
        },
        "barFigure": json.loads(_figure_json(bar_fig)),
        "treemapFigure": json.loads(_figure_json(treemap_fig)),
    }
    data_json = json.dumps(payload, ensure_ascii=False)
    data_json = data_json.replace("</", "<\\/")

    return f"""<!DOCTYPE html>
<html lang="de">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Masterliste — Dashboard</title>
  <script src="https://cdn.plot.ly/plotly-3.6.0.min.js"></script>
  <style>
    :root {{
      color-scheme: light dark;
      --bg: #f4f6f8;
      --panel: #ffffff;
      --text: #1f2933;
      --muted: #52606d;
      --border: #d9e2ec;
      --accent: #2e7d32;
    }}
    @media (prefers-color-scheme: dark) {{
      :root {{
        --bg: #111827;
        --panel: #1f2937;
        --text: #f9fafb;
        --muted: #9ca3af;
        --border: #374151;
        --accent: #4ade80;
      }}
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    header {{
      padding: 20px 24px 8px;
    }}
    h1 {{
      margin: 0 0 6px;
      font-size: 1.5rem;
    }}
    .subtitle {{
      color: var(--muted);
      font-size: 0.95rem;
    }}
    .stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      padding: 0 24px 16px;
    }}
    .stat {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 12px 16px;
      min-width: 140px;
    }}
    .stat strong {{
      display: block;
      font-size: 1.35rem;
    }}
    .stat span {{
      color: var(--muted);
      font-size: 0.85rem;
    }}
    .filters {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 12px;
      padding: 0 24px 16px;
    }}
    .filters label {{
      display: flex;
      flex-direction: column;
      gap: 6px;
      font-size: 0.85rem;
      color: var(--muted);
    }}
    select, input[type="search"], button {{
      font: inherit;
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid var(--border);
      background: var(--panel);
      color: var(--text);
    }}
    .search-row {{
      display: flex;
      gap: 12px;
      align-items: end;
      padding: 0 24px 16px;
      flex-wrap: wrap;
    }}
    .search-row label {{
      flex: 1 1 320px;
      display: flex;
      flex-direction: column;
      gap: 6px;
      font-size: 0.85rem;
      color: var(--muted);
    }}
    button {{
      cursor: pointer;
      background: var(--accent);
      color: #fff;
      border: none;
      min-width: 120px;
    }}
    .charts {{
      display: grid;
      grid-template-columns: minmax(320px, 1fr) minmax(320px, 1.2fr);
      gap: 16px;
      padding: 0 24px 16px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      padding: 8px;
      overflow: hidden;
    }}
    .table-wrap {{
      margin: 0 24px 24px;
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 12px;
      overflow: auto;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 0.88rem;
    }}
    thead th {{
      position: sticky;
      top: 0;
      background: var(--panel);
      border-bottom: 1px solid var(--border);
      text-align: left;
      padding: 10px 12px;
      white-space: nowrap;
    }}
    tbody td {{
      border-bottom: 1px solid var(--border);
      padding: 8px 12px;
      vertical-align: top;
    }}
    tbody tr:hover {{
      background: rgba(46, 125, 50, 0.08);
    }}
    .table-meta {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      padding: 12px 16px;
      color: var(--muted);
      font-size: 0.9rem;
      flex-wrap: wrap;
    }}
    .pager {{
      display: flex;
      gap: 8px;
      align-items: center;
    }}
    .pager button {{
      min-width: auto;
      padding: 6px 12px;
      background: var(--panel);
      color: var(--text);
      border: 1px solid var(--border);
    }}
    .pager button:disabled {{
      opacity: 0.45;
      cursor: default;
    }}
    @media (max-width: 960px) {{
      .charts {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Masterliste Dashboard</h1>
    <div class="subtitle" id="source-label"></div>
  </header>

  <section class="stats" id="stats"></section>

  <section class="filters" id="filter-bar"></section>

  <div class="search-row">
    <label>
      Textsuche
      <input id="search-input" type="search" placeholder="Artikelnummer, Kurztext, Beschreibung …" />
    </label>
    <button type="button" id="reset-filters">Filter zurücksetzen</button>
  </div>

  <section class="charts">
    <div class="panel"><div id="bar-chart" style="height:100%; min-height:320px;"></div></div>
    <div class="panel"><div id="treemap-chart" style="height:520px;"></div></div>
  </section>

  <div class="table-wrap">
    <div class="table-meta">
      <div id="table-summary"></div>
      <div class="pager">
        <button type="button" id="prev-page">Zurück</button>
        <span id="page-label"></span>
        <button type="button" id="next-page">Weiter</button>
      </div>
    </div>
    <table>
      <thead id="table-head"></thead>
      <tbody id="table-body"></tbody>
    </table>
  </div>

  <script id="dashboard-data" type="application/json">{data_json}</script>
  <script>
    const DATA = JSON.parse(document.getElementById("dashboard-data").textContent);
    const PAGE_SIZE = 75;
    let filteredRows = DATA.rows.slice();
    let currentPage = 1;
    const activeFilters = Object.fromEntries(DATA.filterColumns.map((col) => [col, ""]));

    function escapeHtml(value) {{
      return String(value)
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");
    }}

    function matchesSearch(row, query) {{
      if (!query) return true;
      const needle = query.toLowerCase();
      return DATA.searchColumns.some((col) => (row[col] || "").toLowerCase().includes(needle));
    }}

    function applyFilters() {{
      const query = document.getElementById("search-input").value.trim();
      filteredRows = DATA.rows.filter((row) => {{
        if (!matchesSearch(row, query)) return false;
        return DATA.filterColumns.every((col) => {{
          const selected = activeFilters[col];
          if (!selected) return true;
          const value = row[col] || "(leer)";
          return value === selected;
        }});
      }});
      currentPage = 1;
      renderAll();
    }}

    function buildFilterBar() {{
      const container = document.getElementById("filter-bar");
      container.innerHTML = "";
      for (const col of DATA.filterColumns) {{
        const options = DATA.filterOptions[col] || [];
        const label = document.createElement("label");
        label.textContent = col;
        const select = document.createElement("select");
        select.dataset.column = col;
        const allOption = document.createElement("option");
        allOption.value = "";
        allOption.textContent = "Alle";
        select.appendChild(allOption);
        for (const value of options) {{
          const option = document.createElement("option");
          option.value = value;
          option.textContent = value;
          select.appendChild(option);
        }}
        select.addEventListener("change", () => {{
          activeFilters[col] = select.value;
          applyFilters();
        }});
        label.appendChild(select);
        container.appendChild(label);
      }}
    }}

    function renderStats() {{
      const stats = document.getElementById("stats");
      const mainGroups = new Set(filteredRows.map((row) => row.Hauptgruppe || "(leer)")).size;
      const subGroups = new Set(
        filteredRows.map((row) => `${{row.Hauptgruppe || ""}}|${{row.Untergruppe || ""}}`)
      ).size;
      const withArticleNo = filteredRows.filter((row) => row["Prosema Artikelnummer"]).length;
      stats.innerHTML = `
        <div class="stat"><strong>${{filteredRows.length}}</strong><span>Artikel (gefiltert)</span></div>
        <div class="stat"><strong>${{DATA.rows.length}}</strong><span>Artikel (gesamt)</span></div>
        <div class="stat"><strong>${{mainGroups}}</strong><span>Hauptgruppen</span></div>
        <div class="stat"><strong>${{subGroups}}</strong><span>Untergruppen</span></div>
        <div class="stat"><strong>${{withArticleNo}}</strong><span>mit Artikelnummer</span></div>
      `;
    }}

    function countBy(rows, column) {{
      const counts = new Map();
      for (const row of rows) {{
        const key = row[column] || "(leer)";
        counts.set(key, (counts.get(key) || 0) + 1);
      }}
      return [...counts.entries()].sort((a, b) => b[1] - a[1] || a[0].localeCompare(b[0], "de"));
    }}

    function renderCharts() {{
      const mainCounts = countBy(filteredRows, "Hauptgruppe");
      const barFigure = JSON.parse(JSON.stringify(DATA.barFigure));
      barFigure.data[0].x = mainCounts.map(([, count]) => count);
      barFigure.data[0].y = mainCounts.map(([name]) => name);
      barFigure.layout.height = Math.max(320, 28 * mainCounts.length);
      barFigure.layout.yaxis = {{ autorange: "reversed" }};
      Plotly.react("bar-chart", barFigure.data, barFigure.layout, {{ responsive: true }});

      const treemapIds = ["root"];
      const treemapLabels = ["Auswahl"];
      const treemapParents = [""];
      const treemapValues = [filteredRows.length || 1];

      for (const [mainName, mainCount] of mainCounts) {{
        const mainId = `main::${{mainName}}`;
        treemapIds.push(mainId);
        treemapLabels.push(mainName);
        treemapParents.push("root");
        treemapValues.push(mainCount);

        const subCounts = new Map();
        for (const row of filteredRows) {{
          if ((row.Hauptgruppe || "(leer)") !== mainName) continue;
          const subName = row.Untergruppe || "(leer)";
          subCounts.set(subName, (subCounts.get(subName) || 0) + 1);
        }}
        for (const [subName, subCount] of [...subCounts.entries()].sort((a, b) => b[1] - a[1])) {{
          treemapIds.push(`sub::${{mainName}}::${{subName}}`);
          treemapLabels.push(subName);
          treemapParents.push(mainId);
          treemapValues.push(subCount);
        }}
      }}

      const treemapFigure = JSON.parse(JSON.stringify(DATA.treemapFigure));
      treemapFigure.data[0].ids = treemapIds;
      treemapFigure.data[0].labels = treemapLabels;
      treemapFigure.data[0].parents = treemapParents;
      treemapFigure.data[0].values = treemapValues;
      Plotly.react("treemap-chart", treemapFigure.data, treemapFigure.layout, {{ responsive: true }});
    }}

    function renderTable() {{
      const columns = DATA.tableColumns;
      const head = document.getElementById("table-head");
      head.innerHTML = `<tr>${{columns.map((col) => `<th>${{escapeHtml(col)}}</th>`).join("")}}</tr>`;

      const totalPages = Math.max(1, Math.ceil(filteredRows.length / PAGE_SIZE));
      if (currentPage > totalPages) currentPage = totalPages;
      const start = (currentPage - 1) * PAGE_SIZE;
      const pageRows = filteredRows.slice(start, start + PAGE_SIZE);

      const body = document.getElementById("table-body");
      body.innerHTML = pageRows.map((row) => `
        <tr>${{columns.map((col) => `<td>${{escapeHtml(row[col] || "")}}</td>`).join("")}}</tr>
      `).join("");

      document.getElementById("table-summary").textContent =
        `${{filteredRows.length}} Zeilen · Seite ${{currentPage}} von ${{totalPages}}`;
      document.getElementById("page-label").textContent = `${{currentPage}} / ${{totalPages}}`;
      document.getElementById("prev-page").disabled = currentPage <= 1;
      document.getElementById("next-page").disabled = currentPage >= totalPages;
    }}

    function setFilter(column, value) {{
      if (!DATA.filterColumns.includes(column)) return;
      activeFilters[column] = value;
      const select = document.querySelector(`select[data-column="${{column}}"]`);
      if (select) select.value = value;
      applyFilters();
    }}

    function bindTreemapClicks() {{
      const chart = document.getElementById("treemap-chart");
      chart.on("plotly_click", (event) => {{
        const label = event.points?.[0]?.label;
        const id = event.points?.[0]?.id || "";
        if (!label || label === "Auswahl") {{
          DATA.filterColumns.forEach((col) => setFilter(col, ""));
          return;
        }}
        if (id.startsWith("main::")) {{
          setFilter("Hauptgruppe", label);
          setFilter("Untergruppe", "");
          return;
        }}
        if (id.startsWith("sub::")) {{
          const parts = id.split("::");
          setFilter("Hauptgruppe", parts[1] || "");
          setFilter("Untergruppe", parts[2] || "");
        }}
      }});
    }}

    function renderAll() {{
      renderStats();
      renderCharts();
      renderTable();
    }}

    document.getElementById("source-label").textContent =
      `${{DATA.sourceName}} · ${{DATA.columnCount}} Spalten · ${{DATA.rows.length}} Zeilen`;
    document.getElementById("search-input").addEventListener("input", applyFilters);
    document.getElementById("reset-filters").addEventListener("click", () => {{
      DATA.filterColumns.forEach((col) => {{
        activeFilters[col] = "";
        const select = document.querySelector(`select[data-column="${{col}}"]`);
        if (select) select.value = "";
      }});
      document.getElementById("search-input").value = "";
      applyFilters();
    }});
    document.getElementById("prev-page").addEventListener("click", () => {{
      currentPage -= 1;
      renderTable();
    }});
    document.getElementById("next-page").addEventListener("click", () => {{
      currentPage += 1;
      renderTable();
    }});

    buildFilterBar();
    Plotly.newPlot("bar-chart", DATA.barFigure.data, DATA.barFigure.layout, {{ responsive: true }});
    Plotly.newPlot("treemap-chart", DATA.treemapFigure.data, DATA.treemapFigure.layout, {{ responsive: true }});
    bindTreemapClicks();
    renderAll();
  </script>
</body>
</html>
"""


def render_dashboard(
    headers: list[str],
    rows: list[dict[str, str]],
    output: Path,
    *,
    source_name: str,
) -> Path:
    if go is None:
        raise ImportError(
            "Das Paket 'plotly' ist nicht installiert. "
            "Bitte ausführen: pip install plotly"
        )

    bar_fig, treemap_fig = build_overview_figures(rows)
    dashboard_columns = _dashboard_columns(headers)
    output.parent.mkdir(parents=True, exist_ok=True)
    page = _dashboard_html(
        source_name=source_name,
        column_count=len(headers),
        rows=_slim_rows(rows, dashboard_columns),
        bar_fig=bar_fig,
        treemap_fig=treemap_fig,
    )
    output.write_text(page, encoding="utf-8")
    return output


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    input_path = _resolve_path(params["input"])
    output_path = _resolve_path(params["output"])

    headers, rows = load_master_data(input_path)
    output_file = render_dashboard(
        headers,
        rows,
        output_path,
        source_name=input_path.name,
    )

    opened = webbrowser.open(output_file.resolve().as_uri())
    details = [
        f"Zeilen: {len(rows)}",
        f"Spalten: {len(headers)}",
        f"Filter: {', '.join(col for col in FILTER_COLUMNS if col in headers)}",
    ]
    if not opened:
        details.append("Hinweis: Browser konnte nicht automatisch geöffnet werden.")
    return RunResult(
        summary=f"Fertig: {output_file}",
        details=details,
        show_success_dialog=False,
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="master_dashboard",
        title="Masterliste-Dashboard",
        description=(
            "Interaktives Plotly-Dashboard aus der Excel-Masterliste erzeugen. "
            "Filtern nach Kategorien, Textsuche und paginierte Tabellenansicht."
        ),
        fields=(
            FieldSpec(
                "input",
                "Masterdatei",
                FieldKind.FILE_IN,
                "input.xlsx",
            ),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "data/master_dashboard.html",
                output_name="master_dashboard.html",
            ),
        ),
        run=run_job,
    )


def main() -> None:
    _ensure_project_root()
    from gui.job_spec import args_to_params, build_argparser, coerce_params, validate_params

    parser = build_argparser(JOB_SPEC)
    args = parser.parse_args()
    params = coerce_params(JOB_SPEC, args_to_params(JOB_SPEC, args))
    try:
        validate_params(JOB_SPEC, params)
        result = run_job(params)
    except (FileNotFoundError, ValueError, ImportError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    print(result.summary)
    for line in result.details:
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
