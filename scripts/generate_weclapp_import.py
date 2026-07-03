"""
Weclapp-Bezugsquellen-Import aus Masterdatei und Produktgruppen-Rabatten erzeugen.

Liest die Spaltenüberschriften direkt aus der Weclapp-Importvorlage (CSV) und
füllt nur die gemappten Spalten; Rabatte werden anhand des Artikelnummern-Codes
MMM.SSS (mit Fallback auf MMM) aus produktgruppen_rabatte.csv nachgeschlagen.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

DISCOUNT_PRICE_TYPE = "DISCOUNT_PCT"

ARTICLE_PATTERN = re.compile(r"^(\d{3})\.(\d{3})\.(\d{4})$")

MASTER_COLUMNS = {
    "ARTIKELNAME": "PROSEMA Kurztext",
    "Lieferantenartikelnummer": "Artikelnr.",
    "Lieferanten Firmenname": "Lieferanten Firmenname",
    "LIEFERANTENNUMMER": "Lieferantennummer",
    "Bruttokaufpreis": "Verkaufspreis €, BE",
    "Verkaufsartikel-Nummer": "Prosema Artikelnummer",
}

FIXED_COLUMNS = {
    "Zu- und Abschläge Preisart 1": DISCOUNT_PRICE_TYPE,
    "Zu- und Abschläge Preisart 2": DISCOUNT_PRICE_TYPE,
    "Serienartikel": "ja",
    "Dropshipping möglich": "ja",
    "Primäre Bezugsquelle": "ja",
}

DISCOUNT_VALUE_COLUMNS = (
    "Zu- und Abschläge Wert 1",
    "Zu- und Abschläge Wert 2",
)


@dataclass
class GenerationStats:
    rows_read: int = 0
    rows_written: int = 0
    skipped_missing_article: int = 0
    skipped_malformed_article: int = 0
    skipped_no_discount: int = 0

    def summary_lines(self) -> list[str]:
        skipped = (
            self.skipped_missing_article
            + self.skipped_malformed_article
            + self.skipped_no_discount
        )
        return [
            f"Masterzeilen gelesen: {self.rows_read}",
            f"Zeilen geschrieben:   {self.rows_written}",
            f"Zeilen übersprungen:  {skipped}",
            f"  — fehlende Prosema Artikelnummer: {self.skipped_missing_article}",
            f"  — ungültige Artikelnummer:        {self.skipped_malformed_article}",
            f"  — kein Rabatt-Eintrag:            {self.skipped_no_discount}",
        ]


def _detect_csv_format(path: Path) -> tuple[str, str, bool]:
    """Return (encoding, delimiter, use_bom) for a Weclapp template CSV."""
    raw = path.read_bytes()
    has_bom = raw.startswith(b"\xef\xbb\xbf")
    candidates = ("utf-8-sig", "utf-8", "cp1252", "latin-1") if has_bom else (
        "utf-8",
        "cp1252",
        "latin-1",
    )
    for encoding in candidates:
        try:
            raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        write_encoding = "utf-8-sig" if has_bom and encoding.startswith("utf-8") else encoding
        return write_encoding, ";", has_bom
    return ("utf-8-sig" if has_bom else "utf-8"), ";", has_bom


def read_template_headers(path: Path) -> tuple[list[str], str, str, bool]:
    encoding, delimiter, has_bom = _detect_csv_format(path)
    with open(path, encoding=encoding, newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        headers = next(reader)
    if not headers:
        raise ValueError(f"Importvorlage enthält keine Spaltenüberschriften: {path}")
    return headers, encoding, delimiter, has_bom


def parse_discount_percent(value: object) -> int:
    """Parse '50%', '0', '–' etc. into an integer percentage for Weclapp."""
    if value is None:
        return 0
    text = str(value).strip()
    if text in {"", "–", "-", "—", "n/a", "N/A"}:
        return 0
    text = text.replace("%", "").strip()
    text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError as exc:
        raise ValueError(f"Ungültiger Rabattwert: {value!r}") from exc
    return int(number) if number.is_integer() else int(round(number))


def load_discount_table(path: Path) -> dict[str, tuple[int, int]]:
    with open(path, encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    discounts: dict[str, tuple[int, int]] = {}
    for row in rows:
        code = str(row.get("Code", "")).strip()
        if not code:
            continue
        discounts[code] = (
            parse_discount_percent(row.get("Grundrabatt")),
            parse_discount_percent(row.get("Kundenrabatt")),
        )
    return discounts


def lookup_discount(
    discounts: dict[str, tuple[int, int]],
    hauptgruppe: str,
    untergruppe: str,
) -> tuple[int, int] | None:
    # Most specific match first (MMM.SSS), then Hauptgruppe-only (MMM).
    full_code = f"{hauptgruppe}.{untergruppe}"
    if full_code in discounts:
        return discounts[full_code]
    if hauptgruppe in discounts:
        return discounts[hauptgruppe]
    return None


def parse_article_number(value: object) -> tuple[str, str] | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = ARTICLE_PATTERN.match(text)
    if not match:
        return None
    return match.group(1), match.group(2)


def format_output_value(column: str, value: object) -> str:
    if value is None:
        return ""
    if column == "Bruttokaufpreis" and isinstance(value, (int, float)):
        number = float(value)
        formatted = f"{number:.2f}".rstrip("0").rstrip(".")
        return formatted.replace(".", ",")
    if column == "LIEFERANTENNUMMER" and isinstance(value, (int, float)):
        return str(int(value))
    return str(value).strip()


def read_master_rows(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)
    header_row = next(row_iter)
    headers = ["" if cell is None else str(cell) for cell in header_row]
    missing = [src for src in MASTER_COLUMNS.values() if src not in headers]
    if missing:
        wb.close()
        raise ValueError(
            f"Masterdatei fehlen Spalten: {', '.join(missing)} ({path})"
        )
    col_index = {name: headers.index(name) for name in MASTER_COLUMNS.values()}
    rows: list[dict[str, object]] = []
    for row in row_iter:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        rows.append({name: row[col_index[name]] for name in MASTER_COLUMNS.values()})
    wb.close()
    return rows


def generate_weclapp_import(
    master_path: Path,
    template_path: Path,
    rabatte_path: Path,
    output_path: Path,
) -> GenerationStats:
    headers, encoding, delimiter, _has_bom = read_template_headers(template_path)
    discounts = load_discount_table(rabatte_path)
    master_rows = read_master_rows(master_path)

    stats = GenerationStats(rows_read=len(master_rows))
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", encoding=encoding, newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=headers,
            delimiter=delimiter,
            lineterminator="\n",
            extrasaction="ignore",
        )
        writer.writeheader()

        for master_row in master_rows:
            article_raw = master_row["Prosema Artikelnummer"]
            parsed = parse_article_number(article_raw)
            if article_raw is None or str(article_raw).strip() == "":
                stats.skipped_missing_article += 1
                continue
            if parsed is None:
                stats.skipped_malformed_article += 1
                continue

            haupt, unter = parsed
            discount = lookup_discount(discounts, haupt, unter)
            if discount is None:
                stats.skipped_no_discount += 1
                continue

            grundrabatt, kundenrabatt = discount
            out_row = {column: "" for column in headers}

            for out_col, master_col in MASTER_COLUMNS.items():
                if out_col not in headers:
                    raise ValueError(
                        f"Spalte {out_col!r} fehlt in der Importvorlage."
                    )
                out_row[out_col] = format_output_value(
                    out_col, master_row[master_col]
                )

            for out_col, fixed in FIXED_COLUMNS.items():
                if out_col not in headers:
                    raise ValueError(
                        f"Spalte {out_col!r} fehlt in der Importvorlage."
                    )
                out_row[out_col] = fixed

            wert1, wert2 = DISCOUNT_VALUE_COLUMNS
            if wert1 not in headers or wert2 not in headers:
                raise ValueError("Rabatt-Wert-Spalten fehlen in der Importvorlage.")
            out_row[wert1] = str(grundrabatt)
            out_row[wert2] = str(kundenrabatt)

            writer.writerow(out_row)
            stats.rows_written += 1

    return stats


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _resolve_path(path: str | Path) -> Path:
    p = Path(path)
    if not p.is_absolute():
        p = _project_root() / p
    return p


def _ensure_project_root() -> None:
    root = _project_root()
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def run_job(params: dict):
    from gui.job_spec import RunResult, coerce_params, validate_params

    params = coerce_params(JOB_SPEC, params)
    validate_params(JOB_SPEC, params)

    try:
        stats = generate_weclapp_import(
            _resolve_path(params["input"]),
            _resolve_path(params["template"]),
            _resolve_path(params["rabatte"]),
            _resolve_path(params["output"]),
        )
    except PermissionError as exc:
        raise PermissionError(
            f"Konnte {params['output']} nicht speichern — ist die Datei geöffnet?"
        ) from exc

    return RunResult(
        summary=f"Fertig: {params['output']}  ({stats.rows_written} Zeilen)",
        details=stats.summary_lines(),
    )


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    template_default = (
        "data/SupplySourcesWeclapp DemoImportfile_de (28.10.2024)(1).csv"
    )
    return JobSpec(
        id="weclapp_import",
        title="Weclapp-Import erzeugen",
        description=(
            "Bezugsquellen-Import für Weclapp aus der Masterdatei erzeugen. "
            "Rabatte werden anhand der Prosema-Artikelnummer nachgeschlagen."
        ),
        fields=(
            FieldSpec(
                "input",
                "Masterdatei",
                FieldKind.FILE_IN,
                "230703-masterdatei.xlsx",
            ),
            FieldSpec(
                "output",
                "Ausgabedatei",
                FieldKind.FILE_OUT,
                "weclapp_import.csv",
                output_name="weclapp_import.csv",
            ),
            FieldSpec(
                "template",
                "Weclapp-Importvorlage",
                FieldKind.FILE_IN,
                template_default,
                advanced=True,
            ),
            FieldSpec(
                "rabatte",
                "Produktgruppen-Rabatte",
                FieldKind.FILE_IN,
                "data/produktgruppen_rabatte.csv",
                advanced=True,
            ),
        ),
        run=run_job,
    )


def build_argparser() -> argparse.ArgumentParser:
    root = _project_root()
    parser = argparse.ArgumentParser(
        description="Weclapp-Bezugsquellen-Import aus Masterdatei erzeugen.",
    )
    parser.add_argument(
        "master",
        nargs="?",
        default=str(root / "230703-masterdatei.xlsx"),
        help="Masterdatei (.xlsx)",
    )
    parser.add_argument(
        "template",
        nargs="?",
        default=str(
            root
            / "data"
            / "SupplySourcesWeclapp DemoImportfile_de (28.10.2024)(1).csv"
        ),
        help="Weclapp-Importvorlage (.csv)",
    )
    parser.add_argument(
        "rabatte",
        nargs="?",
        default=str(root / "data" / "produktgruppen_rabatte.csv"),
        help="Produktgruppen-Rabatte (.csv)",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(root / "weclapp_import.csv"),
        help="Ausgabe-CSV",
    )
    return parser


def main() -> None:
    _ensure_project_root()
    parser = build_argparser()
    args = parser.parse_args()

    master_path = _resolve_path(args.master)
    template_path = _resolve_path(args.template)
    rabatte_path = _resolve_path(args.rabatte)
    output_path = _resolve_path(args.output)

    for label, path in (
        ("Masterdatei", master_path),
        ("Importvorlage", template_path),
        ("Rabattdatei", rabatte_path),
    ):
        if not path.exists():
            sys.exit(f"{label} nicht gefunden: {path}")

    try:
        stats = generate_weclapp_import(
            master_path, template_path, rabatte_path, output_path
        )
    except (ValueError, OSError) as exc:
        sys.exit(f"Abbruch: {exc}")

    print(f"Fertig: {output_path}")
    for line in stats.summary_lines():
        print(line)


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    main()
