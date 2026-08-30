"""Create weclapp articles from the standardized import CSV template."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from core.article_fields import IMPORT_COLUMNS
from core.article_payload import (
    BOOLEAN_CUSTOM_ATTRS,
    DEFAULTS,
    FALSE_VALUES,
    LIST_CUSTOM_ATTRS,
    NUMBER_PLACEHOLDER,
    STRING_CUSTOM_ATTRS,
    TRUE_VALUES,
    ARTICLE_NAME_FIELD,
    ARTICLE_NUMBER_FIELD,
    LONG_TEXT_FIELD,
    _norm,
    _parse_bool,
    _row_value,
    get_row_value,
    label_variants,
    row_to_payload,
)

GROUP_CODE_RE = re.compile(r"-\s*(\d+)\s*$")
_GROUP_LABEL_RE = re.compile(r"^(.*?)\s*-\s*(\d+)\s*$")
RESTRICTED_SELECT_COLUMNS: tuple[str, ...] = (
    "Artikeltyp",
    "Einheit",
    "Aktiv",
    "Im Verkauf",
    "Steuersatz",
    "Im Shop verfügbar",
    "Im Shop aktiv",
    "Bestand übertragen",
    "Gewichtseinheit",
    "Bodenleger",
    "Dachdecker",
    "Landschaftsgärtner",
    "Plattenleger",
)


@dataclass
class ImportErrorRow:
    article_number: str
    message: str


@dataclass
class ImportStats:
    rows_read: int = 0
    created: int = 0
    skipped: int = 0
    errors: list[ImportErrorRow] = field(default_factory=list)
    created_ids: list[tuple[str, str]] = field(default_factory=list)

    def summary_lines(self) -> list[str]:
        lines = [
            f"Zeilen gelesen: {self.rows_read}",
            f"Erstellt:       {self.created}",
            f"Übersprungen:   {self.skipped}",
            f"Fehler:         {len(self.errors)}",
        ]
        for article_number, article_id in self.created_ids:
            lines.append(f"  OK {article_number} -> {article_id}")
        for error in self.errors:
            lines.append(f"  FEHLER {error.article_number}: {error.message}")
        return lines


def _ensure_project_root() -> None:
    root = Path(__file__).resolve().parents[2]
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))


def _load_schema() -> dict[str, Any]:
    from scripts.paths import DATA_DIR

    path = DATA_DIR / "weclapp_article_create_schema.json"
    return json.loads(path.read_text(encoding="utf-8"))


def _split_group_label(text: str) -> tuple[str, int | None]:
    """Split ``Name - NNN`` into (name, integer code). Code padding is ignored."""
    cleaned = _norm(text)
    match = _GROUP_LABEL_RE.match(cleaned)
    if match is None:
        return cleaned, None
    return match.group(1).strip(), int(match.group(2))


class LookupTables:
    def __init__(self, schema: dict[str, Any]) -> None:
        lookups = schema.get("lookups") or {}
        self.units_by_key: dict[str, str] = {}
        self.unit_names: list[str] = []
        for unit in lookups.get("units") or []:
            unit_id = str(unit.get("id") or "")
            name = _norm(unit.get("name"))
            if name:
                self.unit_names.append(name)
            for key in (unit.get("name"), unit.get("description"), unit_id):
                text = _norm(key).lower()
                if text:
                    self.units_by_key[text] = unit_id
        self.unit_names = sorted(set(self.unit_names), key=str.lower)

        self.categories_by_name: dict[str, str] = {}
        self.category_names: list[str] = []
        for category in lookups.get("articleCategories") or []:
            name = _norm(category.get("name"))
            category_id = str(category.get("id") or "")
            if name and category_id:
                self.categories_by_name[name.lower()] = category_id
                self.category_names.append(name)
        self.category_names = sorted(set(self.category_names), key=str.lower)

        self.attrs_by_label: dict[str, dict[str, Any]] = {}
        for attr in lookups.get("customAttributes") or []:
            label = _norm(attr.get("label"))
            if label:
                self.attrs_by_label[label] = attr

    def unit_id(self, value: str) -> str:
        key = _norm(value).lower()
        if not key:
            raise ValueError("Einheit fehlt")
        unit_id = self.units_by_key.get(key)
        if not unit_id:
            raise ValueError(f"Unbekannte Einheit: {value}")
        return unit_id

    def category_id(self, value: str) -> str | None:
        key = _norm(value).lower()
        if not key:
            return None
        category_id = self.categories_by_name.get(key)
        if not category_id:
            raise ValueError(f"Unbekannte Kategorie: {value}")
        return category_id

    def list_value_id(self, attr_label: str, value: str) -> str:
        """Resolve a selectable-value id.

        Group labels may disagree on zero-padding (``Nivelliersystem - 010`` in
        the registry vs ``Nivelliersystem - 10`` in weclapp). Match on the
        display name plus the integer code; never rewrite either side's label.
        Returns weclapp's own option id so the payload uses weclapp's literal.
        """
        attr = self.attrs_by_label.get(attr_label)
        if not attr:
            raise ValueError(f"Zusatzfeld nicht gefunden: {attr_label}")
        wanted = _norm(value)
        if not wanted:
            raise ValueError(f"Ungültiger Wert für {attr_label}: {value}")
        wanted_lower = wanted.lower()
        wanted_name, wanted_code = _split_group_label(wanted)

        exact_id: str | None = None
        code_name_id: str | None = None
        prefix_id: str | None = None
        for option in attr.get("selectableValues") or []:
            option_value = _norm(option.get("value"))
            option_id = str(option.get("id") or "")
            if not option_value or not option_id:
                continue
            if option_value.lower() == wanted_lower:
                exact_id = option_id
                break
            opt_name, opt_code = _split_group_label(option_value)
            if (
                wanted_code is not None
                and opt_code is not None
                and wanted_code == opt_code
                and opt_name.lower() == wanted_name.lower()
            ):
                code_name_id = option_id
            if opt_name.lower() == wanted_lower:
                prefix_id = option_id

        if exact_id is not None:
            return exact_id
        if code_name_id is not None:
            return code_name_id
        if prefix_id is not None:
            return prefix_id
        raise ValueError(f"Ungültiger Wert für {attr_label}: {value}")

    def list_value_literal(self, attr_label: str, value: str) -> str:
        """Return weclapp's own selectable-value string for ``value``."""
        attr = self.attrs_by_label.get(attr_label)
        if not attr:
            raise ValueError(f"Zusatzfeld nicht gefunden: {attr_label}")
        option_id = self.list_value_id(attr_label, value)
        for option in attr.get("selectableValues") or []:
            if str(option.get("id") or "") == option_id:
                return _norm(option.get("value"))
        raise ValueError(f"Ungültiger Wert für {attr_label}: {value}")

    def attr_id(self, label: str) -> str:
        attr = self.attrs_by_label.get(label)
        if not attr:
            raise ValueError(f"Zusatzfeld nicht gefunden: {label}")
        return str(attr.get("id"))

    def list_values(self, attr_label: str) -> list[str]:
        attr = self.attrs_by_label.get(attr_label) or {}
        values = [
            _norm(option.get("value"))
            for option in attr.get("selectableValues") or []
            if _norm(option.get("value"))
        ]
        return values


def dropdown_options(lookups: LookupTables | None = None) -> dict[str, list[str]]:
    lookups = lookups or LookupTables(_load_schema())
    schema = _load_schema()
    article_types = []
    tax_rates = []
    for item in schema.get("defaultedIfOmitted") or []:
        if item.get("field") == "articleType":
            article_types = list(item.get("allowed") or [])
        if item.get("field") == "taxRateType":
            tax_rates = list(item.get("allowed") or [])
    yes_no = ["Ja", "Nein"]
    return {
        "Artikeltyp": article_types,
        "Einheit": lookups.unit_names,
        "Kategorie": lookups.category_names,
        "Steuersatz": tax_rates,
        "Aktiv": yes_no,
        "Im Verkauf": yes_no,
        "Im Shop verfügbar": yes_no,
        "Im Shop aktiv": yes_no,
        "Bestand übertragen": yes_no,
        "Bodenleger": yes_no,
        "Dachdecker": yes_no,
        "Landschaftsgärtner": yes_no,
        "Plattenleger": yes_no,
        "Gewichtseinheit": ["kg", "g", "lb"],
        "Hauptgruppe": lookups.list_values("Hauptwarengruppe (Auswahl)"),
        "Untergruppe": lookups.list_values("Warengruppe (Auswahl)"),
    }


def parse_group_code(value: str, width: int) -> str | None:
    text = _norm(value)
    if not text:
        return None
    match = GROUP_CODE_RE.search(text)
    if not match:
        return None
    return f"{int(match.group(1)):0{width}d}"


def _master_path() -> Path:
    from scripts.paths import INPUT_DIR

    return INPUT_DIR / "input.xlsx"


def load_existing_running_numbers(master_path: Path | None = None) -> dict[tuple[str, str], int]:
    from core.numbering import Scheme

    scheme = Scheme()
    pattern = scheme.pattern()
    counters: dict[tuple[str, str], int] = {}
    path = master_path or _master_path()
    if not path.is_file():
        return counters

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(min_row=1, values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            return counters
        number_names = set(label_variants(ARTICLE_NUMBER_FIELD))
        try:
            article_idx = next(
                i for i, header in enumerate(headers) if header in number_names
            )
        except StopIteration:
            return counters
        for values in rows:
            if article_idx >= len(values):
                continue
            value = values[article_idx]
            match = pattern.match(str(value).strip()) if value not in (None, "") else None
            if not match:
                continue
            key = (match.group(1), match.group(2))
            counters[key] = max(counters.get(key, 0), int(match.group(3)))
    finally:
        workbook.close()
    return counters


def generate_article_numbers(
    rows: list[dict[str, str]],
    *,
    master_path: Path | None = None,
) -> tuple[list[dict[str, str]], dict[str, int]]:
    from core.numbering import Scheme

    scheme = Scheme()
    pattern = scheme.pattern()
    counters = load_existing_running_numbers(master_path)
    assigned = 0
    placeholders = 0
    kept = 0

    prepared: list[dict[str, str]] = []
    for row in rows:
        prepared.append(dict(row))

    for row in prepared:
        existing = get_row_value(row, ARTICLE_NUMBER_FIELD)
        match = pattern.match(existing) if existing else None
        if not match:
            continue
        main = parse_group_code(row.get("Hauptgruppe", ""), scheme.main_width)
        sub = parse_group_code(row.get("Untergruppe", ""), scheme.sub_width)
        if main and sub and match.group(1) == main and match.group(2) == sub:
            key = (main, sub)
            counters[key] = max(counters.get(key, 0), int(match.group(3)))

    errors: list[str] = []
    for index, row in enumerate(prepared, start=2):
        main = parse_group_code(row.get("Hauptgruppe", ""), scheme.main_width)
        sub = parse_group_code(row.get("Untergruppe", ""), scheme.sub_width)
        supplier = _norm(row.get("Lieferantenartikelnummer")) or f"Zeile {index}"
        if not main or not sub:
            row[ARTICLE_NUMBER_FIELD] = NUMBER_PLACEHOLDER
            placeholders += 1
            missing = []
            if not main:
                missing.append("Hauptgruppe")
            if not sub:
                missing.append("Untergruppe")
            errors.append(
                f"{supplier}: bitte { ' und '.join(missing) } ausfüllen, "
                "dann Artikelnummern erneut erzeugen."
            )
            continue

        existing = get_row_value(row, ARTICLE_NUMBER_FIELD)
        match = pattern.match(existing) if existing else None
        if match and match.group(1) == main and match.group(2) == sub:
            kept += 1
            continue

        key = (main, sub)
        current = counters.get(key)
        nxt = scheme.start if current is None else current + scheme.step
        if nxt > scheme.max_running:
            raise OverflowError(
                f"Gruppe {main}.{sub} hat das Maximum überschritten."
            )
        counters[key] = nxt
        row[ARTICLE_NUMBER_FIELD] = scheme.format(main, sub, nxt)
        assigned += 1

    return prepared, {
        "assigned": assigned,
        "kept": kept,
        "placeholders": placeholders,
        "errors": errors,
    }


def save_import_rows(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=IMPORT_COLUMNS, delimiter=";")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _norm(row.get(column, "")) for column in IMPORT_COLUMNS})


def validate_import_rows(rows: list[dict[str, str]]) -> list[ImportErrorRow]:
    lookups = LookupTables(_load_schema())
    errors: list[ImportErrorRow] = []
    seen: set[str] = set()
    for row in rows:
        article_number = _row_value(row, ARTICLE_NUMBER_FIELD) or "(ohne Nummer)"
        try:
            payload = row_to_payload(row, lookups)
        except ValueError as exc:
            errors.append(ImportErrorRow(article_number, str(exc)))
            continue
        number = payload["articleNumber"]
        if number in seen:
            errors.append(ImportErrorRow(number, "Artikelnummer ist in der Datei doppelt"))
        seen.add(number)
    return errors


COLUMN_ALIASES = {
    "Artikelnr.": "Lieferantenartikelnummer",
    "Hauptwarengruppe": "Hauptgruppe",
    "Warengruppe": "Untergruppe",
    "Prosema Artikelnummer": ARTICLE_NUMBER_FIELD,
    "PROSEMA Kurztext": ARTICLE_NAME_FIELD,
    "Prosema Kurztext": ARTICLE_NAME_FIELD,
    "PROSEMA Langtext": LONG_TEXT_FIELD,
    "Prosema Langtext": LONG_TEXT_FIELD,
}


def load_import_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        if not reader.fieldnames:
            raise ValueError("CSV ohne Kopfzeile")
        name_headers = set(label_variants(ARTICLE_NAME_FIELD))
        if not any(name in name_headers for name in reader.fieldnames):
            raise ValueError(f"Pflichtspalten fehlen: {ARTICLE_NAME_FIELD}")
        rows: list[dict[str, str]] = []
        for raw in reader:
            if not any(_norm(value) for value in raw.values()):
                continue
            normalized: dict[str, str] = {}
            for key, value in raw.items():
                target = COLUMN_ALIASES.get(key or "", key or "")
                text = _norm(value)
                if target and (target not in normalized or not normalized[target]):
                    normalized[target] = text
            row = {column: normalized.get(column, "") for column in IMPORT_COLUMNS}
            rows.append(row)
        return rows


def write_template(path: Path, *, include_dummy: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    dummy = {
        ARTICLE_NUMBER_FIELD: NUMBER_PLACEHOLDER,
        "Lieferantenartikelnummer": "TEST-SUP-001",
        "Hauptgruppe": "",
        "Untergruppe": "",
        ARTICLE_NAME_FIELD: "TEST Dummy Artikel Import Pipeline",
        LONG_TEXT_FIELD: "Testdatensatz für den weclapp-Artikelimport. Kann gelöscht werden.",
        "Kurzbeschreibung": "TEST Dummy Artikel Import Pipeline",
        "Referenz (Matchcode)": "TEST-DUMMY",
        "GTIN (EAN-Nummer)": "",
        "Artikeltyp": "BASIC",
        "Einheit": "Stk.",
        "Kategorie": "Zubehör allgemein",
        "Aktiv": "Ja",
        "Im Verkauf": "Ja",
        "Steuersatz": "STANDARD",
        "Im Shop verfügbar": "Ja",
        "Im Shop aktiv": "Ja",
        "Bestand übertragen": "Ja",
        "Gewichtseinheit": "kg",
        "Grundmaterial": "Testdaten",
        "Oberfläche": "",
        "Farbe": "Testfarbe",
        "Produktfamilie": "",
        "Rabattcode": "",
        "Verkaufseinheit": "Stk.",
        "Verpackung": "1",
        "VPE 1": "",
        "VPE 2": "",
        "VPE 3": "",
        "Breite in mm": "",
        "Länge in cm": "",
        "Höhe in mm": "",
        "Bodenleger": "Nein",
        "Dachdecker": "Nein",
        "Landschaftsgärtner": "Nein",
        "Plattenleger": "Nein",
        "Artikelbeschreibung HTML": "<p>Testdatensatz für den weclapp-Artikelimport.</p>",
        "Nettogewicht kg": "0.1",
        "Produkt-ID (Prosema)": "",
        "Varianten-ID (Prosema)": "",
    }
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=IMPORT_COLUMNS, delimiter=";")
        writer.writeheader()
        if include_dummy:
            writer.writerow(dummy)


MSG_CLI_RETIRED = (
    "Der CLI-/Desktop-Artikelimport ist abgelöst. "
    "Neue Artikel über die Web-App anlegen: /artikel-registrierung"
)


def _article_exists(client, article_number: str) -> dict[str, Any] | None:
    data = client.get(
        "/article",
        params={"pageSize": 1, "articleNumber-eq": article_number},
    )
    rows = (data or {}).get("result") or []
    return rows[0] if rows else None


def import_articles(
    input_path: Path,
    *,
    dry_run: bool = True,
    limit: int | None = None,
    article_numbers: set[str] | None = None,
) -> ImportStats:
    """Retired: do not POST articles from the CLI or desktop GUI.

    Shared helpers in this module (columns, lookups, validation) remain for the
    web Artikelregistrierung. Offline CSV checks can use ``validate_import_rows``.
    """
    raise RuntimeError(MSG_CLI_RETIRED)


def run_job(params: dict):
    from gui.job_spec import RunResult

    return RunResult(summary=MSG_CLI_RETIRED, details=[])


def _build_job_spec():
    from gui.job_spec import FieldKind, FieldSpec, JobSpec

    return JobSpec(
        id="weclapp_import_articles",
        title="weclapp-Artikel anlegen (abgelöst)",
        description=MSG_CLI_RETIRED,
        fields=(
            FieldSpec(
                "input",
                "Import-CSV",
                FieldKind.FILE_IN,
                "data/weclapp_article_import_template.csv",
            ),
        ),
        run=run_job,
    )


def main(argv: list[str] | None = None) -> int:
    _ensure_project_root()
    from scripts.paths import DATA_DIR, resolve_path

    parser = argparse.ArgumentParser(
        description=(
            "ABGELÖST — Artikel anlegen über /artikel-registrierung. "
            "Dieses Skript schreibt nur noch das Template oder bricht ab."
        ),
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DATA_DIR / "weclapp_article_import_template.csv",
        help="Import-CSV (Semikolon, utf-8-sig)",
    )
    parser.add_argument(
        "--create",
        action="store_true",
        help="(abgelöst)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="(abgelöst)",
    )
    parser.add_argument(
        "--article",
        action="append",
        default=[],
        help="(abgelöst)",
    )
    parser.add_argument(
        "--write-template",
        action="store_true",
        help="Template-CSV schreiben und beenden",
    )
    args = parser.parse_args(argv)

    input_path = resolve_path(args.input)
    if args.write_template:
        write_template(input_path)
        print(f"Template geschrieben: {input_path}")
        return 0

    print(MSG_CLI_RETIRED, file=sys.stderr)
    return 1


_ensure_project_root()
JOB_SPEC = _build_job_spec()


if __name__ == "__main__":
    raise SystemExit(main())
