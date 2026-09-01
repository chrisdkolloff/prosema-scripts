"""Map weclapp article records to PROSEMA master-list column names."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from html import unescape
from pathlib import Path
from typing import Any

MASTER_COLUMNS: tuple[str, ...] = (
    "Prosema Artikelnummer",
    "Kategorie",
    "Lieferanten Firmenname",
    "Artikelnr.",
    "Beschreibung",
    "PROSEMA Kurztext",
    "PROSEMA Langtext",
    "SEO Meta Title",
    "SEO Meta Description",
    "Fokus-Keyword",
    "Shopify Tags",
    "Datenstatus",
    "Referenz (Matchcode)",
    "Hauptgruppe",
    "Untergruppe",
    "Produktfamilie",
    "Grundmaterial",
    "Oberfläche",
    "Farbe",
    "Breite mm",
    "Länge cm",
    "Höhe mm",
    "Basiseinheitencode",
    "Nettogewicht kg",
    "Verpackung",
    "GTIN (EAN-Nummer)",
    "Zolltarifnummer",
    "Einkaufspreis EUR netto",
    "Nettoverkaufspreis CHF",
    "Einkaufspreis Prosema",
    "Verkaufspreis",
    "Verkaufspreis o. Verkn.",
    "Verkaufseinheit",
    "Rabattkategorie_Lieferant",
    "VPE 1",
    "VPE 2",
    "VPE 3",
    "Lieferantennummer",
    "Produktfoto 1",
)

WECLAPP_EXTRA_COLUMNS: tuple[str, ...] = (
    "weclapp Artikel-ID",
    "weclapp Aktiv",
    "weclapp Artikeltyp",
    "weclapp Erstellt am",
    "weclapp Geändert am",
    "weclapp Kategorie-ID",
    "weclapp Bezugsquelle-ID",
    "weclapp Einheit-ID",
    "weclapp Version",
    "weclapp Im Verkauf",
    "weclapp Steuersatz",
    "weclapp Kurzbeschreibung",
    "weclapp Breite (m)",
    "weclapp Im Shop aktiv (Prosema)",
    "weclapp Im Shop verfügbar (Prosema)",
    "weclapp Bestand übertragen (Prosema)",
    "weclapp Produkt-ID (Prosema)",
    "weclapp Varianten-ID (Prosema)",
    "weclapp Gewichtseinheit",
)

EXPORT_COLUMNS: tuple[str, ...] = MASTER_COLUMNS + WECLAPP_EXTRA_COLUMNS

# Historic master-list headers still present in older files and snapshots.
MASTER_COLUMN_RENAMES: dict[str, str] = {
    "Verkaufspreis €, BE": "Nettoverkaufspreis CHF",
}


def apply_master_column_renames(
    headers: list[str],
    rows: list[dict[str, str]],
) -> tuple[list[str], list[dict[str, str]]]:
    """Rewrite historic headers onto the current master-list names."""
    if not MASTER_COLUMN_RENAMES:
        return headers, rows
    if not any(
        old in headers or any(old in row for row in rows)
        for old in MASTER_COLUMN_RENAMES
    ):
        return headers, rows
    seen: set[str] = set()
    new_headers: list[str] = []
    for header in headers:
        current = MASTER_COLUMN_RENAMES.get(header, header)
        if current in seen:
            continue
        new_headers.append(current)
        seen.add(current)
    remapped_rows: list[dict[str, str]] = []
    for row in rows:
        out: dict[str, str] = {}
        for key, value in row.items():
            current = MASTER_COLUMN_RENAMES.get(key, key)
            if current not in out or not out[current]:
                out[current] = value
        remapped_rows.append(out)
    return new_headers, remapped_rows


# (display column name, source column in raw export, or None for empty column)
_EXPORT_DISPLAY_MAPPING: tuple[tuple[str, str | None], ...] = (
    ("Prosema-Artikelnummer", "Prosema Artikelnummer"),
    ("Lieferanten-Artikelnummer", "Artikelnr."),
    ("Matchcode", "Referenz (Matchcode)"),
    ("Hauptgruppe", "Hauptgruppe"),
    ("Untergruppe", "Untergruppe"),
    ("Produktfamilie", "Produktfamilie"),
    ("Zielgruppe", "Datenstatus"),
    ("Lieferantennummer", "Lieferantennummer"),
    ("Lieferanten Firmenname", "Lieferanten Firmenname"),
    ("UVP Lieferant", "Einkaufspreis EUR netto"),
    ("UVP Lieferant Währung", None),
    ("Rabattkategorie Lieferant", "Rabattkategorie_Lieferant"),
    ("Rabatt 1", None),
    ("Rabatt 2", None),
    ("Einkaufspreis Prosema", "Einkaufspreis Prosema"),
    ("Einkaufspreis Prosema Währung", None),
    ("Zuschlag (%)", None),
    ("Verkaufspreis Prosema", "Verkaufspreis"),
    ("Verkaufspreis Prosema Währung", None),
    ("Basiseinheiten", "Basiseinheitencode"),
    ("Verkaufseinheit", "Verkaufseinheit"),
    ("Prosema Kurztext", "PROSEMA Kurztext"),
    ("Prosema Langtext", "PROSEMA Langtext"),
    ("SEO Meta Title", "SEO Meta Title"),
    ("Grundmaterial", "Grundmaterial"),
    ("Oberfläche", "Oberfläche"),
    ("Farbe", "Farbe"),
    ("Breite (mm)", "Breite mm"),
    ("Länge (cm)", "Länge cm"),
    ("Höhe (mm)", "Höhe mm"),
    ("Nettogewicht (kg)", "Nettogewicht kg"),
    ("Verpackung (Stk)", "Verpackung"),
    ("VPE 1", "VPE 1"),
    ("VPE 2", "VPE 2"),
    ("VPE 3", "VPE 3"),
    ("EAN-Nummer", "GTIN (EAN-Nummer)"),
    ("Zolltarifnummer", "Zolltarifnummer"),
    ("Shopify Produkte", "Kategorie"),
    ("weclapp Produkt-ID (Prosema)", "weclapp Produkt-ID (Prosema)"),
    ("weclapp Varianten-ID (Prosema)", "weclapp Varianten-ID (Prosema)"),
    ("weclapp Im Shop aktiv (Prosema)", "weclapp Im Shop aktiv (Prosema)"),
    ("weclapp Im Shop verfügbar (Prosema)", "weclapp Im Shop verfügbar (Prosema)"),
    ("weclapp Bestand übertragen (Prosema)", "weclapp Bestand übertragen (Prosema)"),
    ("weclapp Artikel-ID", "weclapp Artikel-ID"),
    ("weclapp Aktiv", "weclapp Aktiv"),
    ("weclapp Artikeltyp", "weclapp Artikeltyp"),
    ("weclapp Im Verkauf", "weclapp Im Verkauf"),
    ("weclapp Steuersatz", "weclapp Steuersatz"),
    ("weclapp Kurzbeschreibung", "weclapp Kurzbeschreibung"),
    ("weclapp Breite (m)", "weclapp Breite (m)"),
    ("weclapp Gewichtseinheit", "weclapp Gewichtseinheit"),
    ("weclapp Kategorie-ID", "weclapp Kategorie-ID"),
    ("weclapp Bezugsquelle-ID", "weclapp Bezugsquelle-ID"),
    ("weclapp Einheit-ID", "weclapp Einheit-ID"),
    ("weclapp Version", "weclapp Version"),
    ("weclapp Erstellt am", "weclapp Erstellt am"),
    ("weclapp Geändert am", "weclapp Geändert am"),
)

EXPORT_DISPLAY_COLUMNS: tuple[str, ...] = tuple(
    display for display, _ in _EXPORT_DISPLAY_MAPPING
)


def transform_export_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Map raw master-list export rows to import-template column names."""
    transformed: list[dict[str, str]] = []
    for raw in rows:
        row = {
            display: raw.get(source, "") if source else ""
            for display, source in _EXPORT_DISPLAY_MAPPING
        }
        if row.get("EAN-Nummer"):
            row["EAN-Nummer"] = _format_ean(row["EAN-Nummer"])
        transformed.append(row)
    return transformed


def master_list_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Strip weclapp-only columns and normalize master-list fields."""
    normalized: list[dict[str, str]] = []
    for row in rows:
        master = {column: row.get(column, "") for column in MASTER_COLUMNS}
        for old, new in MASTER_COLUMN_RENAMES.items():
            if not master.get(new) and row.get(old):
                master[new] = row[old]
        if master["GTIN (EAN-Nummer)"]:
            master["GTIN (EAN-Nummer)"] = _format_ean(master["GTIN (EAN-Nummer)"])
        normalized.append(master)
    return normalized


def write_master_list_xlsx(rows: list[dict[str, str]], path: Path) -> None:
    """Write a PROSEMA master list (.xlsx) with EAN stored as text."""
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    master_rows = master_list_rows(rows)
    ean_col = MASTER_COLUMNS.index("GTIN (EAN-Nummer)") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Masterliste"
    for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(master_rows, start=2):
        for col_idx, header in enumerate(MASTER_COLUMNS, start=1):
            value = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value or None)
            if col_idx == ean_col and value:
                cell.number_format = "@"
    wb.save(path)


CUSTOM_ATTRIBUTE_TO_MASTER: dict[str, str] = {
    "Grundmaterial": "Grundmaterial",
    "Oberfläche": "Oberfläche",
    "Farbe": "Farbe",
    "Breite in mm": "Breite mm",
    "Länge in cm": "Länge cm",
    "Höhe in mm": "Höhe mm",
    "Verpackung": "Verpackung",
    "Verkaufseinheit": "Verkaufseinheit",
    "Rabattcode": "Rabattkategorie_Lieferant",
    "Produktfamilie": "Produktfamilie",
    "VPE 1": "VPE 1",
    "VPE 2": "VPE 2",
    "VPE 3": "VPE 3",
    "Artikelbeschreibung (Prosema)": "Beschreibung",
    "Produkt-ID (Prosema)": "Kategorie",
}

CUSTOM_ATTRIBUTE_TO_EXTRA: dict[str, str] = {
    "Im Shop aktiv (Prosema)": "weclapp Im Shop aktiv (Prosema)",
    "Im Shop verfügbar (Prosema)": "weclapp Im Shop verfügbar (Prosema)",
    "Bestand übertragen (Prosema)": "weclapp Bestand übertragen (Prosema)",
    "Produkt-ID (Prosema)": "weclapp Produkt-ID (Prosema)",
    "Varianten-ID (Prosema)": "weclapp Varianten-ID (Prosema)",
    "Gewichtseinheit": "weclapp Gewichtseinheit",
}


def normalize_verkaufseinheit(value: str) -> str:
    return value.replace("€", "CHF")


def strip_html(value: object) -> str:
    if value is None:
        return ""
    text = unescape(str(value))
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p\s*>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _format_ean(value: object) -> str:
    """Format EAN/GTIN as integer string without scientific notation."""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(round(value)))
    text = str(value).strip()
    if not text:
        return ""
    normalized = text.replace(",", ".")
    if "e" in normalized.lower():
        try:
            return str(int(round(float(normalized))))
        except (ValueError, OverflowError):
            return text
    if normalized.replace(".", "", 1).isdigit():
        number = float(normalized)
        return str(int(round(number)))
    return text


def _format_decimal(value: object) -> str:
    if value is None or value == "":
        return ""
    text = str(value).strip()
    try:
        number = float(text.replace(",", "."))
    except ValueError:
        return text
    formatted = f"{number:.3f}".rstrip("0").rstrip(".")
    return formatted.replace(".", ",")


def _format_weclapp_timestamp(value: object) -> str:
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if number >= 1_000_000_000_000:
        number /= 1000
    try:
        from datetime import datetime

        return datetime.fromtimestamp(number).strftime("%d.%m.%Y %H:%M")
    except (OSError, OverflowError, ValueError):
        return str(value)


def _format_bool(value: object) -> str:
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    text = str(value or "").strip().lower()
    if text in {"true", "ja", "1"}:
        return "Ja"
    if text in {"false", "nein", "0"}:
        return "Nein"
    return str(value or "")


def _custom_attribute_value(entry: dict[str, Any]) -> str:
    for key in (
        "stringValue",
        "numberValue",
        "booleanValue",
        "dateValue",
        "selectedValueId",
        "entityId",
    ):
        if key not in entry:
            continue
        value = entry[key]
        if value is None:
            continue
        if isinstance(value, bool):
            return "Ja" if value else "Nein"
        return str(value)
    return ""


def _custom_attributes_by_label(
    article: dict[str, Any],
    attribute_labels: dict[str, str],
) -> dict[str, str]:
    values: dict[str, str] = {}
    for entry in article.get("customAttributes") or []:
        if not isinstance(entry, dict):
            continue
        attr_id = str(entry.get("attributeDefinitionId", "")).strip()
        label = attribute_labels.get(attr_id, "")
        if not label:
            continue
        values[label] = _custom_attribute_value(entry)
    return values


def _first_price(prices: object, *, sales_channel: str | None = None) -> str:
    if not isinstance(prices, list):
        return ""
    for entry in prices:
        if not isinstance(entry, dict):
            continue
        if sales_channel and entry.get("salesChannel") != sales_channel:
            continue
        price = entry.get("price")
        if price is not None:
            return str(price)
    if prices and isinstance(prices[0], dict):
        price = prices[0].get("price")
        return "" if price is None else str(price)
    return ""


def _category_names(category_id: str, categories: dict[str, dict[str, Any]]) -> tuple[str, str]:
    if not category_id:
        return "", ""
    chain: list[str] = []
    current_id = category_id
    seen: set[str] = set()
    while current_id and current_id not in seen:
        seen.add(current_id)
        category = categories.get(current_id)
        if not category:
            break
        name = str(category.get("name") or "").strip()
        if name:
            chain.append(name)
        current_id = str(category.get("parentCategoryId") or "").strip()
    if not chain:
        return "", ""
    untergruppe = chain[0]
    hauptgruppe = chain[-1] if len(chain) > 1 else chain[0]
    return hauptgruppe, untergruppe


@dataclass
class MasterLookups:
    categories: dict[str, dict[str, Any]] = field(default_factory=dict)
    units: dict[str, dict[str, Any]] = field(default_factory=dict)
    customs_tariffs: dict[str, dict[str, Any]] = field(default_factory=dict)
    supply_sources: dict[str, dict[str, Any]] = field(default_factory=dict)
    parties: dict[str, dict[str, Any]] = field(default_factory=dict)
    attribute_labels: dict[str, str] = field(default_factory=dict)

    def category_names(self, category_id: str) -> tuple[str, str]:
        return _category_names(category_id, self.categories)

    def unit_name(self, unit_id: str) -> str:
        unit = self.units.get(unit_id or "")
        return str(unit.get("name") or "").strip() if unit else ""

    def customs_tariff_name(self, tariff_id: str) -> str:
        tariff = self.customs_tariffs.get(tariff_id or "")
        return str(tariff.get("name") or "").strip() if tariff else ""

    def supply_source(self, supply_source_id: str) -> dict[str, Any]:
        return self.supply_sources.get(supply_source_id or "", {})

    def party(self, party_id: str) -> dict[str, Any]:
        return self.parties.get(party_id or "", {})


def article_to_master_row(
    article: dict[str, Any],
    lookups: MasterLookups,
) -> dict[str, str]:
    attrs = _custom_attributes_by_label(article, lookups.attribute_labels)
    row = {column: "" for column in EXPORT_COLUMNS}

    hauptgruppe, untergruppe = lookups.category_names(
        str(article.get("articleCategoryId") or "")
    )
    supply = lookups.supply_source(str(article.get("primarySupplySourceId") or ""))
    supplier = lookups.party(str(supply.get("supplierId") or ""))

    row["Prosema Artikelnummer"] = str(article.get("articleNumber") or "")
    row["PROSEMA Kurztext"] = str(article.get("name") or "")
    row["PROSEMA Langtext"] = strip_html(article.get("longText"))
    row["Referenz (Matchcode)"] = str(article.get("matchCode") or "")
    row["GTIN (EAN-Nummer)"] = _format_ean(article.get("ean"))
    row["Hauptgruppe"] = hauptgruppe
    row["Untergruppe"] = untergruppe
    row["Basiseinheitencode"] = lookups.unit_name(str(article.get("unitId") or ""))
    row["Nettogewicht kg"] = _format_decimal(article.get("articleNetWeight"))
    row["Zolltarifnummer"] = lookups.customs_tariff_name(
        str(article.get("customsTariffNumberId") or "")
    )
    row["Nettoverkaufspreis CHF"] = _first_price(
        article.get("articlePrices"),
        sales_channel="GROSS1",
    )
    row["Einkaufspreis EUR netto"] = _first_price(supply.get("articlePrices"))
    row["Artikelnr."] = str(supply.get("articleNumber") or "")
    row["Lieferanten Firmenname"] = str(supplier.get("company") or "")
    row["Lieferantennummer"] = str(supplier.get("supplierNumber") or "")

    for attr_label, master_column in CUSTOM_ATTRIBUTE_TO_MASTER.items():
        value = attrs.get(attr_label, "")
        if value and not row.get(master_column):
            row[master_column] = value

    if not row["Länge cm"] and article.get("articleLength") not in (None, ""):
        try:
            row["Länge cm"] = str(int(round(float(article["articleLength"]) * 100)))
        except (TypeError, ValueError):
            pass
    if not row["Höhe mm"] and article.get("articleHeight") not in (None, ""):
        try:
            row["Höhe mm"] = str(int(round(float(article["articleHeight"]) * 1000)))
        except (TypeError, ValueError):
            pass

    row["weclapp Artikel-ID"] = str(article.get("id") or "")
    row["weclapp Aktiv"] = _format_bool(article.get("active"))
    row["weclapp Artikeltyp"] = str(article.get("articleType") or "")
    row["weclapp Erstellt am"] = _format_weclapp_timestamp(article.get("createdDate"))
    row["weclapp Geändert am"] = _format_weclapp_timestamp(article.get("lastModifiedDate"))
    row["weclapp Kategorie-ID"] = str(article.get("articleCategoryId") or "")
    row["weclapp Bezugsquelle-ID"] = str(article.get("primarySupplySourceId") or "")
    row["weclapp Einheit-ID"] = str(article.get("unitId") or "")
    row["weclapp Version"] = str(article.get("version") or "")
    row["weclapp Im Verkauf"] = _format_bool(article.get("availableInSale"))
    row["weclapp Steuersatz"] = str(article.get("taxRateType") or "")
    row["weclapp Kurzbeschreibung"] = str(article.get("shortDescription1") or "")
    if article.get("articleWidth") not in (None, ""):
        row["weclapp Breite (m)"] = _format_decimal(article.get("articleWidth"))

    for attr_label, extra_column in CUSTOM_ATTRIBUTE_TO_EXTRA.items():
        value = attrs.get(attr_label, "")
        if value:
            row[extra_column] = _format_bool(value)

    row["Verkaufseinheit"] = normalize_verkaufseinheit(row["Verkaufseinheit"])

    return row


def flat_weclapp_csv_to_master_row(row: dict[str, str]) -> dict[str, str]:
    """Best-effort mapping from a raw weclapp CSV row without API lookups."""
    master = {column: "" for column in EXPORT_COLUMNS}

    direct_map = {
        "articleNumber": "Prosema Artikelnummer",
        "name": "PROSEMA Kurztext",
        "matchCode": "Referenz (Matchcode)",
        "ean": "GTIN (EAN-Nummer)",
    }
    for source, target in direct_map.items():
        if row.get(source):
            value = row[source]
            if target == "GTIN (EAN-Nummer)":
                value = _format_ean(value)
            master[target] = value

    if row.get("longText"):
        master["PROSEMA Langtext"] = strip_html(row["longText"])
    if row.get("articleNetWeight"):
        master["Nettogewicht kg"] = _format_decimal(row["articleNetWeight"])
    if row.get("articlePrices"):
        master["Nettoverkaufspreis CHF"] = _first_price(json.loads(row["articlePrices"]))
    if row.get("active", "").lower() == "true":
        master["Datenstatus"] = "Aktiv"
        master["weclapp Aktiv"] = "Ja"
    elif row.get("active", "").lower() == "false":
        master["Datenstatus"] = "Inaktiv"
        master["weclapp Aktiv"] = "Nein"

    for key in ("id", "articleType", "unitId", "articleCategoryId", "primarySupplySourceId", "version"):
        if row.get(key):
            target = {
                "id": "weclapp Artikel-ID",
                "articleType": "weclapp Artikeltyp",
                "unitId": "weclapp Einheit-ID",
                "articleCategoryId": "weclapp Kategorie-ID",
                "primarySupplySourceId": "weclapp Bezugsquelle-ID",
                "version": "weclapp Version",
            }[key]
            master[target] = row[key]
    if row.get("createdDate"):
        master["weclapp Erstellt am"] = _format_weclapp_timestamp(row["createdDate"])
    if row.get("lastModifiedDate"):
        master["weclapp Geändert am"] = _format_weclapp_timestamp(row["lastModifiedDate"])
    if row.get("availableInSale"):
        master["weclapp Im Verkauf"] = _format_bool(row["availableInSale"])
    if row.get("taxRateType"):
        master["weclapp Steuersatz"] = row["taxRateType"]
    if row.get("shortDescription1"):
        master["weclapp Kurzbeschreibung"] = row["shortDescription1"]
    if row.get("articleWidth"):
        master["weclapp Breite (m)"] = _format_decimal(row["articleWidth"])

    for key, value in row.items():
        if not key.startswith("attr_") or not value:
            continue
        label = key.removeprefix("attr_")
        master_column = CUSTOM_ATTRIBUTE_TO_MASTER.get(label)
        if master_column:
            master[master_column] = value
        extra_column = CUSTOM_ATTRIBUTE_TO_EXTRA.get(label)
        if extra_column and value:
            master[extra_column] = _format_bool(value)
        elif label == "Produktfamilie":
            master["Produktfamilie"] = value
            if not master["Hauptgruppe"]:
                master["Hauptgruppe"] = value

    if row.get("attr_Produktfamilie") and not master["Hauptgruppe"]:
        master["Hauptgruppe"] = row["attr_Produktfamilie"]
    if row.get("attr_Rabattcode") and not master["Untergruppe"]:
        master["Untergruppe"] = row["attr_Rabattcode"]

    master["Verkaufseinheit"] = normalize_verkaufseinheit(master["Verkaufseinheit"])

    return master


def ensure_export_columns(
    headers: list[str],
    rows: list[dict[str, str]],
) -> tuple[list[str], list[dict[str, str]]]:
    """Ensure weclapp export snapshots expose all export columns in order."""
    headers, rows = apply_master_column_renames(headers, rows)
    header_set = set(headers)
    ordered = [column for column in EXPORT_COLUMNS if column in header_set]
    ordered.extend(header for header in headers if header not in ordered)
    missing = [column for column in EXPORT_COLUMNS if column not in header_set]
    if missing:
        ordered.extend(missing)
    if not missing:
        return ordered, rows
    return ordered, [{column: row.get(column, "") for column in ordered} for row in rows]


def is_master_format(headers: list[str]) -> bool:
    header_set = set(headers)
    return "Prosema Artikelnummer" in header_set and "articleNumber" not in header_set


def is_raw_weclapp_export(headers: list[str]) -> bool:
    header_set = set(headers)
    return "articleNumber" in header_set and "articleType" in header_set


def build_lookups(client, articles: list[dict[str, Any]]) -> MasterLookups:
    """Load reference data needed to map articles to master columns."""
    lookups = MasterLookups()

    for definition in client.iter_pages("customAttributeDefinition"):
        attr_id = str(definition.get("id", "")).strip()
        label = str(
            definition.get("label") or definition.get("attributeKey") or attr_id
        ).strip()
        if attr_id and label:
            lookups.attribute_labels[attr_id] = label

    for entity, store in (
        ("articleCategory", lookups.categories),
        ("unit", lookups.units),
        ("customsTariffNumber", lookups.customs_tariffs),
    ):
        for row in client.iter_pages(entity):
            store[str(row.get("id", ""))] = row

    for row in client.iter_pages("articleSupplySource"):
        lookups.supply_sources[str(row.get("id", ""))] = row

    supplier_ids = {
        str(row.get("supplierId") or "")
        for row in lookups.supply_sources.values()
        if row.get("supplierId")
    }
    for supplier_id in sorted(supplier_ids):
        lookups.parties[supplier_id] = client.get(f"/party/id/{supplier_id}")

    return lookups
