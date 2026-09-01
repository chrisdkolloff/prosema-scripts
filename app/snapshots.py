"""Artikelübersicht: snapshot pull, filtering, grid config, Excel export."""

from __future__ import annotations

import logging
import uuid
from dataclasses import dataclass
from datetime import datetime
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, false, func, or_, select, text
from sqlalchemy.orm import Session

from app.batches import JSPREADSHEET_CE_VERSION, JSUITES_VERSION
from app.config import settings
from app.excel_export import workbook_bytes, write_cell
from app.models import ArticleSnapshot, ArticleSnapshotRow, Job
from core.article_flatten import flatten_articles, snapshot_column_title
from core.article_payload import ARTICLE_NAME_FIELD, ARTICLE_NUMBER_FIELD, label_variants
from core.numbering import Scheme

logger = logging.getLogger(__name__)

ZURICH = ZoneInfo("Europe/Zurich")
GRID_PAGE_SIZE = 250
EXCEL_MAX_ROWS = 50_000

# Postgres is capped at 32 GiB with autogrow off. When the volume fills, the
# database goes read-only and the application fails in a way that looks like a
# bug rather than a capacity problem. Snapshots, uploads, and generated exports
# share that disk. Each article snapshot is ~8 MB; keep a short rolling window
# plus monthly archives, not unbounded history.
RETENTION_KEEP_RECENT = 20
RETENTION_KEEP_DAYS = 14
RETENTION_KEEP_MONTHS = 12
RETENTION_ORPHAN_DAYS = 7


@dataclass
class SnapshotFilters:
    query: str = ""
    hauptgruppe: str = ""
    untergruppe: str = ""
    nur_aktive: bool = True
    page: int = 1
    assistant_query_id: uuid.UUID | None = None
    assistant_article_numbers: list[str] | None = None


def format_swiss_number(value: int) -> str:
    return f"{value:,}".replace(",", "\u202f")


def format_snapshot_timestamp(when: datetime) -> str:
    local = when.astimezone(ZURICH)
    return local.strftime("%d.%m.%Y, %H:%M Uhr")


def excel_filename_timestamp(when: datetime) -> str:
    local = when.astimezone(ZURICH)
    return local.strftime("%Y-%m-%d_%H%M")


def running_snapshot(db: Session) -> ArticleSnapshot | None:
    return db.scalars(
        select(ArticleSnapshot).where(ArticleSnapshot.status == "running").limit(1)
    ).first()


def list_snapshots(db: Session, *, tenant: str | None = None) -> list[ArticleSnapshot]:
    tenant = tenant or settings.weclapp_tenant.strip()
    return list(
        db.scalars(
            select(ArticleSnapshot)
            .where(ArticleSnapshot.weclapp_tenant == tenant)
            .order_by(ArticleSnapshot.created_at.desc())
        )
    )


def _base_row_query(snapshot_id: uuid.UUID, filters: SnapshotFilters):
    stmt = select(ArticleSnapshotRow).where(ArticleSnapshotRow.snapshot_id == snapshot_id)
    if filters.nur_aktive:
        stmt = stmt.where(ArticleSnapshotRow.active.is_(True))
    if filters.hauptgruppe:
        stmt = stmt.where(ArticleSnapshotRow.hauptgruppe_code == filters.hauptgruppe)
    if filters.untergruppe:
        stmt = stmt.where(ArticleSnapshotRow.untergruppe_code == filters.untergruppe)
    needle = filters.query.strip().lower()
    if needle:
        pattern = f"%{needle}%"
        stmt = stmt.where(
            or_(
                func.lower(ArticleSnapshotRow.article_number).like(pattern),
                func.lower(ArticleSnapshotRow.article_name).like(pattern),
            )
        )
    if filters.assistant_article_numbers is not None:
        numbers = filters.assistant_article_numbers
        if not numbers:
            stmt = stmt.where(false())
        else:
            stmt = stmt.where(ArticleSnapshotRow.article_number.in_(numbers))
    return stmt


def count_filtered_rows(db: Session, snapshot_id: uuid.UUID, filters: SnapshotFilters) -> int:
    stmt = select(func.count()).select_from(_base_row_query(snapshot_id, filters).subquery())
    return int(db.scalar(stmt) or 0)


def fetch_filtered_rows(
    db: Session,
    snapshot_id: uuid.UUID,
    filters: SnapshotFilters,
) -> tuple[list[ArticleSnapshotRow], int, int]:
    total = count_filtered_rows(db, snapshot_id, filters)
    page = max(1, filters.page)
    start = (page - 1) * GRID_PAGE_SIZE
    stmt = (
        _base_row_query(snapshot_id, filters)
        .order_by(ArticleSnapshotRow.position)
        .offset(start)
        .limit(GRID_PAGE_SIZE)
    )
    rows = list(db.scalars(stmt))
    pages = max(1, (total + GRID_PAGE_SIZE - 1) // GRID_PAGE_SIZE) if total else 1
    return rows, total, pages


def fetch_all_filtered_rows(
    db: Session,
    snapshot_id: uuid.UUID,
    filters: SnapshotFilters,
) -> list[ArticleSnapshotRow]:
    stmt = _base_row_query(snapshot_id, filters).order_by(ArticleSnapshotRow.position)
    return list(db.scalars(stmt))


def distinct_hauptgruppen(db: Session, snapshot_id: uuid.UUID) -> list[str]:
    stmt = (
        select(ArticleSnapshotRow.hauptgruppe_code)
        .where(
            ArticleSnapshotRow.snapshot_id == snapshot_id,
            ArticleSnapshotRow.hauptgruppe_code != "",
        )
        .distinct()
        .order_by(ArticleSnapshotRow.hauptgruppe_code)
    )
    return [row[0] for row in db.execute(stmt).all()]


def distinct_untergruppen(
    db: Session,
    snapshot_id: uuid.UUID,
    *,
    hauptgruppe: str = "",
) -> list[str]:
    stmt = select(ArticleSnapshotRow.untergruppe_code).where(
        ArticleSnapshotRow.snapshot_id == snapshot_id,
        ArticleSnapshotRow.untergruppe_code != "",
    )
    if hauptgruppe:
        stmt = stmt.where(ArticleSnapshotRow.hauptgruppe_code == hauptgruppe)
    stmt = stmt.distinct().order_by(ArticleSnapshotRow.untergruppe_code)
    return [row[0] for row in db.execute(stmt).all()]


def _freeze_column_count(columns: list[dict[str, Any]]) -> int:
    keys = {col.get("key", "") for col in columns}
    count = 0
    for target in (ARTICLE_NUMBER_FIELD, ARTICLE_NAME_FIELD):
        if keys & set(label_variants(target)):
            count += 1
    return max(count, 1)


def build_grid_config(
    snapshot: ArticleSnapshot,
    rows: list[ArticleSnapshotRow],
) -> dict[str, Any]:
    columns = snapshot.columns or []
    jss_columns: list[dict[str, Any]] = []
    keys: list[str] = []
    for col in columns:
        key = str(col.get("key", ""))
        keys.append(key)
        stored_title = str(col.get("title", key) or key)
        jss_columns.append(
            {
                "type": "text",
                "title": snapshot_column_title(key, stored_title),
                "width": int(col.get("width", 140)),
                "readOnly": True,
                "name": key,
            }
        )
    data = [[row.data.get(key, "") if isinstance(row.data, dict) else "" for key in keys] for row in rows]
    return {
        "editable": False,
        "parseFormulas": False,
        "freezeColumns": _freeze_column_count(columns),
        "columns": jss_columns,
        "data": data,
        "fields": keys,
    }


def build_excel_workbook(
    snapshot: ArticleSnapshot,
    rows: list[ArticleSnapshotRow],
    filters: SnapshotFilters,
    *,
    question_de: str | None = None,
) -> Workbook:
    columns = snapshot.columns or []
    keys = [str(col.get("key", "")) for col in columns]
    headers = [
        snapshot_column_title(str(col.get("key", "")), str(col.get("title") or col.get("key") or ""))
        for col in columns
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Artikel"

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, start=2):
        data = row.data if isinstance(row.data, dict) else {}
        for col_idx, key in enumerate(keys, start=1):
            write_cell(ws.cell(row=row_idx, column=col_idx), key, data.get(key, ""))

    if keys:
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(keys))
        ws.auto_filter.ref = f"A1:{last_col}{max(1, len(rows) + 1)}"

    meta = wb.create_sheet("Abfrage")
    meta.append(["Merkmal", "Wert"])
    meta.append(["Stand", format_snapshot_timestamp(snapshot.created_at)])
    meta.append(["Tenant", snapshot.weclapp_tenant])
    meta.append(["Zeilen im Snapshot", snapshot.row_count or 0])
    meta.append(["Zeilen in dieser Datei", len(rows)])
    meta.append(["Suche", filters.query or "(keine)"])
    meta.append(["Hauptgruppe", filters.hauptgruppe or "(alle)"])
    meta.append(["Untergruppe", filters.untergruppe or "(alle)"])
    meta.append(["Nur aktive Artikel", "Ja" if filters.nur_aktive else "Nein"])
    if question_de:
        meta.append(["Frage", question_de])
        meta.append(["Datenstand", format_snapshot_timestamp(snapshot.created_at)])
    return wb


def excel_bytes(
    snapshot: ArticleSnapshot,
    rows: list[ArticleSnapshotRow],
    filters: SnapshotFilters,
    *,
    question_de: str | None = None,
) -> bytes:
    return workbook_bytes(
        build_excel_workbook(snapshot, rows, filters, question_de=question_de)
    )


def apply_retention(db: Session, *, tenant: str) -> list[uuid.UUID]:
    """Delete complete snapshots outside the keep set, oldest first, by id.

    Keep the union of: the newest ``RETENTION_KEEP_RECENT`` complete snapshots
    per tenant, every complete snapshot from the last ``RETENTION_KEEP_DAYS``
    days, and the newest complete snapshot of each UTC calendar month in the
    last ``RETENTION_KEEP_MONTHS`` months. Also delete incomplete (orphan)
    snapshots older than ``RETENTION_ORPHAN_DAYS``. Row deletion is CASCADE
    from the header. Does not commit; the caller must use a transaction
    separate from the pull.
    """
    rows = db.execute(
        text(
            """
            WITH complete AS (
                SELECT id, created_at
                FROM article_snapshots
                WHERE weclapp_tenant = :tenant
                  AND status = 'complete'
            ),
            keep_recent AS (
                SELECT id
                FROM complete
                ORDER BY created_at DESC, id DESC
                LIMIT :keep_recent
            ),
            keep_days AS (
                SELECT id
                FROM complete
                WHERE created_at >= now()
                    - CAST(:keep_days AS integer) * INTERVAL '1 day'
            ),
            keep_monthly AS (
                SELECT DISTINCT ON (
                    date_trunc('month', created_at AT TIME ZONE 'UTC')
                )
                    id
                FROM complete
                WHERE created_at >= (
                    date_trunc('month', now() AT TIME ZONE 'UTC')
                    - (CAST(:keep_months AS integer) - 1) * INTERVAL '1 month'
                ) AT TIME ZONE 'UTC'
                ORDER BY
                    date_trunc('month', created_at AT TIME ZONE 'UTC'),
                    created_at DESC,
                    id DESC
            ),
            keep AS (
                SELECT id FROM keep_recent
                UNION
                SELECT id FROM keep_days
                UNION
                SELECT id FROM keep_monthly
            )
            SELECT complete.id
            FROM complete
            WHERE complete.id NOT IN (SELECT id FROM keep)
            ORDER BY complete.created_at ASC, complete.id ASC
            """
        ),
        {
            "tenant": tenant,
            "keep_recent": RETENTION_KEEP_RECENT,
            "keep_days": RETENTION_KEEP_DAYS,
            "keep_months": RETENTION_KEEP_MONTHS,
        },
    ).all()
    ids = [row[0] if isinstance(row[0], uuid.UUID) else uuid.UUID(str(row[0])) for row in rows]
    if not ids:
        logger.info("snapshot retention removed 0 snapshots, 0 rows")
    else:
        n_rows = int(
            db.scalar(
                select(func.count()).where(ArticleSnapshotRow.snapshot_id.in_(ids))
            )
            or 0
        )
        for snapshot_id in ids:
            db.execute(delete(ArticleSnapshot).where(ArticleSnapshot.id == snapshot_id))
        logger.info(
            "snapshot retention removed %s snapshots, %s rows",
            len(ids),
            n_rows,
        )

    orphan_rows = db.execute(
        text(
            """
            SELECT id
            FROM article_snapshots
            WHERE weclapp_tenant = :tenant
              AND status <> 'complete'
              AND created_at < now()
                  - CAST(:orphan_days AS integer) * INTERVAL '1 day'
            ORDER BY created_at ASC, id ASC
            """
        ),
        {
            "tenant": tenant,
            "orphan_days": RETENTION_ORPHAN_DAYS,
        },
    ).all()
    orphan_ids = [
        row[0] if isinstance(row[0], uuid.UUID) else uuid.UUID(str(row[0]))
        for row in orphan_rows
    ]
    if not orphan_ids:
        logger.info("snapshot retention removed 0 incomplete snapshots, 0 rows")
    else:
        orphan_row_count = int(
            db.scalar(
                select(func.count()).where(
                    ArticleSnapshotRow.snapshot_id.in_(orphan_ids)
                )
            )
            or 0
        )
        for snapshot_id in orphan_ids:
            db.execute(delete(ArticleSnapshot).where(ArticleSnapshot.id == snapshot_id))
        logger.info(
            "snapshot retention removed %s incomplete snapshots, %s rows",
            len(orphan_ids),
            orphan_row_count,
        )

    return ids


def pull_snapshot_rows(
    db: Session,
    snapshot: ArticleSnapshot,
    *,
    oid: str,
) -> dict[str, Any]:
    """Fetch from weclapp (read-only) and persist rows, then retain separately."""
    from app.weclapp import weclapp_client_for

    client = weclapp_client_for(db, oid)
    articles = list(client.iter_pages("article"))
    data_rows, indexed, columns = flatten_articles(client, articles)

    db.execute(
        delete(ArticleSnapshotRow).where(ArticleSnapshotRow.snapshot_id == snapshot.id)
    )
    for position, (data, fields) in enumerate(zip(data_rows, indexed, strict=True)):
        db.add(
            ArticleSnapshotRow(
                snapshot_id=snapshot.id,
                position=position,
                data=data,
                article_number=fields["article_number"],
                article_name=fields["article_name"],
                hauptgruppe_code=fields["hauptgruppe_code"],
                untergruppe_code=fields["untergruppe_code"],
                active=fields["active"],
                weclapp_id=fields["weclapp_id"],
                weclapp_version=fields["weclapp_version"],
            )
        )

    pattern = Scheme().pattern()
    non_conforming = 0
    for fields in indexed:
        number = str(fields.get("article_number") or "").strip()
        if number and pattern.match(number) is None:
            non_conforming += 1

    snapshot.columns = columns
    snapshot.row_count = len(data_rows)
    snapshot.non_conforming_number_count = non_conforming
    snapshot.status = "complete"
    snapshot.error = None
    db.commit()

    deleted: list[uuid.UUID] = []
    try:
        deleted = apply_retention(db, tenant=snapshot.weclapp_tenant)
        db.commit()
    except Exception:
        db.rollback()
        logger.exception(
            "snapshot retention failed tenant=%s snapshot=%s",
            snapshot.weclapp_tenant,
            snapshot.id,
        )
    return {"row_count": len(data_rows), "deleted_snapshots": [str(i) for i in deleted]}


def fail_snapshot(db: Session, snapshot: ArticleSnapshot, message: str) -> None:
    db.execute(
        delete(ArticleSnapshotRow).where(ArticleSnapshotRow.snapshot_id == snapshot.id)
    )
    snapshot.status = "failed"
    snapshot.error = message
    snapshot.row_count = None
    db.commit()


def create_snapshot_pull(
    db: Session,
    user: dict[str, Any],
) -> tuple[ArticleSnapshot, Job] | ArticleSnapshot:
    """Enqueue a pull or return the snapshot that is already running."""
    tenant = settings.weclapp_tenant.strip()
    existing = running_snapshot(db)
    if existing is not None:
        return existing

    from app.jobs import enqueue

    snapshot = ArticleSnapshot(
        status="running",
        created_by_oid=user["oid"],
        created_by_name=user["name"],
        weclapp_tenant=tenant,
    )
    db.add(snapshot)
    db.flush()

    job = enqueue(
        db,
        "weclapp_article_snapshot",
        {"snapshot_id": str(snapshot.id)},
        user,
    )
    snapshot.job_id = job.id
    db.commit()
    db.refresh(snapshot)
    return snapshot, job
