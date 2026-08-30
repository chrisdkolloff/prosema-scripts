"""Parse an uploaded .xlsx or .csv into an article registration batch."""

from __future__ import annotations

import csv
import hashlib
import io
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import UTC, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.article_templates import get_active_template
from app.batches import (
    ARTICLE_NUMBER_FIELD,
    HAUPTGRUPPE_FIELD,
    UNTERGRUPPE_FIELD,
    effective_values,
    group_label,
    resolve_row_groups,
    validate_effective,
)
from app.models import ArticleBatch, ArticleBatchRow, ArticleTemplate
from app.numbering_high_water import assign_proposed_numbers, seed_high_water
from core.article_fields import IMPORT_COLUMNS, find_field, normalize_label
from core.article_payload import DEFAULTS, get_row_value, label_variants, label_variants

MAX_UPLOAD_ROWS = 2000
MAX_MANUAL_ROWS = 200
DEFAULT_MANUAL_ROWS = 20
MAX_BATCH_ROWS = 2000

MSG_UPLOAD_FORMAT = "Nur .xlsx- und .csv-Dateien werden akzeptiert."
MSG_CSV_READ = "CSV-Datei konnte nicht gelesen werden: {detail}"
MSG_DUPLICATE_HEADERS = "Doppelte Spaltenüberschriften sind nicht erlaubt: {headers}."
MSG_MISSING_HEADERS = "Pflichtspalten fehlen: {headers}."
MSG_TOO_MANY_ROWS = "Maximal 2000 Zeilen pro Batch. Diese Datei enthält {n}."
MSG_EMPTY_FILE = "Die Datei enthält keine Datenzeilen."
MSG_MANUAL_ROW_RANGE = "Zeilenanzahl muss zwischen 1 und {max} liegen."
MSG_BATCH_CAP = "Maximal {max} Zeilen pro Batch."
MSG_NOT_DRAFT = "Nur Entwürfe können ergänzt werden."


class BatchUploadError(Exception):
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


@dataclass
class DuplicateUpload:
    batch: ArticleBatch
    created_at: datetime


@dataclass
class UploadResult:
    batch: ArticleBatch
    blank_skipped: int = 0
    unknown_columns: list[str] = field(default_factory=list)
    notices: list[str] = field(default_factory=list)


def _cell_as_text(value: object) -> str:
    """Read cell values as text; never let codes become floats."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Ja" if value else "Nein"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, "f").rstrip("0").rstrip(".")
        return text
    return str(value).strip()


def _is_blank_row(values: dict[str, str]) -> bool:
    return not any(str(v or "").strip() for v in values.values())


def _template_required_from_columns(columns: list) -> list[str]:
    return [
        str(col.get("label") or "")
        for col in columns
        if col.get("required") and col.get("label")
    ]


def _template_optional_from_columns(columns: list) -> list[str]:
    return [
        str(col.get("label") or "")
        for col in columns
        if not col.get("required") and col.get("label")
    ]


def _folds_for_label(label: str) -> set[str]:
    return {normalize_label(name).casefold() for name in label_variants(label) if name}


def _parse_upload_table(
    headers: list[str],
    body_rows: Iterable[Sequence[object]],
    *,
    template: ArticleTemplate,
) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """Return (rows as raw_data dicts, unknown headers, missing optional labels)."""
    columns = template.columns if isinstance(template.columns, list) else []
    label_to_key: dict[str, str] = {}
    for col in columns:
        label = normalize_label(col.get("label"))
        key = str(col.get("key") or label)
        if not label:
            continue
        for fold in _folds_for_label(label):
            label_to_key.setdefault(fold, key)
        field = find_field(label)
        if field is not None:
            for fold in _folds_for_label(field.label):
                label_to_key.setdefault(fold, key)

    required_labels = _template_required_from_columns(columns)
    optional_labels = _template_optional_from_columns(columns)

    seen_fold: dict[str, int] = {}
    duplicates: list[str] = []
    by_index: dict[int, str] = {}  # col index -> raw_data key
    present_fold: set[str] = set()
    unknown: list[str] = []

    for index, name in enumerate(headers):
        if not name:
            continue
        fold = name.casefold()
        if fold in seen_fold:
            duplicates.append(name)
            continue
        seen_fold[fold] = index
        present_fold.update(_folds_for_label(name))
        canonical = label_to_key.get(fold)
        if canonical is not None:
            by_index[index] = canonical
        else:
            by_index[index] = name
            unknown.append(name)

    if duplicates:
        uniq = sorted(set(duplicates))
        raise BatchUploadError(MSG_DUPLICATE_HEADERS.format(headers=", ".join(uniq)))

    missing_required = [
        label
        for label in required_labels
        if normalize_label(label).casefold() not in present_fold
        and not (_folds_for_label(label) & present_fold)
    ]
    if missing_required:
        raise BatchUploadError(
            MSG_MISSING_HEADERS.format(headers=", ".join(missing_required))
        )

    missing_optional = [
        label
        for label in optional_labels
        if normalize_label(label).casefold() not in present_fold
        and not (_folds_for_label(label) & present_fold)
    ]

    parsed: list[dict[str, str]] = []
    number_keys = set(label_variants(ARTICLE_NUMBER_FIELD))
    for values in body_rows:
        raw: dict[str, str] = {}
        for index, key in by_index.items():
            cell = values[index] if index < len(values) else None
            raw[key] = _cell_as_text(cell)
        if _is_blank_row(raw):
            parsed.append({})
            continue
        # Never keep a typed article number from the file — numbers are derived.
        for number_key in number_keys:
            raw.pop(number_key, None)
        parsed.append(raw)
    return parsed, unknown, missing_optional


def parse_workbook_bytes(
    data: bytes,
    *,
    template: ArticleTemplate,
) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """Return (rows as raw_data dicts, unknown headers, missing optional labels).

    Always reads worksheet index 0 — never ``wb.active`` and never the first
    non-empty sheet — so a ``Beispiel`` sheet cannot be mistaken for data.
    """
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:
        raise BatchUploadError(f"Excel-Datei konnte nicht gelesen werden: {exc}") from exc
    try:
        if not wb.worksheets:
            raise BatchUploadError(MSG_EMPTY_FILE)
        sheet = wb.worksheets[0]
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise BatchUploadError(MSG_EMPTY_FILE) from exc

        headers = [normalize_label(h) for h in header_row]
        return _parse_upload_table(headers, rows_iter, template=template)
    finally:
        wb.close()


def _detect_csv_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        if ";" in sample.splitlines()[0]:
            return ";"
        return ","


def parse_csv_bytes(
    data: bytes,
    *,
    template: ArticleTemplate,
) -> tuple[list[dict[str, str]], list[str], list[str]]:
    """Return (rows as raw_data dicts, unknown headers, missing optional labels)."""
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise BatchUploadError(MSG_CSV_READ.format(detail=exc)) from exc

    delimiter = _detect_csv_delimiter(text)
    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        try:
            header_row = next(reader)
        except StopIteration as exc:
            raise BatchUploadError(MSG_EMPTY_FILE) from exc
        headers = [normalize_label(h) for h in header_row]
        body_rows = list(reader)
    except csv.Error as exc:
        raise BatchUploadError(MSG_CSV_READ.format(detail=exc)) from exc

    return _parse_upload_table(headers, body_rows, template=template)


def _upload_format(filename: str) -> str | None:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return "xlsx"
    if lower.endswith(".csv"):
        return "csv"
    return None


def find_duplicate_batch(db: Session, sha256: str) -> ArticleBatch | None:
    return db.scalars(
        select(ArticleBatch)
        .where(
            ArticleBatch.source_sha256 == sha256,
            ArticleBatch.status != "discarded",
        )
        .order_by(ArticleBatch.created_at.desc())
        .limit(1)
    ).first()


def _prepare_row(
    db: Session,
    *,
    batch: ArticleBatch,
    position: int,
    raw: dict[str, str],
) -> ArticleBatchRow:
    row = ArticleBatchRow(
        batch_id=batch.id,
        position=position,
        raw_data=raw,
        edits={},
        proposed_article_number="",
        include=True,
        validation_error="",
    )
    values = {col: get_row_value(raw, col) for col in IMPORT_COLUMNS}
    for col, default in DEFAULTS.items():
        if not values.get(col):
            values[col] = default
    haupt, unter, group_error = resolve_row_groups(db, values)
    row.resolved_hauptgruppe_id = haupt.id if haupt is not None else None
    row.resolved_untergruppe_id = unter.id if unter is not None else None
    edits: dict[str, str] = {}
    if haupt is not None:
        label = group_label(haupt.name, haupt.code)
        if values.get(HAUPTGRUPPE_FIELD, "") != label:
            edits[HAUPTGRUPPE_FIELD] = label
    if unter is not None:
        label = group_label(unter.name, unter.code)
        if values.get(UNTERGRUPPE_FIELD, "") != label:
            edits[UNTERGRUPPE_FIELD] = label
    if edits:
        row.edits = edits
    row._resolved_haupt = haupt
    row._resolved_unter = unter
    row._group_error = group_error
    return row


def create_batch_from_upload(
    db: Session,
    *,
    filename: str,
    data: bytes,
    user: Mapping[str, Any],
    confirmed: bool = False,
) -> UploadResult | DuplicateUpload:
    upload_format = _upload_format(filename)
    if upload_format is None:
        raise BatchUploadError(MSG_UPLOAD_FORMAT)

    digest = hashlib.sha256(data).hexdigest()
    existing = find_duplicate_batch(db, digest)
    if existing is not None and not confirmed:
        return DuplicateUpload(batch=existing, created_at=existing.created_at)

    template = get_active_template(db)
    if upload_format == "csv":
        rows_raw, unknown, missing_optional = parse_csv_bytes(data, template=template)
    else:
        rows_raw, unknown, missing_optional = parse_workbook_bytes(data, template=template)
    blank_skipped = sum(1 for row in rows_raw if not row)
    data_rows = [row for row in rows_raw if row]
    if not data_rows:
        raise BatchUploadError(MSG_EMPTY_FILE)
    if len(data_rows) > MAX_UPLOAD_ROWS:
        raise BatchUploadError(MSG_TOO_MANY_ROWS.format(n=len(data_rows)))

    batch = ArticleBatch(
        status="draft",
        filename=filename,
        source_bytes=data,
        source_sha256=digest,
        template_id=template.id,
        created_by_oid=str(user["oid"]),
        created_by_name=str(user.get("name") or user["oid"]),
    )
    db.add(batch)
    db.flush()

    reserved = seed_high_water(db, exclude_batch_id=None)
    created_rows: list[ArticleBatchRow] = []
    for position, raw in enumerate(data_rows, start=1):
        row = _prepare_row(db, batch=batch, position=position, raw=raw)
        created_rows.append(row)
        db.add(row)

    db.flush()
    assign_proposed_numbers(created_rows, reserved)

    for row in created_rows:
        values = effective_values(row)
        row.validation_error = validate_effective(
            values, getattr(row, "_group_error", None)
        )

    notices: list[str] = []
    if blank_skipped:
        notices.append(f"{blank_skipped} leere Zeilen übersprungen.")
    if missing_optional:
        notices.append(
            f"{len(missing_optional)} optionale Spalten fehlen: "
            f"{', '.join(missing_optional)}."
        )
    if unknown:
        notices.append(
            f"{len(unknown)} unbekannte Spalten wurden übernommen, aber nicht geprüft: "
            f"{', '.join(unknown)}."
        )

    batch.updated_at = datetime.now(UTC)
    return UploadResult(
        batch=batch,
        blank_skipped=blank_skipped,
        unknown_columns=unknown,
        notices=notices,
    )


def create_manual_batch(
    db: Session,
    *,
    user: Mapping[str, Any],
    row_count: int = DEFAULT_MANUAL_ROWS,
) -> ArticleBatch:
    if row_count < 1 or row_count > MAX_MANUAL_ROWS:
        raise BatchUploadError(MSG_MANUAL_ROW_RANGE.format(max=MAX_MANUAL_ROWS))

    template = get_active_template(db)
    batch = ArticleBatch(
        status="draft",
        filename=None,
        source_bytes=None,
        source_sha256=None,
        template_id=template.id,
        created_by_oid=str(user["oid"]),
        created_by_name=str(user.get("name") or user["oid"]),
    )
    db.add(batch)
    db.flush()

    created_rows: list[ArticleBatchRow] = []
    for position in range(1, row_count + 1):
        row = ArticleBatchRow(
            batch_id=batch.id,
            position=position,
            raw_data={},
            edits={},
            proposed_article_number="",
            include=True,
            validation_error="",
        )
        values = {col: DEFAULTS.get(col, "") for col in IMPORT_COLUMNS}
        haupt, unter, group_error = resolve_row_groups(db, values)
        row.resolved_hauptgruppe_id = haupt.id if haupt is not None else None
        row.resolved_untergruppe_id = unter.id if unter is not None else None
        row._resolved_haupt = haupt
        row._resolved_unter = unter
        row._group_error = group_error
        row.validation_error = validate_effective(effective_values(row), group_error)
        created_rows.append(row)
        db.add(row)

    batch.updated_at = datetime.now(UTC)
    return batch


def append_empty_rows(
    db: Session,
    batch: ArticleBatch,
    *,
    count: int = DEFAULT_MANUAL_ROWS,
) -> int:
    if batch.status != "draft":
        raise BatchUploadError(MSG_NOT_DRAFT)
    if count < 1 or count > MAX_MANUAL_ROWS:
        raise BatchUploadError(MSG_MANUAL_ROW_RANGE.format(max=MAX_MANUAL_ROWS))

    existing = list(
        db.scalars(
            select(ArticleBatchRow)
            .where(ArticleBatchRow.batch_id == batch.id)
            .order_by(ArticleBatchRow.position)
        )
    )
    if len(existing) + count > MAX_BATCH_ROWS:
        raise BatchUploadError(MSG_BATCH_CAP.format(max=MAX_BATCH_ROWS))

    start = (existing[-1].position + 1) if existing else 1
    for offset in range(count):
        position = start + offset
        row = ArticleBatchRow(
            batch_id=batch.id,
            position=position,
            raw_data={},
            edits={},
            proposed_article_number="",
            include=True,
            validation_error="",
        )
        values = {col: DEFAULTS.get(col, "") for col in IMPORT_COLUMNS}
        haupt, unter, group_error = resolve_row_groups(db, values)
        row.resolved_hauptgruppe_id = haupt.id if haupt is not None else None
        row.resolved_untergruppe_id = unter.id if unter is not None else None
        row.validation_error = validate_effective(
            {**values, ARTICLE_NUMBER_FIELD: ""}, group_error
        )
        db.add(row)

    batch.updated_at = datetime.now(UTC)
    return count


def _row_has_user_content(row: ArticleBatchRow) -> bool:
    """True when raw_data or edits carry any non-blank value."""
    raw = row.raw_data if isinstance(row.raw_data, dict) else {}
    edits = row.edits if isinstance(row.edits, dict) else {}
    if any(str(v or "").strip() for v in raw.values()):
        return True
    if any(str(v or "").strip() for v in edits.values()):
        return True
    return bool((row.proposed_article_number or "").strip())


def exclude_empty_rows(db: Session, batch: ArticleBatch) -> int:
    """Set include=False on draft rows with no user-entered content. Does not delete."""
    if batch.status != "draft":
        raise BatchUploadError(MSG_NOT_DRAFT)

    rows = list(
        db.scalars(
            select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
        )
    )
    excluded = 0
    for row in rows:
        if not row.include:
            continue
        if _row_has_user_content(row):
            continue
        row.include = False
        excluded += 1

    batch.updated_at = datetime.now(UTC)
    return excluded
