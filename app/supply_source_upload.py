"""Parse Bezugsquellen uploads against a supplier template.

Fail loudly: collect every file-level and row-level problem, then reject.
Comma-decimal and dot-decimal Listenpreis each have their own cast.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import WeclappUnit
from app.supply_source_templates import KEY_BY_LABEL, LABEL_BY_KEY

MSG_EMPTY_FILE = "Die Datei ist leer. Mindestens eine Datenzeile ist nötig."
MSG_HEADER_ONLY = (
    "Die Datei enthält nur die Kopfzeile. Mindestens eine Datenzeile eintragen."
)
MSG_UNREADABLE_XLSX = "Excel-Datei konnte nicht gelesen werden: {detail}"
MSG_UNREADABLE_CSV = "CSV-Datei konnte nicht gelesen werden: {detail}"
MSG_UNSUPPORTED = "Nur .xlsx- oder .csv-Dateien werden akzeptiert."


class SupplySourceParseError(Exception):
    """File-level rejection. ``messages`` lists every problem found."""

    def __init__(self, messages: Sequence[str]):
        self.messages = [str(m) for m in messages if str(m).strip()]
        super().__init__("\n".join(self.messages))


@dataclass
class ParsedUploadRow:
    excel_row: int
    supplier_article_number: str
    name: str | None
    listenpreis: Decimal
    ean: str | None
    unit_id: str | None
    unit_raw: str | None
    rabattcode: str | None
    min_purchase_qty: Decimal | None
    procurement_lead_days: int | None


@dataclass
class ParseResult:
    rows: list[ParsedUploadRow] = field(default_factory=list)
    row_errors: list[str] = field(default_factory=list)
    unmatched_units: list[dict[str, str]] = field(default_factory=list)


def parse_listenpreis(raw: object) -> Decimal:
    """Parse Swiss/EU Listenpreis. Comma decimal and dot decimal are separate casts."""
    if raw is None:
        raise ValueError("Listenpreis fehlt.")
    if isinstance(raw, (int, float, Decimal)) and not isinstance(raw, bool):
        value = Decimal(str(raw))
        if value < 0:
            raise ValueError("Listenpreis darf nicht negativ sein.")
        return value
    compact = (
        str(raw)
        .strip()
        .replace(" ", "")
        .replace("\u00a0", "")
        .replace("'", "")
        .replace("’", "")
    )
    if compact == "":
        raise ValueError("Listenpreis fehlt.")
    has_comma = "," in compact
    has_dot = "." in compact
    if has_comma and has_dot:
        if compact.rindex(",") > compact.rindex("."):
            try:
                value = Decimal(compact.replace(".", "").replace(",", "."))
            except InvalidOperation as exc:
                raise ValueError(
                    "Listenpreis mit Komma-Dezimalstelle ist keine Zahl."
                ) from exc
        else:
            try:
                value = Decimal(compact.replace(",", ""))
            except InvalidOperation as exc:
                raise ValueError(
                    "Listenpreis mit Punkt-Dezimalstelle ist keine Zahl."
                ) from exc
    elif has_comma:
        try:
            value = Decimal(compact.replace(".", "").replace(",", "."))
        except InvalidOperation as exc:
            raise ValueError(
                "Listenpreis mit Komma-Dezimalstelle ist keine Zahl."
            ) from exc
    elif has_dot:
        try:
            value = Decimal(compact)
        except InvalidOperation as extra:
            raise ValueError(
                "Listenpreis mit Punkt-Dezimalstelle ist keine Zahl."
            ) from extra
    else:
        try:
            value = Decimal(compact)
        except InvalidOperation as exc:
            raise ValueError("Listenpreis ist keine Zahl.") from exc
    if value < 0:
        raise ValueError("Listenpreis darf nicht negativ sein.")
    return value


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _blank_row(values: dict[str, str]) -> bool:
    return all(not (v or "").strip() for v in values.values())


def _detect_csv_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        if sample.splitlines() and ";" in sample.splitlines()[0]:
            return ";"
        return ","


def _header_map(headers: Sequence[object], columns: Sequence[Mapping[str, Any]]) -> dict[str, int]:
    wanted = {
        str(col.get("label") or "").strip(): str(col.get("key") or "")
        for col in columns
        if col.get("label") and col.get("key")
    }
    found: dict[str, int] = {}
    for index, raw in enumerate(headers):
        label = _cell_text(raw)
        key = wanted.get(label) or KEY_BY_LABEL.get(label)
        if key and key not in found:
            found[key] = index
    return found


def _read_table(headers: Sequence[object], body: Iterable[Sequence[object]]) -> tuple[list[str], list[list[object]]]:
    return [_cell_text(h) for h in headers], [list(row) for row in body]


def parse_upload_bytes(
    db: Session,
    data: bytes,
    *,
    filename: str,
    columns: Sequence[Mapping[str, Any]],
) -> ParseResult:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx"):
        headers, body = _read_xlsx(data)
    elif lower.endswith(".csv"):
        headers, body = _read_csv(data)
    else:
        raise SupplySourceParseError([MSG_UNSUPPORTED])

    file_errors: list[str] = []
    if not any(headers):
        raise SupplySourceParseError([MSG_EMPTY_FILE])

    required = [
        str(col.get("key"))
        for col in columns
        if col.get("required") and col.get("key")
    ]
    found = _header_map(headers, columns)
    missing = [LABEL_BY_KEY.get(k, k) for k in required if k not in found]
    if missing:
        file_errors.append(
            "Pflichtspalten fehlen: "
            + ", ".join(missing)
            + ". Kopfzeile an die aktive Vorlage anpassen."
        )

    data_rows = [row for row in body if any(_cell_text(c) for c in row)]
    if not data_rows:
        file_errors.append(MSG_HEADER_ONLY)

    if file_errors:
        raise SupplySourceParseError(file_errors)

    units_by_name = {
        (u.name or ""): u.weclapp_id
        for u in db.scalars(select(WeclappUnit)).all()
        if u.name
    }

    result = ParseResult()
    sans: dict[str, list[int]] = {}
    for offset, raw_row in enumerate(data_rows, start=2):
        cells = {
            key: _cell_text(raw_row[idx] if idx < len(raw_row) else None)
            for key, idx in found.items()
        }
        if _blank_row(cells):
            continue
        excel_row = offset
        san = cells.get("supplier_article_number", "").strip()
        if not san:
            result.row_errors.append(
                f"Zeile {excel_row}: Lieferantenartikelnummer fehlt. "
                "Nummer eintragen oder die Zeile löschen."
            )
            continue
        sans.setdefault(san, []).append(excel_row)
        listen_raw = cells.get("listenpreis", "")
        try:
            listenpreis = parse_listenpreis(listen_raw)
        except ValueError as exc:
            result.row_errors.append(
                f"Zeile {excel_row} ({san}): {exc} "
                "Komma- oder Punkt-Dezimalstelle verwenden."
            )
            continue
        unit_raw = cells.get("unit") or None
        unit_id = None
        if unit_raw:
            unit_id = units_by_name.get(unit_raw)
            if unit_id is None:
                result.unmatched_units.append(
                    {"row": str(excel_row), "san": san, "value": unit_raw}
                )
        min_qty = None
        min_raw = cells.get("min_purchase_qty") or ""
        if min_raw:
            try:
                min_qty = parse_listenpreis(min_raw)
            except ValueError:
                result.row_errors.append(
                    f"Zeile {excel_row} ({san}): Mindestbestellmenge ist keine Zahl."
                )
                continue
        lead = None
        lead_raw = cells.get("procurement_lead_days") or ""
        if lead_raw:
            try:
                lead = int(Decimal(str(lead_raw).replace(",", ".")))
            except (InvalidOperation, ValueError):
                result.row_errors.append(
                    f"Zeile {excel_row} ({san}): Lieferzeit (Tage) ist keine ganze Zahl."
                )
                continue
        ean = cells.get("ean") or None
        name = cells.get("name") or None
        rabattcode = cells.get("rabattcode") or None
        result.rows.append(
            ParsedUploadRow(
                excel_row=excel_row,
                supplier_article_number=san,
                name=name,
                listenpreis=listenpreis,
                ean=ean,
                unit_id=unit_id,
                unit_raw=unit_raw,
                rabattcode=rabattcode,
                min_purchase_qty=min_qty,
                procurement_lead_days=lead,
            )
        )

    duplicates = {san: rows for san, rows in sans.items() if len(rows) > 1}
    if duplicates:
        parts = [
            f"{san} (Zeilen {', '.join(str(n) for n in nums)})"
            for san, nums in sorted(duplicates.items())
        ]
        raise SupplySourceParseError(
            [
                "Doppelte Lieferantenartikelnummer in der Datei: "
                + "; ".join(parts)
                + ". Jede Nummer darf nur einmal vorkommen."
            ]
        )
    if not result.rows:
        messages = list(result.row_errors) or [MSG_HEADER_ONLY]
        raise SupplySourceParseError(messages)
    return result


def _read_xlsx(data: bytes) -> tuple[list[str], list[list[object]]]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise SupplySourceParseError([MSG_UNREADABLE_XLSX.format(detail=exc)]) from exc
    try:
        if not wb.worksheets:
            raise SupplySourceParseError([MSG_EMPTY_FILE])
        sheet = wb.worksheets[0]
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise SupplySourceParseError([MSG_EMPTY_FILE]) from exc
        return _read_table(header_row, rows_iter)
    finally:
        wb.close()


def _read_csv(data: bytes) -> tuple[list[str], list[list[object]]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise SupplySourceParseError([MSG_UNREADABLE_CSV.format(detail=exc)]) from exc
    delimiter = _detect_csv_delimiter(text)
    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        try:
            header_row = next(reader)
        except StopIteration as exc:
            raise SupplySourceParseError([MSG_EMPTY_FILE]) from exc
        return _read_table(header_row, list(reader))
    except csv.Error as exc:
        raise SupplySourceParseError([MSG_UNREADABLE_CSV.format(detail=exc)]) from exc
