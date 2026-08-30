"""Shared Excel writer helpers for Artikelübersicht and batch downloads.

Code-shaped columns are always forced to text format ``@`` so Excel does not
mangle values like ``010.020.0010`` into floats.
"""

from __future__ import annotations

import io
import re
from collections.abc import Iterable, Mapping, Sequence
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

TEXT_EXCEL_COLUMNS = frozenset(
    {
        "Prosema Artikelnummer",
        "Prosema-Artikelnummer",
        "Artikelnummer",
        "Hauptgruppe",
        "Untergruppe",
        "GTIN (EAN-Nummer)",
        "Lieferantenartikelnummer",
        "weclapp-ID",
    }
)

_PRICE_RE = re.compile(r"preis|€", re.IGNORECASE)


def is_price_column(key: str) -> bool:
    return bool(_PRICE_RE.search(key))


def parse_price(value: object) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text.replace(" ", "").replace("'", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def write_cell(cell: Any, key: str, raw: object, *, text_columns: frozenset[str] | None = None) -> None:
    columns = text_columns if text_columns is not None else TEXT_EXCEL_COLUMNS
    if key in columns:
        cell.value = str(raw) if raw is not None else ""
        cell.number_format = "@"
        return
    if is_price_column(key):
        number = parse_price(raw)
        if number is not None:
            cell.value = number
            cell.number_format = "#,##0.00"
        else:
            cell.value = str(raw) if raw else None
        return
    cell.value = str(raw) if raw is not None and raw != "" else None


def write_header_row(ws: Worksheet, headers: Sequence[str]) -> None:
    for col_idx, key in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=key)


def write_data_rows(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Iterable[Mapping[str, object]],
    *,
    text_columns: frozenset[str] | None = None,
    start_row: int = 2,
) -> int:
    row_idx = start_row
    for row in rows:
        for col_idx, key in enumerate(headers, start=1):
            write_cell(
                ws.cell(row=row_idx, column=col_idx),
                key,
                row.get(key, ""),
                text_columns=text_columns,
            )
        row_idx += 1
    return row_idx - start_row


def freeze_and_filter(ws: Worksheet, headers: Sequence[str], row_count: int) -> None:
    if not headers:
        return
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, row_count + 1)}"


def workbook_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_template_workbook(
    headers: Sequence[str],
    *,
    examples: Mapping[str, object] | None = None,
) -> bytes:
    """Header-only sheet ``Vorlage``; optional examples on sheet ``Beispiel``.

    Article uploads must never see example rows on sheet 1 — users paste under
    them and create junk articles. Template-replacement uploads ignore data
    rows on purpose; article uploads do not.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vorlage"
    write_header_row(ws, headers)
    freeze_and_filter(ws, headers, row_count=0)
    if examples is not None:
        beispiel = wb.create_sheet("Beispiel")
        write_header_row(beispiel, headers)
        write_data_rows(beispiel, headers, [examples])
        freeze_and_filter(beispiel, headers, row_count=1)
    return workbook_bytes(wb)
