"""Approve, discard, and Excel-download helpers for article batches."""

from __future__ import annotations

from collections.abc import Mapping
from datetime import UTC, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.audit import record_audit_log
from app.batches import (
    ARTICLE_NUMBER_FIELD,
    effective_values,
    load_batch_rows,
)
from app.excel_export import TEXT_EXCEL_COLUMNS, workbook_bytes, write_cell
from app.models import ArticleBatch, ArticleBatchRow, ArticleSnapshot
from app.numbering_high_water import latest_completed_snapshot
from app.snapshots import format_snapshot_timestamp, running_snapshot
from core.article_fields import IMPORT_COLUMNS
from core.article_payload import row_to_payload
from scripts.weclapp.article_import import LookupTables, _load_schema

MSG_NOT_DRAFT = "Nur Entwürfe können freigegeben werden."
MSG_HAS_ERRORS = "Freigabe nicht möglich: {n} Zeilen mit Fehlern"
MSG_NO_SNAPSHOT = (
    "Es existiert noch keine Artikelübersicht. Bitte zuerst eine Abfrage starten."
)
MSG_NULL_OR_DUP_NUMBER = (
    "Freigabe nicht möglich: Artikelnummern fehlen oder sind doppelt."
)
MSG_NOT_DRAFT_DISCARD = "Nur Entwürfe können verworfen werden."
MSG_SNAPSHOT_STALE = (
    "Die Artikelübersicht ist {n} Stunden alt. Nummern könnten mit neueren Artikeln kollidieren."
)

STATUS_LABELS = {
    "draft": "Entwurf",
    "approved": "Freigegeben",
    "submitting": "Wird gesendet",
    "submitted": "Gesendet",
    "discarded": "Verworfen",
}


class BatchActionError(Exception):
    def __init__(self, message: str, *, status_code: int = 400):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def _lookups() -> LookupTables:
    return LookupTables(_load_schema())


def batch_counts(rows: list[ArticleBatchRow]) -> dict[str, int]:
    included = [row for row in rows if row.include]
    return {
        "row_count": len(rows),
        "error_count": sum(
            1 for row in included if (row.validation_error or "").strip()
        ),
        "written_count": sum(1 for row in included if row.weclapp_article_id),
        "include_count": len(included),
    }


def snapshot_age_warning(snapshot: ArticleSnapshot | None) -> str | None:
    if snapshot is None:
        return None
    age = datetime.now(UTC) - snapshot.created_at.astimezone(UTC)
    hours = int(age.total_seconds() // 3600)
    if age >= timedelta(hours=24):
        return MSG_SNAPSHOT_STALE.format(n=max(hours, 24))
    return None


def snapshot_banner_state(db: Session, *, after_refresh: bool = False) -> dict[str, Any]:
    """Action-bar flags for stale warning, in-flight pull, and just-finished success."""
    snapshot = latest_completed_snapshot(db)
    running = running_snapshot(db) is not None
    warning = snapshot_age_warning(snapshot)
    succeeded = bool(after_refresh and not running and warning is None and snapshot is not None)
    error = ""
    if after_refresh and not running and not succeeded:
        failed = db.scalars(
            select(ArticleSnapshot)
            .where(ArticleSnapshot.status == "failed")
            .order_by(ArticleSnapshot.created_at.desc())
            .limit(1)
        ).first()
        if failed is not None and (
            snapshot is None or failed.created_at >= snapshot.created_at
        ):
            error = failed.error or "Aktualisierung fehlgeschlagen."
    return {
        "snapshot": snapshot,
        "snapshot_ts": format_snapshot_timestamp(snapshot.created_at) if snapshot else "",
        "snapshot_warning": warning,
        "snapshot_refresh_running": running,
        "snapshot_refresh_succeeded": succeeded,
        "snapshot_refresh_error": error,
    }


def approve_batch(
    db: Session,
    batch: ArticleBatch,
    *,
    actor: Mapping[str, Any],
) -> ArticleBatch:
    locked = db.scalars(
        select(ArticleBatch).where(ArticleBatch.id == batch.id).with_for_update()
    ).first()
    if locked is None:
        raise BatchActionError("Stapel nicht gefunden", status_code=404)
    if locked.status != "draft":
        raise BatchActionError(MSG_NOT_DRAFT)

    snapshot = latest_completed_snapshot(db)
    if snapshot is None:
        raise BatchActionError(MSG_NO_SNAPSHOT)

    rows = load_batch_rows(db, locked.id)
    included = [row for row in rows if row.include]
    error_count = sum(1 for row in included if (row.validation_error or "").strip())
    if error_count:
        raise BatchActionError(MSG_HAS_ERRORS.format(n=error_count))

    numbers = [(row.proposed_article_number or "").strip() for row in included]
    if any(not number for number in numbers) or len(numbers) != len(set(numbers)):
        raise BatchActionError(MSG_NULL_OR_DUP_NUMBER)

    lookups = _lookups()
    for row in included:
        values = effective_values(row)
        payload = row_to_payload(values, lookups)
        row.approved_payload = payload

    now = datetime.now(UTC)
    locked.status = "approved"
    locked.approved_at = now
    locked.approved_by_oid = str(actor["oid"])
    locked.approved_by_name = str(actor.get("name") or actor["oid"])
    locked.updated_at = now

    record_audit_log(
        db,
        actor=actor,
        entity_type="article_batch",
        entity_id=locked.id,
        action="approved",
        detail={
            "row_count": len(included),
            "article_numbers": numbers,
            "snapshot_id": str(snapshot.id),
            "snapshot_created_at": snapshot.created_at.isoformat(),
        },
    )
    return locked


def discard_batch(
    db: Session,
    batch: ArticleBatch,
    *,
    actor: Mapping[str, Any],
) -> ArticleBatch:
    locked = db.scalars(
        select(ArticleBatch).where(ArticleBatch.id == batch.id).with_for_update()
    ).first()
    if locked is None:
        raise BatchActionError("Stapel nicht gefunden", status_code=404)
    if locked.status != "draft":
        raise BatchActionError(MSG_NOT_DRAFT_DISCARD)
    locked.status = "discarded"
    locked.updated_at = datetime.now(UTC)
    record_audit_log(
        db,
        actor=actor,
        entity_type="article_batch",
        entity_id=locked.id,
        action="discarded",
        detail={},
    )
    return locked


def build_batch_excel(db: Session, batch: ArticleBatch) -> bytes:
    rows = load_batch_rows(db, batch.id)
    snapshot = latest_completed_snapshot(db)
    headers = [
        *IMPORT_COLUMNS,
        "Artikelnummer",
        "Validierungsfehler",
        "weclapp-ID",
        "Schreibfehler",
    ]
    text_columns = TEXT_EXCEL_COLUMNS | {
        "Artikelnummer",
        "weclapp-ID",
        ARTICLE_NUMBER_FIELD,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Artikel"
    for col_idx, key in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=key)

    for row_idx, row in enumerate(rows, start=2):
        values = effective_values(row)
        data = {key: values.get(key, "") for key in IMPORT_COLUMNS}
        data["Artikelnummer"] = row.proposed_article_number or ""
        data["Validierungsfehler"] = row.validation_error or ""
        data["weclapp-ID"] = row.weclapp_article_id or ""
        data["Schreibfehler"] = row.write_error or ""
        for col_idx, key in enumerate(headers, start=1):
            write_cell(
                ws.cell(row=row_idx, column=col_idx),
                key,
                data.get(key, ""),
                text_columns=text_columns,
            )

    meta = wb.create_sheet("Batch")
    meta.append(["Merkmal", "Wert"])
    meta.append(["Batch-ID", str(batch.id)])
    meta.append(["Dateiname", batch.filename or ""])
    meta.append(
        [
            "Hochgeladen",
            format_snapshot_timestamp(batch.created_at) if batch.created_at else "",
        ]
    )
    meta.append(["Hochgeladen von", batch.created_by_name or ""])
    meta.append(
        [
            "Freigegeben",
            format_snapshot_timestamp(batch.approved_at) if batch.approved_at else "",
        ]
    )
    meta.append(["Freigegeben von", batch.approved_by_name or ""])
    meta.append(
        [
            "Gesendet",
            format_snapshot_timestamp(batch.submitted_at) if batch.submitted_at else "",
        ]
    )
    meta.append(["Gesendet von", batch.submitted_by_name or ""])
    counts = batch_counts(rows)
    meta.append(["Zeilen", counts["row_count"]])
    meta.append(["Fehler", counts["error_count"]])
    meta.append(["Geschrieben", counts["written_count"]])
    meta.append(
        [
            "Stand der Artikelübersicht",
            format_snapshot_timestamp(snapshot.created_at) if snapshot else "(keine)",
        ]
    )
    if snapshot is not None:
        meta.append(
            [
                "Nicht-konforme Nummern in Übersicht",
                snapshot.non_conforming_number_count,
            ]
        )
    return workbook_bytes(wb)


def approval_dialog_context(batch: ArticleBatch, rows: list[ArticleBatchRow]) -> dict[str, Any]:
    included = [row for row in rows if row.include]
    numbers = sorted(
        (row.proposed_article_number or "") for row in included if row.proposed_article_number
    )
    return {
        "approve_count": len(included),
        "first_number": numbers[0] if numbers else "",
        "last_number": numbers[-1] if numbers else "",
    }
