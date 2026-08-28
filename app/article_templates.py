"""Versioned article upload templates: validate, activate, download."""

from __future__ import annotations

import hashlib
import io
import uuid
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.audit import record_audit_log
from app.excel_export import build_template_workbook
from app.models import ArticleBatch, ArticleTemplate
from core.article_fields import (
    FIELDS,
    PROTECTED_FIELDS,
    find_field,
    normalize_label,
    seed_template_columns,
)

MSG_ALREADY_ACTIVE = "Diese Vorlage ist bereits aktiv."
MSG_EMPTY_SHEET = "Die Datei enthält keine Spalten."
MSG_UNKNOWN_HEADERS = (
    "Unbekannte Spalten: {list}. Zulässige Spaltennamen siehe Feldübersicht."
)
MSG_MISSING_PROTECTED = "Pflichtspalten fehlen: {list}"
MSG_DUPLICATE_HEADERS = "Doppelte Spaltenüberschriften sind nicht erlaubt: {headers}."
MSG_NOTE_REQUIRED = "Grund der Änderung ist erforderlich."
MSG_XLSX_ONLY = "Nur .xlsx-Dateien werden akzeptiert."
MSG_NOT_ADMIN = "Nur Admins dürfen die Artikel-Vorlage ersetzen."
MSG_NO_PENDING = "Keine ausstehende Vorlage zur Bestätigung."
MSG_NO_ACTIVE = "Keine aktive Vorlage vorhanden."


class TemplateError(Exception):
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


class TemplatePermissionError(TemplateError):
    """Raised when a non-admin calls a mutating template operation."""


@dataclass
class TemplateDiff:
    added: list[str] = field(default_factory=list)
    removed: list[str] = field(default_factory=list)
    order_changed: bool = False
    draft_count: int = 0
    from_version: int = 0
    to_version: int = 0


@dataclass
class PendingTemplate:
    xlsx_bytes: bytes
    sha256: str
    note: str
    columns: list[dict[str, object]]
    diff: TemplateDiff


def require_template_admin(user: Mapping[str, Any]) -> None:
    """Thin service-layer guard; routes also use require_admin."""
    if "admin" not in (user.get("roles") or []):
        raise TemplatePermissionError(MSG_NOT_ADMIN)


def get_active_template(db: Session) -> ArticleTemplate:
    row = db.scalars(
        select(ArticleTemplate).where(ArticleTemplate.is_active.is_(True))
    ).first()
    if row is None:
        raise TemplateError(MSG_NO_ACTIVE)
    return row


def list_templates(db: Session) -> list[ArticleTemplate]:
    return list(
        db.scalars(
            select(ArticleTemplate).order_by(ArticleTemplate.version.desc())
        )
    )


def get_template(db: Session, template_id: uuid.UUID) -> ArticleTemplate | None:
    return db.get(ArticleTemplate, template_id)


def template_column_labels(template: ArticleTemplate) -> list[str]:
    columns = template.columns if isinstance(template.columns, list) else []
    return [str(col.get("label") or "") for col in columns if col.get("label")]


def template_column_keys(template: ArticleTemplate) -> list[str]:
    columns = template.columns if isinstance(template.columns, list) else []
    out: list[str] = []
    for col in columns:
        key = str(col.get("key") or col.get("label") or "")
        if key:
            out.append(key)
    return out


def template_required_labels(template: ArticleTemplate) -> list[str]:
    columns = template.columns if isinstance(template.columns, list) else []
    return [
        str(col.get("label") or "")
        for col in columns
        if col.get("required") and col.get("label")
    ]


def download_filename(template: ArticleTemplate) -> str:
    return f"prosema-artikel-vorlage-v{template.version}.xlsx"


def build_seed_workbook_bytes() -> bytes:
    headers = [field.label for field in FIELDS]
    examples = {field.label: field.example for field in FIELDS}
    return build_template_workbook(headers, examples=examples)


def _read_sheet1_headers(data: bytes) -> list[str]:
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except Exception as exc:
        raise TemplateError(f"Excel-Datei konnte nicht gelesen werden: {exc}") from exc
    try:
        if not wb.worksheets:
            raise TemplateError(MSG_EMPTY_SHEET)
        sheet = wb.worksheets[0]
        rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise TemplateError(MSG_EMPTY_SHEET) from exc
        return [normalize_label(cell) for cell in header_row]
    finally:
        wb.close()


def parse_template_headers(headers: Sequence[str]) -> list[dict[str, object]]:
    """Validate headers against the catalogue; return ordered column dicts."""
    non_empty = [h for h in headers if h]
    if not non_empty:
        raise TemplateError(MSG_EMPTY_SHEET)

    seen_fold: dict[str, str] = {}
    duplicates: list[str] = []
    unknown: list[str] = []
    columns: list[dict[str, object]] = []

    for header in non_empty:
        fold = header.casefold()
        if fold in seen_fold:
            duplicates.append(header)
            continue
        field = find_field(header)
        if field is None:
            unknown.append(header)
            continue
        seen_fold[fold] = field.label
        columns.append(
            {
                "key": field.key,
                "label": field.label,
                "required": field.required_for_upload,
            }
        )

    if duplicates:
        raise TemplateError(
            MSG_DUPLICATE_HEADERS.format(headers=", ".join(sorted(set(duplicates))))
        )
    if unknown:
        raise TemplateError(MSG_UNKNOWN_HEADERS.format(list=", ".join(unknown)))

    present = {col["label"] for col in columns}
    missing = [f.label for f in PROTECTED_FIELDS if f.label not in present]
    if missing:
        raise TemplateError(MSG_MISSING_PROTECTED.format(list=", ".join(missing)))

    return columns


def _labels(columns: Sequence[Mapping[str, object]]) -> list[str]:
    return [str(col.get("label") or "") for col in columns if col.get("label")]


def compute_diff(
    db: Session,
    *,
    old: ArticleTemplate,
    new_columns: Sequence[Mapping[str, object]],
) -> TemplateDiff:
    old_labels = _labels(old.columns if isinstance(old.columns, list) else [])
    new_labels = _labels(new_columns)
    old_set = set(old_labels)
    new_set = set(new_labels)
    draft_count = db.scalar(
        select(func.count())
        .select_from(ArticleBatch)
        .where(
            ArticleBatch.template_id == old.id,
            ArticleBatch.status == "draft",
        )
    )
    return TemplateDiff(
        added=sorted(new_set - old_set),
        removed=sorted(old_set - new_set),
        order_changed=old_labels != new_labels and old_set == new_set,
        draft_count=int(draft_count or 0),
        from_version=old.version,
        to_version=old.version + 1,
    )


def prepare_template_replacement(
    db: Session,
    *,
    user: Mapping[str, Any],
    filename: str,
    data: bytes,
    note: str,
) -> PendingTemplate:
    require_template_admin(user)
    if not filename.lower().endswith(".xlsx"):
        raise TemplateError(MSG_XLSX_ONLY)
    cleaned_note = str(note or "").strip()
    if not cleaned_note:
        raise TemplateError(MSG_NOTE_REQUIRED)

    active = get_active_template(db)
    digest = hashlib.sha256(data).hexdigest()
    if digest == active.sha256:
        raise TemplateError(MSG_ALREADY_ACTIVE)

    headers = _read_sheet1_headers(data)
    columns = parse_template_headers(headers)
    diff = compute_diff(db, old=active, new_columns=columns)
    return PendingTemplate(
        xlsx_bytes=data,
        sha256=digest,
        note=cleaned_note,
        columns=columns,
        diff=diff,
    )


def activate_template(
    db: Session,
    *,
    user: Mapping[str, Any],
    pending: PendingTemplate,
) -> ArticleTemplate:
    require_template_admin(user)
    active = get_active_template(db)
    if pending.sha256 == active.sha256:
        raise TemplateError(MSG_ALREADY_ACTIVE)

    # Re-validate in case the catalogue moved under a pending session.
    parse_template_headers(_labels(pending.columns))

    next_version = int(
        db.scalar(select(func.coalesce(func.max(ArticleTemplate.version), 0))) or 0
    ) + 1
    diff = compute_diff(db, old=active, new_columns=pending.columns)

    db.execute(
        update(ArticleTemplate)
        .where(ArticleTemplate.id == active.id)
        .values(is_active=False)
    )
    row = ArticleTemplate(
        version=next_version,
        is_active=True,
        columns=list(pending.columns),
        xlsx_bytes=pending.xlsx_bytes,
        sha256=pending.sha256,
        created_by_oid=str(user["oid"]),
        created_by_name=str(user.get("name") or user["oid"]),
        note=pending.note,
    )
    db.add(row)
    db.flush()

    record_audit_log(
        db,
        actor=user,
        entity_type="article_template",
        entity_id=row.id,
        action="activated",
        detail={
            "from_version": diff.from_version,
            "to_version": row.version,
            "note": pending.note,
            "added": diff.added,
            "removed": diff.removed,
            "order_changed": diff.order_changed,
            "column_count": len(pending.columns),
        },
    )
    return row


def pending_to_session(pending: PendingTemplate) -> dict[str, Any]:
    """Session payload without xlsx bytes (cookie size). Confirm re-uploads the file."""
    return {
        "sha256": pending.sha256,
        "note": pending.note,
        "columns": pending.columns,
        "diff": {
            "added": pending.diff.added,
            "removed": pending.diff.removed,
            "order_changed": pending.diff.order_changed,
            "draft_count": pending.diff.draft_count,
            "from_version": pending.diff.from_version,
            "to_version": pending.diff.to_version,
        },
    }


def pending_meta_from_session(payload: Mapping[str, Any] | None) -> dict[str, Any]:
    if not payload or "sha256" not in payload:
        raise TemplateError(MSG_NO_PENDING)
    return dict(payload)


def activate_from_upload(
    db: Session,
    *,
    user: Mapping[str, Any],
    data: bytes,
    session_pending: Mapping[str, Any],
) -> ArticleTemplate:
    """Second-step activation: bytes must match the sha256 from the confirm session."""
    require_template_admin(user)
    meta = pending_meta_from_session(session_pending)
    digest = hashlib.sha256(data).hexdigest()
    if digest != str(meta["sha256"]):
        raise TemplateError(
            "Die hochgeladene Datei stimmt nicht mit der bestätigten Vorlage überein."
        )
    pending = PendingTemplate(
        xlsx_bytes=data,
        sha256=digest,
        note=str(meta["note"]),
        columns=list(meta["columns"]),
        diff=TemplateDiff(
            added=list(meta["diff"].get("added") or []),
            removed=list(meta["diff"].get("removed") or []),
            order_changed=bool(meta["diff"].get("order_changed")),
            draft_count=int(meta["diff"].get("draft_count") or 0),
            from_version=int(meta["diff"].get("from_version") or 0),
            to_version=int(meta["diff"].get("to_version") or 0),
        ),
    )
    return activate_template(db, user=user, pending=pending)


def catalogue_for_display() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for entry in FIELDS:
        rows.append(
            {
                "label": entry.label,
                "required_for_upload": entry.required_for_upload,
                "in_payload": entry.in_payload,
                "never_editable": entry.never_editable,
                "protected": entry.protected,
                "protected_reason": entry.protected_reason,
                "field_type": entry.field_type,
                "example": entry.example,
                "description": entry.description,
            }
        )
    return rows


# Re-export for migration / tests
__all__ = [
    "MSG_ALREADY_ACTIVE",
    "MSG_NOT_ADMIN",
    "PendingTemplate",
    "TemplateDiff",
    "TemplateError",
    "TemplatePermissionError",
    "activate_template",
    "build_seed_workbook_bytes",
    "catalogue_for_display",
    "compute_diff",
    "download_filename",
    "get_active_template",
    "get_template",
    "list_templates",
    "parse_template_headers",
    "pending_meta_from_session",
    "pending_to_session",
    "prepare_template_replacement",
    "require_template_admin",
    "seed_template_columns",
    "template_column_keys",
    "template_column_labels",
    "template_required_labels",
]
