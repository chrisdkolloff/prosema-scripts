"""Artikelregistrierung: upload → approve → submit."""

from __future__ import annotations

import io
import csv
import json
from datetime import UTC, datetime, timedelta
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.batch_actions import MSG_NO_SNAPSHOT, MSG_SNAPSHOT_STALE, approve_batch, snapshot_age_warning
from app.batch_submit import MSG_DRY_RUN_FAILED, run_batch_submit
from app.batch_upload import BatchUploadError, MAX_UPLOAD_ROWS, create_batch_from_upload, create_manual_batch
from app.batches import CellEdit, apply_edits
from app.db import engine, get_db
from app.groups_service import create_hauptgruppe, create_untergruppe
from app.main import app
from app.models import (
    ArticleBatch,
    ArticleBatchRow,
    ArticleSnapshot,
    ArticleSnapshotRow,
    GruppenAudit,
    Hauptgruppe,
    Job,
)
from scripts.weclapp.article_import import IMPORT_COLUMNS
from scripts.weclapp.client import WeclappError

# Live weclapp POSTs (if any) must use this pair only — see docs/artikel-registrierung.md § acceptance 7.
TEST_WRITE_GROUP = ("999", "999")

ACTOR = {"oid": "reg-oid", "name": "Reg User"}
PLAIN_USER = {
    "oid": "reg-oid",
    "name": "Reg User",
    "email": "reg@example.com",
    "roles": ["user"],
}


@pytest.fixture
def db_session():
    connection = engine.connect()
    trans = connection.begin()
    session = Session(
        bind=connection,
        join_transaction_mode="create_savepoint",
        expire_on_commit=False,
    )
    try:
        yield session
    finally:
        session.close()
        trans.rollback()
        connection.close()


@pytest.fixture
def user_client(db_session):
    def override_user():
        return PLAIN_USER

    def override_db():
        yield db_session

    app.dependency_overrides[get_current_user] = override_user
    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    try:
        yield client
    finally:
        app.dependency_overrides.clear()


def _unused_code(db_session, prefix: str = "7") -> str:
    used = {row[0] for row in db_session.execute(select(Hauptgruppe.code)).all()}
    for index in range(100):
        code = f"{prefix}{index:02d}"
        if code not in used:
            return code
    raise RuntimeError("No free test code remaining")


def _make_groups(db_session):
    haupt_code = _unused_code(db_session)
    unter_code = "010"
    haupt = create_hauptgruppe(
        db_session, code=haupt_code, name=f"Haupt {haupt_code}", actor=ACTOR
    )
    unter = create_untergruppe(
        db_session,
        haupt,
        code=unter_code,
        name=f"Unter {unter_code}",
        actor=ACTOR,
    )
    db_session.flush()
    return haupt, unter


def _xlsx_bytes(rows: list[dict[str, str]], *, headers: list[str] | None = None) -> bytes:
    headers = headers or list(IMPORT_COLUMNS)
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    for row_idx, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row_idx, col, row.get(header, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(
    rows: list[dict[str, str]],
    *,
    headers: list[str] | None = None,
    delimiter: str = ",",
) -> bytes:
    headers = headers or list(IMPORT_COLUMNS)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, delimiter=delimiter, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: row.get(header, "") for header in headers})
    return buf.getvalue().encode("utf-8")


def _row(haupt_label: str, unter_label: str, **extra: str) -> dict[str, str]:
    data = {
        "Prosema-Artikelname": "Testartikel",
        "Hauptgruppe": haupt_label,
        "Untergruppe": unter_label,
        "Einheit": "Stk.",
        "Artikeltyp": "BASIC",
        "Aktiv": "Ja",
        "Im Verkauf": "Ja",
        "Steuersatz": "STANDARD",
    }
    data.update(extra)
    return data


def _make_snapshot(db_session, *, numbers: list[str] | None = None) -> ArticleSnapshot:
    snapshot = ArticleSnapshot(
        status="complete",
        created_by_oid=ACTOR["oid"],
        created_by_name=ACTOR["name"],
        weclapp_tenant="test",
        row_count=len(numbers or []),
        columns=[],
        non_conforming_number_count=0,
    )
    db_session.add(snapshot)
    db_session.flush()
    for position, number in enumerate(numbers or [], start=1):
        db_session.add(
            ArticleSnapshotRow(
                snapshot_id=snapshot.id,
                position=position,
                data={"Prosema Artikelnummer": number},
                article_number=number,
                article_name="Existing",
                active=True,
            )
        )
    db_session.flush()
    return snapshot


def _group_labels(haupt, unter) -> tuple[str, str]:
    return f"{haupt.name} - {haupt.code}", f"{unter.name} - {unter.code}"


def _clear_row_errors(db_session, batch_id) -> None:
    """Registry test groups are not weclapp list values; clear payload validation noise."""
    for row in db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch_id)
    ):
        row.validation_error = ""
    db_session.flush()


def _fake_payload(values, _lookups):
    number = values.get("Prosema-Artikelnummer") or values.get("Prosema Artikelnummer") or ""
    return {
        "articleNumber": number,
        "name": values.get("Prosema-Artikelname") or values.get("PROSEMA Kurztext") or "Test",
        "articleType": "BASIC",
        "unitId": "unit-1",
        "taxRateType": "STANDARD",
        "active": True,
        "availableInSale": True,
    }


def _approve(db_session, batch):
    _clear_row_errors(db_session, batch.id)
    with patch("app.batch_actions.row_to_payload", side_effect=_fake_payload):
        return approve_batch(db_session, batch, actor=ACTOR)


def test_upload_rejects_unsupported_format(db_session):
    with pytest.raises(BatchUploadError, match="Nur .xlsx- und .csv"):
        create_batch_from_upload(
            db_session,
            filename="bad.txt",
            data=b"not a spreadsheet",
            user=ACTOR,
            confirmed=True,
        )


def test_upload_csv_comma_delimiter(db_session):
    haupt, unter = _make_groups(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="import.csv",
        data=_csv_bytes([_row(h, u)], delimiter=","),
        user=ACTOR,
        confirmed=True,
    )
    rows = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == result.batch.id)
    ).all()
    assert len(rows) == 1
    assert rows[0].raw_data["PROSEMA Kurztext"] == "Testartikel"


def test_upload_csv_semicolon_delimiter(db_session):
    haupt, unter = _make_groups(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="import.csv",
        data=_csv_bytes([_row(h, u)], delimiter=";"),
        user=ACTOR,
        confirmed=True,
    )
    rows = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == result.batch.id)
    ).all()
    assert len(rows) == 1


def test_upload_missing_required_header_inserts_nothing(db_session):
    before_batches = db_session.scalar(select(ArticleBatch.id).limit(1))
    data = _xlsx_bytes(
        [{"PROSEMA Kurztext": "x"}],
        headers=["PROSEMA Kurztext", "Einheit"],
    )
    with pytest.raises(BatchUploadError, match="Pflichtspalten fehlen"):
        create_batch_from_upload(
            db_session, filename="bad.xlsx", data=data, user=ACTOR, confirmed=True
        )
    db_session.rollback()
    assert db_session.scalar(select(ArticleBatch.id).limit(1)) == before_batches


def test_upload_row_limit(db_session):
    haupt, unter = _make_groups(db_session)
    h, u = _group_labels(haupt, unter)
    too_many = [
        _row(h, u, **{"Prosema-Artikelname": f"A{i}"}) for i in range(MAX_UPLOAD_ROWS + 1)
    ]
    with pytest.raises(BatchUploadError, match="Maximal 2000"):
        create_batch_from_upload(
            db_session,
            filename="big.xlsx",
            data=_xlsx_bytes(too_many),
            user=ACTOR,
            confirmed=True,
        )
    ok_rows = [
        _row(h, u, **{"Prosema-Artikelname": f"B{i}"}) for i in range(MAX_UPLOAD_ROWS)
    ]
    result = create_batch_from_upload(
        db_session,
        filename="ok.xlsx",
        data=_xlsx_bytes(ok_rows),
        user=ACTOR,
        confirmed=True,
    )
    n = db_session.scalar(
        select(func.count()).where(ArticleBatchRow.batch_id == result.batch.id)
    )
    assert n == MAX_UPLOAD_ROWS


def test_article_number_survives_upload_grid_excel(db_session, user_client):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session, numbers=[f"{haupt.code}.{unter.code}.0010"])
    h, u = _group_labels(haupt, unter)
    # Force a code-looking supplier number into an unknown column and into GTIN.
    result = create_batch_from_upload(
        db_session,
        filename="codes.xlsx",
        data=_xlsx_bytes(
            [
                _row(
                    h,
                    u,
                    **{
                        "GTIN (EAN-Nummer)": "010.020.0010",
                        "Lieferantenartikelnummer": "010.020.0010",
                    },
                )
            ]
        ),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    batch = result.batch
    row = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
    ).one()
    assert row.raw_data["Lieferantenartikelnummer"] == "010.020.0010"
    assert row.proposed_article_number == f"{haupt.code}.{unter.code}.0020"

    response = user_client.get(f"/batches/{batch.id}/excel")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Artikel"]
    headers = [cell.value for cell in ws[1]]
    liefer_idx = headers.index("Lieferantenartikelnummer") + 1
    art_idx = headers.index("Artikelnummer") + 1
    assert ws.cell(2, liefer_idx).value == "010.020.0010"
    assert ws.cell(2, liefer_idx).number_format == "@"
    assert ws.cell(2, art_idx).value == row.proposed_article_number
    assert ws.cell(2, art_idx).number_format == "@"


def test_two_draft_batches_never_share_number(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session, numbers=[])
    h, u = _group_labels(haupt, unter)
    first = create_batch_from_upload(
        db_session,
        filename="a.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    second = create_batch_from_upload(
        db_session,
        filename="b.xlsx",
            data=_xlsx_bytes([_row(h, u, **{"Prosema-Artikelname": "Andere"})]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    n1 = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == first.batch.id)
    ).one().proposed_article_number
    n2 = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == second.batch.id)
    ).one().proposed_article_number
    assert n1 != n2
    assert n1.endswith(".0010")
    assert n2.endswith(".0020")


def test_freigeben_with_validation_error_refused(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="err.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    row = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == result.batch.id)
    ).one()
    row.validation_error = "Kaputt"
    db_session.flush()
    with pytest.raises(Exception, match="Freigabe nicht möglich"):
        approve_batch(db_session, result.batch, actor=ACTOR)


def test_approved_payload_survives_failed_submit(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="pay.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _approve(db_session, result.batch)
    db_session.flush()
    row = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == result.batch.id)
    ).one()
    before = json.dumps(row.approved_payload, sort_keys=True)

    client = MagicMock()
    client.get.return_value = {"result": [{"id": "existing"}]}
    with patch("app.batch_submit.weclapp_client_for", return_value=client):
        with pytest.raises(ValueError, match="Probelauf"):
            run_batch_submit(
                db_session,
                batch_id=result.batch.id,
                actor_oid=ACTOR["oid"],
                actor_name=ACTOR["name"],
            )
    db_session.refresh(row)
    assert json.dumps(row.approved_payload, sort_keys=True) == before
    assert db_session.get(ArticleBatch, result.batch.id).status == "approved"


def test_dry_run_collision_zero_posts(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="col.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _approve(db_session, result.batch)
    db_session.flush()

    client = MagicMock()
    client.get.return_value = {"result": [{"id": "taken"}]}
    with patch("app.batch_submit.weclapp_client_for", return_value=client):
        with pytest.raises(ValueError, match=MSG_DRY_RUN_FAILED):
            run_batch_submit(
                db_session,
                batch_id=result.batch.id,
                actor_oid=ACTOR["oid"],
            )
    client.post.assert_not_called()


def test_submit_interrupted_retry_posts_remaining(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    rows = [_row(h, u, **{"Prosema-Artikelname": f"Art {i}"}) for i in range(4)]
    result = create_batch_from_upload(
        db_session,
        filename="part.xlsx",
        data=_xlsx_bytes(rows),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _approve(db_session, result.batch)
    db_session.flush()

    batch_rows = list(
        db_session.scalars(
            select(ArticleBatchRow)
            .where(ArticleBatchRow.batch_id == result.batch.id)
            .order_by(ArticleBatchRow.position)
        )
    )
    for row in batch_rows[:2]:
        row.weclapp_article_id = f"wc-{row.position}"
        row.submitted_at = datetime.now(UTC)
        row.submitted_by_oid = ACTOR["oid"]
    db_session.flush()

    client = MagicMock()
    client.get.return_value = {"result": []}
    client.post.side_effect = [{"id": "new-3"}, {"id": "new-4"}]
    with patch("app.batch_submit.weclapp_client_for", return_value=client):
        run_batch_submit(
            db_session,
            batch_id=result.batch.id,
            actor_oid=ACTOR["oid"],
            actor_name=ACTOR["name"],
        )
    assert client.post.call_count == 2
    db_session.expire_all()
    batch = db_session.get(ArticleBatch, result.batch.id)
    assert batch.status == "submitted"
    ids = {
        row.weclapp_article_id
        for row in db_session.scalars(
            select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
        )
    }
    assert ids == {"wc-1", "wc-2", "new-3", "new-4"}


def test_licence_failure_keeps_written_ids(db_session):
    haupt, unter = _make_groups(db_session)
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    rows = [_row(h, u, **{"Prosema-Artikelname": f"Art {i}"}) for i in range(3)]
    result = create_batch_from_upload(
        db_session,
        filename="lic.xlsx",
        data=_xlsx_bytes(rows),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _approve(db_session, result.batch)
    db_session.flush()

    client = MagicMock()
    client.get.return_value = {"result": []}
    client.post.side_effect = [
        {"id": "ok-1"},
        WeclappError("forbidden", status_code=403),
    ]
    with patch("app.batch_submit.weclapp_client_for", return_value=client):
        with patch(
            "app.batch_submit.job_error_message",
            return_value="Keine weclapp-Lizenz zugewiesen. Aktuell hat vermutlich jemand anderes die Lizenz.",
        ):
            with pytest.raises(Exception):
                run_batch_submit(
                    db_session,
                    batch_id=result.batch.id,
                    actor_oid=ACTOR["oid"],
                )
    db_session.expire_all()
    batch = db_session.get(ArticleBatch, result.batch.id)
    assert batch.status == "approved"
    written = [
        row.weclapp_article_id
        for row in db_session.scalars(
            select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
        )
        if row.weclapp_article_id
    ]
    assert written == ["ok-1"]


def test_edit_submitted_rejected_at_service(db_session):
    from app.article_templates import get_active_template

    batch = ArticleBatch(
        status="submitted",
        created_by_oid=ACTOR["oid"],
        created_by_name=ACTOR["name"],
        filename="x.xlsx",
        template_id=get_active_template(db_session).id,
    )
    db_session.add(batch)
    db_session.flush()
    row = ArticleBatchRow(
        batch_id=batch.id,
        position=1,
        raw_data={"PROSEMA Kurztext": "x"},
        edits={},
    )
    db_session.add(row)
    db_session.flush()
    with pytest.raises(Exception, match="genehmigt"):
        apply_edits(
            db_session,
            batch,
            [CellEdit(row_id=row.id, field="PROSEMA Kurztext", value="y")],
        )


def test_freigeben_without_snapshot_refused(db_session):
    haupt, unter = _make_groups(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="nosnap.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _clear_row_errors(db_session, result.batch.id)
    with patch("app.batch_actions.latest_completed_snapshot", return_value=None):
        with pytest.raises(Exception, match=MSG_NO_SNAPSHOT):
            approve_batch(db_session, result.batch, actor=ACTOR)


def test_first_write_locks_group_and_audits(db_session):
    haupt, unter = _make_groups(db_session)
    assert haupt.locked_at is None
    _make_snapshot(db_session)
    h, u = _group_labels(haupt, unter)
    result = create_batch_from_upload(
        db_session,
        filename="lock.xlsx",
        data=_xlsx_bytes([_row(h, u)]),
        user=ACTOR,
        confirmed=True,
    )
    db_session.flush()
    _approve(db_session, result.batch)
    db_session.flush()

    client = MagicMock()
    client.get.return_value = {"result": []}
    client.post.return_value = {"id": "wc-new"}
    with patch("app.batch_submit.weclapp_client_for", return_value=client):
        run_batch_submit(
            db_session,
            batch_id=result.batch.id,
            actor_oid=ACTOR["oid"],
            actor_name=ACTOR["name"],
        )
    db_session.refresh(haupt)
    db_session.refresh(unter)
    assert haupt.locked_at is not None
    assert unter.locked_at is not None
    actions = [
        row.action
        for row in db_session.scalars(
            select(GruppenAudit).where(GruppenAudit.entity_id == haupt.id)
        )
    ]
    assert "locked_by_registration" in actions


def test_stub_text_gone(user_client):
    with patch("app.routes.tools.check_weclapp_access") as mock_access:
        mock_access.return_value = MagicMock(kind="ok", message="")
        response = user_client.get("/artikel-registrierung")
    assert response.status_code == 200
    assert "folgt in Woche 3" not in response.text
    assert "Excel- oder CSV-Datei hochladen" in response.text


def test_snapshot_age_warning_after_24_hours():
    snap = ArticleSnapshot(
        status="complete",
        created_by_oid=ACTOR["oid"],
        created_by_name=ACTOR["name"],
        weclapp_tenant="test",
        created_at=datetime.now(UTC) - timedelta(hours=30),
    )
    warning = snapshot_age_warning(snap)
    assert warning == MSG_SNAPSHOT_STALE.format(n=30)
    snap.created_at = datetime.now(UTC) - timedelta(hours=2)
    assert snapshot_age_warning(snap) is None
    assert snapshot_age_warning(None) is None


def test_stale_snapshot_warning_has_refresh_button(db_session, user_client):
    snapshot = _make_snapshot(db_session)
    snapshot.created_at = datetime.now(UTC) - timedelta(hours=30)
    batch = create_manual_batch(db_session, user=ACTOR, row_count=1)
    db_session.flush()

    response = user_client.get(f"/batches/{batch.id}")
    assert response.status_code == 200
    assert "30 Stunden alt" in response.text
    assert "Aktualisieren" in response.text
    assert f"/batches/{batch.id}/artikeluebersicht-aktualisieren" in response.text


@patch("app.config.settings.weclapp_tenant", "test")
def test_refresh_snapshot_from_batch_enqueues_pull(db_session, user_client):
    snapshot = _make_snapshot(db_session)
    snapshot.created_at = datetime.now(UTC) - timedelta(hours=30)
    batch = create_manual_batch(db_session, user=ACTOR, row_count=1)
    db_session.flush()

    before = db_session.scalar(
        select(func.count()).select_from(Job).where(Job.job_type == "weclapp_article_snapshot")
    )
    response = user_client.post(
        f"/batches/{batch.id}/artikeluebersicht-aktualisieren",
        follow_redirects=False,
    )
    after = db_session.scalar(
        select(func.count()).select_from(Job).where(Job.job_type == "weclapp_article_snapshot")
    )
    assert response.status_code == 303
    assert response.headers["location"] == f"/batches/{batch.id}"
    assert after == before + 1

    running = db_session.scalars(
        select(ArticleSnapshot).where(ArticleSnapshot.status == "running")
    ).all()
    assert len(running) == 1
