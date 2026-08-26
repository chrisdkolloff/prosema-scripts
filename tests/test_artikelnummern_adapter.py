"""Round-trip test for the Excel adapter (wiring only)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.numbering import Scheme
from scripts.processing.artikelnummern import ExcelLayout, assign_article_numbers


def _write_dictionary(path: Path) -> None:
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Hauptgruppen"
    ws_main.append(["Code", "Bezeichnung"])
    ws_main.append(["100", "Dichtungen"])

    ws_sub = wb.create_sheet("Untergruppen")
    ws_sub.append(["Hauptgruppe", "Untergruppe", "Bezeichnung"])
    ws_sub.append(["100", "010", "O-Ring"])
    wb.save(path)


def _write_master(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(
        [
            "Artikelnr.",
            "Hauptgruppe",
            "Untergruppe",
            "Prosema Artikelnummer",
        ]
    )
    ws.append(["A-1", "Dichtungen", "O-Ring", None])
    ws.append(["A-2", "Dichtungen", "O-Ring", "100.010.0030"])
    ws.append(["A-3", "Dichtungen", "O-Ring", None])
    wb.save(path)


def test_excel_adapter_round_trip(tmp_path: Path):
    dictionary = tmp_path / "gruppen.xlsx"
    master_in = tmp_path / "input.xlsx"
    master_out = tmp_path / "output.xlsx"
    _write_dictionary(dictionary)
    _write_master(master_in)

    assigned, ranges = assign_article_numbers(
        str(master_in),
        str(master_out),
        scheme=Scheme(),
        layout=ExcelLayout(),
        dictionary_path=dictionary,
        overwrite_existing=False,
        strict=True,
    )

    assert assigned == 2
    assert ranges["100.010"] == 50

    wb = load_workbook(master_out, data_only=True)
    ws = wb.active
    # Pass 1 saw existing 0030; blanks continue past that high-water mark.
    assert ws.cell(2, 4).value == "100.010.0040"
    assert ws.cell(3, 4).value == "100.010.0030"  # preserved
    assert ws.cell(4, 4).value == "100.010.0050"
    # Fixed master columns created and filled on data rows
    assert ws.cell(1, 5).value == "Währung"
    assert ws.cell(2, 5).value == "EUR"
