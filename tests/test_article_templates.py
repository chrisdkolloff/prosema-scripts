"""Artikel-Vorlage: versioned templates, upload binding, manual batches."""

from __future__ import annotations

import hashlib
import io
import uuid

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.article_templates import (
    MSG_ALREADY_ACTIVE,
    MSG_NOT_ADMIN,
    TemplatePermissionError,
    activate_template,
    get_active_template,
    prepare_template_replacement,
    require_template_admin,
)
from app.auth import get_current_user
from app.batch_upload import (
    create_batch_from_upload,
    create_manual_batch,
    exclude_empty_rows,
    parse_workbook_bytes,
)
from app.batches import CellEdit, apply_edits, effective_values
from app.db import engine, get_db
from app.groups_service import create_hauptgruppe, create_untergruppe
from app.main import app
from app.models import ArticleBatchRow, ArticleTemplate, AuditLog, Hauptgruppe
from core.article_fields import FIELDS

ACTOR = {"oid": "tpl-admin", "name": "Admin User", "roles": ["user", "admin"]}
PLAIN = {"oid": "tpl-user", "name": "Plain User", "email": "u@example.com", "roles": ["user"]}
ADMIN = {
    "oid": "tpl-admin",
    "name": "Admin User",
    "email": "a@example.com",
    "roles": ["user", "admin"],
}


@pytest.fixture
def db_session():
    connection = engine.connect()
    trans = connection.begin()
    session = Session(
        bind=connection,
        join_transaction_mode="create_savepoint",
        expire_on_commit=False,
    )
    try:
        yield session
    finally:
        session.close()
        trans.rollback()
        connection.close()


@pytest.fixture
def admin_client(db_session):
    def override_user():
        return ADMIN

    def override_db():
        yield db_session

    app.dependency_overrides[get_current_user] = override_user
    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    try:
        yield client
    finally:
        app.dependency_overrides.clear()


@pytest.fixture
def user_client(db_session):
    def override_user():
        return PLAIN

    def override_db():
        yield db_session

    app.dependency_overrides[get_current_user] = override_user
    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    try:
        yield client
    finally:
        app.dependency_overrides.clear()


def _unused_code(db_session, prefix: str = "8") -> str:
    used = {row[0] for row in db_session.execute(select(Hauptgruppe.code)).all()}
    for index in range(100):
        code = f"{prefix}{index:02d}"
        if code not in used:
            return code
    raise RuntimeError("No free test code remaining")


def _make_groups(db_session):
    haupt_code = _unused_code(db_session)
    unter_code = "010"
    haupt = create_hauptgruppe(
        db_session, code=haupt_code, name=f"Haupt {haupt_code}", actor=ACTOR
    )
    unter = create_untergruppe(
        db_session,
        haupt,
        code=unter_code,
        name=f"Unter {unter_code}",
        actor=ACTOR,
    )
    return haupt, unter


def _xlsx_from_headers(headers: list[str], rows: list[dict[str, str]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vorlage"
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    for row_idx, row in enumerate(rows or [], start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row_idx, col, row.get(header, ""))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _full_headers() -> list[str]:
    return [field.label for field in FIELDS]


def _minimal_row(haupt, unter) -> dict[str, str]:
    return {
        "Prosema-Artikelname": "Test Artikel",
        "Hauptgruppe": f"{haupt.name} - {haupt.code}",
        "Untergruppe": f"{unter.name} - {unter.code}",
        "Lieferantenartikelnummer": "SUP-1",
        "Einheit": "Stk.",
    }


# --- 1 seed / unique active -------------------------------------------------


def test_seeded_v1_active_and_partial_unique(db_session):
    active = get_active_template(db_session)
    assert active.version == 1
    assert active.is_active is True
    assert active.created_by_name == "System"
    assert len(active.columns) == len(FIELDS)

    with pytest.raises(IntegrityError), db_session.begin_nested():
        db_session.execute(
            text(
                """
                    INSERT INTO article_templates (
                        id, version, is_active, columns, xlsx_bytes, sha256,
                        created_by_name, note
                    ) VALUES (
                        :id, 999, true, CAST('[]' AS jsonb), :bytes, 'x',
                        'Other', 'dup'
                    )
                    """
            ),
            {"id": uuid.uuid4(), "bytes": b"abc"},
        )
        db_session.flush()

    still = get_active_template(db_session)
    assert still.is_active is True
    assert still.version == 1


# --- 2–5 template replace ---------------------------------------------------


def test_unknown_header_rejected(db_session):
    headers = _full_headers() + ["Erfundene Spalte"]
    data = _xlsx_from_headers(headers)
    with pytest.raises(Exception, match="Unbekannte Spalten"):
        prepare_template_replacement(
            db_session,
            user=ACTOR,
            filename="v.xlsx",
            data=data,
            note="Versuch",
        )
    assert get_active_template(db_session).version == 1


def test_hyphenated_prosema_headers_are_catalogue_fields():
    from app.article_templates import parse_template_headers
    from core.article_fields import find_field

    for label in (
        "Prosema-Artikelnummer",
        "Prosema-Artikelname",
        "Prosema-Langtext",
    ):
        field = find_field(label)
        assert field is not None
        assert field.label == label
    assert find_field("PROSEMA Kurztext").label == "Prosema-Artikelname"
    assert find_field("Prosema Artikelnummer").label == "Prosema-Artikelnummer"

    columns = parse_template_headers(
        [
            "Prosema-Artikelnummer",
            "Lieferantenartikelnummer",
            "Hauptgruppe",
            "Untergruppe",
            "Prosema-Artikelname",
            "Prosema-Langtext",
            "Kurzbeschreibung",
            "Referenz (Matchcode)",
            "GTIN (EAN-Nummer)",
            "Artikeltyp",
            "Einheit",
            "Kategorie",
            "Aktiv",
            "Im Verkauf",
            "Steuersatz",
            "Im Shop verfügbar",
            "Im Shop aktiv",
            "Bestand übertragen",
            "Gewichtseinheit",
            "Grundmaterial",
            "Oberfläche",
            "Farbe",
            "Produktfamilie",
            "Rabattcode",
            "Verkaufseinheit",
            "Verpackung",
            "VPE 1",
            "VPE 2",
            "VPE 3",
            "Breite in mm",
            "Länge in cm",
            "Höhe in mm",
            "Bodenleger",
            "Dachdecker",
            "Landschaftsgärtner",
            "Plattenleger",
            "Artikelbeschreibung HTML",
            "Nettogewicht kg",
            "Produkt-ID (Prosema)",
            "Varianten-ID (Prosema)",
        ]
    )
    assert [col["label"] for col in columns][0:6] == [
        "Prosema-Artikelnummer",
        "Lieferantenartikelnummer",
        "Hauptgruppe",
        "Untergruppe",
        "Prosema-Artikelname",
        "Prosema-Langtext",
    ]


def test_missing_protected_rejected(db_session):
    headers = [f.label for f in FIELDS if f.label != "Lieferantenartikelnummer"]
    data = _xlsx_from_headers(headers)
    with pytest.raises(Exception, match="Pflichtspalten fehlen"):
        prepare_template_replacement(
            db_session,
            user=ACTOR,
            filename="v.xlsx",
            data=data,
            note="ohne Lieferant",
        )


def test_first_post_never_activates_only_bestaetigt(db_session):
    drop = "Farbe"
    headers = [f.label for f in FIELDS if f.label != drop]
    data = _xlsx_from_headers(headers)
    before = get_active_template(db_session).version

    pending = prepare_template_replacement(
        db_session,
        user=ACTOR,
        filename="v.xlsx",
        data=data,
        note="Farbe raus",
    )
    assert get_active_template(db_session).version == before

    activated = activate_template(db_session, user=ACTOR, pending=pending)
    db_session.flush()
    assert activated.version == before + 1
    assert activated.is_active is True

    actives = list(
        db_session.scalars(select(ArticleTemplate).where(ArticleTemplate.is_active.is_(True)))
    )
    assert len(actives) == 1
    assert actives[0].id == activated.id

    inactive = db_session.scalars(
        select(ArticleTemplate).where(
            ArticleTemplate.version == before,
            ArticleTemplate.is_active.is_(False),
        )
    ).first()
    assert inactive is not None

    audits = list(
        db_session.scalars(
            select(AuditLog).where(
                AuditLog.entity_type == "article_template",
                AuditLog.action == "activated",
            )
        )
    )
    assert len(audits) == 1
    assert audits[0].detail["from_version"] == before
    assert audits[0].detail["to_version"] == activated.version
    assert drop in audits[0].detail["removed"]


# --- 6–8 / 12 batch pinning -------------------------------------------------


def test_draft_on_v1_survives_v2(db_session):
    haupt, unter = _make_groups(db_session)
    v1 = get_active_template(db_session)
    row_data = _minimal_row(haupt, unter)
    # include Farbe so v1 has it in raw_data
    row_data["Farbe"] = "Rot"
    data = _xlsx_from_headers(_full_headers(), [row_data])
    result = create_batch_from_upload(
        db_session, filename="a.xlsx", data=data, user=PLAIN, confirmed=True
    )
    batch = result.batch
    assert batch.template_id == v1.id
    batch_row = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
    ).first()
    assert batch_row.raw_data.get("Farbe") == "Rot"

    # activate v2 without Farbe
    headers = [f.label for f in FIELDS if f.label != "Farbe"]
    pending = prepare_template_replacement(
        db_session,
        user=ACTOR,
        filename="v2.xlsx",
        data=_xlsx_from_headers(headers),
        note="Farbe entfernt",
    )
    v2 = activate_template(db_session, user=ACTOR, pending=pending)
    db_session.flush()

    db_session.refresh(batch)
    assert batch.template_id == v1.id
    values = effective_values(batch_row)
    assert values.get("Farbe") == "Rot"

    # new upload binds to v2 and does not populate Farbe
    data2 = _xlsx_from_headers(headers, [{**row_data, "Farbe": "Blau"}])
    # Farbe header absent from file — if we add unknown Farbe it is kept but not template
    result2 = create_batch_from_upload(
        db_session, filename="b.xlsx", data=data2, user=PLAIN, confirmed=True
    )
    assert result2.batch.template_id == v2.id
    row2 = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == result2.batch.id)
    ).first()
    assert "Farbe" not in row2.raw_data


def test_download_bytes_identical(db_session):
    headers = [f.label for f in FIELDS if f.label != "Farbe"]
    data = _xlsx_from_headers(headers)
    pending = prepare_template_replacement(
        db_session,
        user=ACTOR,
        filename="v.xlsx",
        data=data,
        note="bytes",
    )
    activated = activate_template(db_session, user=ACTOR, pending=pending)
    db_session.flush()
    assert bytes(activated.xlsx_bytes) == data
    assert activated.sha256 == hashlib.sha256(data).hexdigest()


# --- 9 service-layer admin --------------------------------------------------


def test_non_admin_refused_at_service_layer(db_session):
    with pytest.raises(TemplatePermissionError, match=MSG_NOT_ADMIN):
        require_template_admin(PLAIN)
    with pytest.raises(TemplatePermissionError):
        prepare_template_replacement(
            db_session,
            user=PLAIN,
            filename="v.xlsx",
            data=_xlsx_from_headers(_full_headers()),
            note="nope",
        )


# --- 10–11 manual -----------------------------------------------------------


def test_manual_batch_empty_raw_edits(db_session):
    batch = create_manual_batch(db_session, user=PLAIN, row_count=20)
    rows = list(
        db_session.scalars(
            select(ArticleBatchRow)
            .where(ArticleBatchRow.batch_id == batch.id)
            .order_by(ArticleBatchRow.position)
        )
    )
    assert len(rows) == 20
    assert all(row.raw_data == {} for row in rows)
    assert batch.filename is None
    assert batch.source_bytes is None

    target = rows[0]
    apply_edits(
        db_session,
        batch,
        [CellEdit(row_id=target.id, field="PROSEMA Kurztext", value="Manuell")],
    )
    db_session.refresh(target)
    assert target.edits.get("PROSEMA Kurztext") == "Manuell"
    assert target.raw_data == {}


def test_exclude_empty_sets_include_false(db_session):
    batch = create_manual_batch(db_session, user=PLAIN, row_count=5)
    rows = list(
        db_session.scalars(
            select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
        )
    )
    apply_edits(
        db_session,
        batch,
        [CellEdit(row_id=rows[0].id, field="PROSEMA Kurztext", value="Keep")],
    )
    excluded = exclude_empty_rows(db_session, batch)
    assert excluded == 4
    db_session.refresh(rows[0])
    assert rows[0].include is True
    for row in rows[1:]:
        db_session.refresh(row)
        assert row.include is False
    assert db_session.scalar(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)
    ) is not None
    assert (
        db_session.scalar(
            select(ArticleBatchRow)
            .where(ArticleBatchRow.batch_id == batch.id)
            .with_only_columns(ArticleBatchRow.id)
        )
        is not None
    )
    count = len(
        list(db_session.scalars(select(ArticleBatchRow).where(ArticleBatchRow.batch_id == batch.id)))
    )
    assert count == 5


# --- 13 numbering identical under v1 and v2 ---------------------------------


def test_numbering_identical_v1_v2(db_session):
    haupt, unter = _make_groups(db_session)
    row_data = _minimal_row(haupt, unter)
    data_v1 = _xlsx_from_headers(_full_headers(), [row_data])
    r1 = create_batch_from_upload(
        db_session, filename="n1.xlsx", data=data_v1, user=PLAIN, confirmed=True
    )
    num1 = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == r1.batch.id)
    ).first().proposed_article_number

    # discard so high-water doesn't advance for a second live draft of same number
    r1.batch.status = "discarded"
    db_session.flush()

    headers = [f.label for f in FIELDS if f.label != "Farbe"]
    pending = prepare_template_replacement(
        db_session,
        user=ACTOR,
        filename="v2.xlsx",
        data=_xlsx_from_headers(headers),
        note="v2",
    )
    activate_template(db_session, user=ACTOR, pending=pending)
    db_session.flush()

    data_v2 = _xlsx_from_headers(headers, [row_data])
    r2 = create_batch_from_upload(
        db_session, filename="n2.xlsx", data=data_v2, user=PLAIN, confirmed=True
    )
    num2 = db_session.scalars(
        select(ArticleBatchRow).where(ArticleBatchRow.batch_id == r2.batch.id)
    ).first().proposed_article_number
    assert num1 == num2
    assert num1.startswith(f"{haupt.code}.{unter.code}.")


def test_sheet1_only_ignores_beispiel(db_session):
    template = get_active_template(db_session)
    haupt, unter = _make_groups(db_session)
    # build workbook with junk on Beispiel that would fail if read
    headers = _full_headers()
    wb = Workbook()
    ws = wb.active
    ws.title = "Vorlage"
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    row = _minimal_row(haupt, unter)
    for col, header in enumerate(headers, start=1):
        ws.cell(2, col, row.get(header, ""))
    beispiel = wb.create_sheet("Beispiel")
    beispiel.cell(1, 1, "NotARealHeader")
    beispiel.cell(2, 1, "junk")
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    parsed, unknown, _missing_opt = parse_workbook_bytes(data, template=template)
    assert len([r for r in parsed if r]) == 1
    assert unknown == []


def test_already_active_sha_rejected(db_session):
    active = get_active_template(db_session)
    with pytest.raises(Exception, match=MSG_ALREADY_ACTIVE):
        prepare_template_replacement(
            db_session,
            user=ACTOR,
            filename="same.xlsx",
            data=bytes(active.xlsx_bytes),
            note="same",
        )


def test_user_can_download_active_template(user_client, db_session):
    response = user_client.get("/artikel-registrierung/vorlage")
    assert response.status_code == 200
    active = get_active_template(db_session)
    assert response.content == bytes(active.xlsx_bytes)
