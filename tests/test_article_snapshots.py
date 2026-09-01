"""Artikelübersicht: snapshot pull, filtering, Excel, retention."""

from __future__ import annotations

import io
import json
import re
import uuid
from datetime import UTC, datetime, timedelta
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.assistant.examples import EXAMPLE_QUESTIONS
from app.assistant.service import MSG_UNVERIFIED
from app.auth import get_current_user
from app.config import settings
from app.db import engine, get_db
from app.jobs import HANDLERS
from app.main import app
from app.models import ArticleSnapshot, ArticleSnapshotRow, AssistantQuery, Job
from app.routes.snapshots import (
    MSG_FRAGE_NOT_FOUND,
    MSG_FRAGE_OTHER_SNAPSHOT,
    MSG_FRAGE_OTHER_USER,
    MSG_SELECTION_TRUNCATED,
    _viewer_context,
)
from app.snapshots import (
    GRID_PAGE_SIZE,
    SnapshotFilters,
    build_excel_workbook,
    build_grid_config,
    count_filtered_rows,
    fetch_filtered_rows,
    pull_snapshot_rows,
)
from core.article_flatten import (
    build_snapshot_columns,
    extract_indexed_fields,
    master_row_to_snapshot_data,
    snapshot_column_title,
)

PLAIN_USER = {
    "oid": "user-oid-snapshots",
    "name": "Christopher Kolloff",
    "email": "user@example.com",
    "roles": ["user"],
}

TENANT = "test-tenant"


@pytest.fixture
def db_session():
    connection = engine.connect()
    trans = connection.begin()
    session = Session(
        bind=connection,
        join_transaction_mode="create_savepoint",
        expire_on_commit=False,
    )
    try:
        yield session
    finally:
        session.close()
        trans.rollback()
        connection.close()


@pytest.fixture
def user_client(db_session):
    def override_user():
        return PLAIN_USER

    def override_db():
        yield db_session

    app.dependency_overrides[get_current_user] = override_user
    app.dependency_overrides[get_db] = override_db
    client = TestClient(app)
    try:
        yield client
    finally:
        app.dependency_overrides.clear()


def _sample_master(article_number: str, **extra: str) -> dict[str, str]:
    from scripts.weclapp.master_columns import EXPORT_COLUMNS

    row = {column: "" for column in EXPORT_COLUMNS}
    row["Prosema Artikelnummer"] = article_number
    row["PROSEMA Kurztext"] = extra.get("name", f"Artikel {article_number}")
    row["Hauptgruppe"] = extra.get("hauptgruppe", "Holz - 010")
    row["Untergruppe"] = extra.get("untergruppe", "Bretter - 020")
    row["weclapp Aktiv"] = extra.get("aktiv", "Ja")
    row["weclapp Artikel-ID"] = extra.get("weclapp_id", f"id-{article_number}")
    row["weclapp Version"] = extra.get("version", "42")
    return row


def _make_complete_snapshot(
    db_session,
    *,
    rows: list[dict[str, str]] | None = None,
    columns: list[dict] | None = None,
) -> ArticleSnapshot:
    if rows is None:
        rows = [
            master_row_to_snapshot_data(_sample_master("010.020.0010", hauptgruppe="Holz - 010")),
            master_row_to_snapshot_data(
                _sample_master(
                    "010.030.0001",
                    hauptgruppe="Holz - 010",
                    untergruppe="Latten - 030",
                    aktiv="Nein",
                )
            ),
            master_row_to_snapshot_data(
                _sample_master(
                    "020.010.0005",
                    hauptgruppe="Metall - 020",
                    untergruppe="Schrauben - 010",
                )
            ),
        ]
    columns = columns or build_snapshot_columns(rows)
    snapshot = ArticleSnapshot(
        status="complete",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        row_count=len(rows),
        columns=columns,
    )
    db_session.add(snapshot)
    db_session.flush()
    for position, data in enumerate(rows):
        fields = extract_indexed_fields(data)
        db_session.add(
            ArticleSnapshotRow(
                snapshot_id=snapshot.id,
                position=position,
                data=data,
                article_number=fields["article_number"],
                article_name=fields["article_name"],
                hauptgruppe_code=fields["hauptgruppe_code"],
                untergruppe_code=fields["untergruppe_code"],
                active=fields["active"],
                weclapp_id=fields["weclapp_id"],
                weclapp_version=fields["weclapp_version"],
            )
        )
    db_session.flush()
    return snapshot


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_worker_writes_rows_in_one_transaction_and_failure_leaves_no_rows(db_session):
    snapshot = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(snapshot)
    db_session.commit()

    articles = [
        {"id": "1", "articleNumber": "010.020.0010", "name": "A", "version": 1},
        {"id": "2", "articleNumber": "010.020.0011", "name": "B", "version": 2},
    ]
    client = MagicMock()

    with patch("app.weclapp.weclapp_client_for", return_value=client):
        with patch("app.snapshots.flatten_articles") as flat_mock:
            data_rows = [
                master_row_to_snapshot_data(_sample_master("010.020.0010")),
                master_row_to_snapshot_data(_sample_master("010.020.0011")),
            ]
            indexed = [extract_indexed_fields(row) for row in data_rows]
            columns = build_snapshot_columns(data_rows)
            flat_mock.return_value = (data_rows, indexed, columns)
            result = pull_snapshot_rows(db_session, snapshot, oid=PLAIN_USER["oid"])

    assert result["row_count"] == 2
    db_session.refresh(snapshot)
    assert snapshot.status == "complete"
    assert snapshot.row_count == 2
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id == snapshot.id)
        )
        == 2
    )

    failing = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(failing)
    db_session.commit()

    boom = MagicMock()
    boom.iter_pages.side_effect = RuntimeError("pull aborted")

    from app.snapshots import fail_snapshot

    with patch("app.weclapp.weclapp_client_for", return_value=boom):
        try:
            pull_snapshot_rows(db_session, failing, oid=PLAIN_USER["oid"])
        except RuntimeError:
            db_session.rollback()
            failing = db_session.get(ArticleSnapshot, failing.id)
            fail_snapshot(db_session, failing, "pull aborted")

    db_session.refresh(failing)
    assert failing.status == "failed"
    assert failing.error == "pull aborted"
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id == failing.id)
        )
        == 0
    )


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_post_abfragen_while_running_does_not_enqueue_second_job(db_session, user_client):
    running = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(running)
    db_session.commit()

    before = db_session.scalar(select(func.count()).select_from(Job))
    response = user_client.post("/artikel-uebersicht/abfragen", follow_redirects=False)
    after = db_session.scalar(select(func.count()).select_from(Job))
    assert after == before
    assert response.status_code == 303
    assert str(running.id) in response.headers["location"]


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_stand_banner_fresh_pull_is_green_list_link_is_yellow(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    fresh = user_client.get(f"/artikel-uebersicht/{snapshot.id}?neu=1")
    assert fresh.status_code == 200
    assert "snapshot-stand-banner--current" in fresh.text
    assert 'class="snapshot-stand-banner__close"' in fresh.text

    from_list = user_client.get(f"/artikel-uebersicht/{snapshot.id}")
    assert from_list.status_code == 200
    assert "snapshot-stand-banner--archive" in from_list.text
    assert 'class="snapshot-stand-banner__close"' not in from_list.text


@patch("app.routes.snapshots.create_snapshot_pull")
@patch("app.config.settings.weclapp_tenant", TENANT)
def test_post_abfragen_redirects_with_fresh_pull_flag(pull_mock, user_client):
    snapshot_id = uuid.uuid4()
    snapshot = ArticleSnapshot(
        id=snapshot_id,
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    pull_mock.return_value = (snapshot, MagicMock())

    response = user_client.post("/artikel-uebersicht/abfragen", follow_redirects=False)

    assert response.status_code == 303
    assert response.headers["location"] == f"/artikel-uebersicht/{snapshot_id}?neu=1"


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_running_snapshot_has_no_duplicate_abfrage_laeuft_copy(db_session, user_client):
    snapshot = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(snapshot)
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}")
    assert response.status_code == 200
    assert "Abfrage läuft" not in response.text
    assert "alert-info" not in response.text
    assert "Diese Seite aktualisiert sich automatisch." in response.text
    assert 'id="snapshot-status-panel"' in response.text

    poll = user_client.get(f"/artikel-uebersicht/{snapshot.id}/status")
    assert poll.status_code == 200
    assert poll.text == ""
    assert "HX-Redirect" not in poll.headers


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_status_poll_redirects_when_snapshot_complete(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    poll = user_client.get(f"/artikel-uebersicht/{snapshot.id}/status")
    assert poll.status_code == 200
    assert poll.headers["HX-Redirect"] == f"/artikel-uebersicht/{snapshot.id}"


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_filter_hauptgruppe_matches_count(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    response = user_client.get(
        f"/artikel-uebersicht/{snapshot.id}/zeilen?hauptgruppe=010&nur_aktive=0"
    )
    assert response.status_code == 200
    assert "Gefiltert: 2 von 3 Zeilen" in response.text

    filters = SnapshotFilters(hauptgruppe="010", nur_aktive=False)
    assert count_filtered_rows(db_session, snapshot.id, filters) == 2
    rows, total, _pages = fetch_filtered_rows(db_session, snapshot.id, filters)
    assert total == 2
    assert len(rows) == 2
    assert all(row.hauptgruppe_code == "010" for row in rows)


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_untergruppe_options_narrow_by_hauptgruppe(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    response = user_client.get(
        f"/artikel-uebersicht/{snapshot.id}/untergruppen?hauptgruppe=010"
    )
    assert response.status_code == 200
    assert 'name="untergruppe"' in response.text
    # Holz (010) has Bretter→020 and Latten→030 only.
    assert ">020<" in response.text
    assert ">030<" in response.text
    assert ">010<" not in response.text

    all_groups = user_client.get(f"/artikel-uebersicht/{snapshot.id}/untergruppen")
    assert ">020<" in all_groups.text
    assert ">030<" in all_groups.text
    assert ">010<" in all_groups.text


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_excel_matches_zeilen_filter(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    zeilen = user_client.get(f"/artikel-uebersicht/{snapshot.id}/zeilen?hauptgruppe=020")
    assert "Gefiltert: 1 von 3 Zeilen" in zeilen.text

    excel = user_client.get(f"/artikel-uebersicht/{snapshot.id}/excel?hauptgruppe=020")
    assert excel.status_code == 200
    wb = load_workbook(io.BytesIO(excel.content))
    ws = wb["Artikel"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    artikel_col = headers.index("Prosema-Artikelnummer") + 1
    numbers = {ws.cell(row, artikel_col).value for row in range(2, ws.max_row + 1)}
    assert numbers == {"020.010.0005"}


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_article_number_excel_text_format(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}/excel")
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb["Artikel"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    artikel_col = headers.index("Prosema-Artikelnummer") + 1
    cell = ws.cell(2, artikel_col)
    assert cell.number_format == "@"
    assert cell.value == "010.020.0010"
    assert isinstance(cell.value, str)


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_snapshot_uses_stored_columns_not_current_schema(db_session):
    stored_columns = [
        {"key": "Legacy-Spalte", "title": "Legacy-Spalte", "width": 120},
        {"key": "Prosema Artikelnummer", "title": "Prosema Artikelnummer", "width": 160},
    ]
    data = {"Legacy-Spalte": "alt", "Prosema Artikelnummer": "010.020.0010"}
    snapshot = ArticleSnapshot(
        status="complete",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        row_count=1,
        columns=stored_columns,
    )
    db_session.add(snapshot)
    db_session.flush()
    row = ArticleSnapshotRow(
        snapshot_id=snapshot.id,
        position=0,
        data=data,
        article_number="010.020.0010",
        article_name="Test",
        hauptgruppe_code="010",
        untergruppe_code="020",
        active=True,
        weclapp_id="1",
    )
    db_session.add(row)
    db_session.flush()

    config = build_grid_config(snapshot, [row])
    assert config["fields"] == ["Legacy-Spalte", "Prosema Artikelnummer"]
    assert [col["title"] for col in config["columns"]] == [
        "Legacy-Spalte",
        "Prosema-Artikelnummer",
    ]
    assert config["data"] == [["alt", "010.020.0010"]]


def test_snapshot_column_title_matches_registration_where_applicable():
    assert snapshot_column_title("Prosema Artikelnummer") == "Prosema-Artikelnummer"
    assert snapshot_column_title("PROSEMA Kurztext") == "Prosema-Artikelname"
    assert snapshot_column_title("PROSEMA Langtext") == "Prosema-Langtext"
    assert snapshot_column_title("Prosema-Artikelname") == "Prosema-Artikelname"
    assert snapshot_column_title("Breite mm") == "Breite in mm"
    assert snapshot_column_title("weclapp Aktiv") == "Aktiv"
    assert snapshot_column_title("Artikelnr.") == "Lieferantenartikelnummer"
    assert snapshot_column_title("Einkaufspreis EUR netto") == "Einkaufspreis EUR netto"
    assert snapshot_column_title("Nettoverkaufspreis CHF") == "Nettoverkaufspreis CHF"
    assert snapshot_column_title("Verkaufspreis €, BE") == "Nettoverkaufspreis CHF"
    assert snapshot_column_title("Legacy-Spalte") == "Legacy-Spalte"


def test_apply_master_column_renames_rewrites_legacy_price_header():
    from scripts.weclapp.master_columns import apply_master_column_renames

    headers, rows = apply_master_column_renames(
        ["Prosema Artikelnummer", "Verkaufspreis €, BE"],
        [{"Prosema Artikelnummer": "1", "Verkaufspreis €, BE": "9.00"}],
    )
    assert headers == ["Prosema Artikelnummer", "Nettoverkaufspreis CHF"]
    assert rows == [{"Prosema Artikelnummer": "1", "Nettoverkaufspreis CHF": "9.00"}]


def test_flatten_uses_registration_column_keys_and_titles():
    data = master_row_to_snapshot_data(_sample_master("010.020.0010"))
    assert data["Prosema-Artikelnummer"] == "010.020.0010"
    assert data["Prosema-Artikelname"] == "Artikel 010.020.0010"
    columns = build_snapshot_columns([data])
    by_key = {col["key"]: col["title"] for col in columns}
    assert by_key["Prosema-Artikelnummer"] == "Prosema-Artikelnummer"
    assert by_key["Prosema-Artikelname"] == "Prosema-Artikelname"
    assert by_key["Prosema-Langtext"] == "Prosema-Langtext"


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_grid_and_excel_show_registration_titles_for_historic_keys(db_session):
    stored_columns = [
        {"key": "Prosema Artikelnummer", "title": "Prosema Artikelnummer", "width": 160},
        {"key": "PROSEMA Kurztext", "title": "PROSEMA Kurztext", "width": 220},
        {"key": "PROSEMA Langtext", "title": "PROSEMA Langtext", "width": 280},
        {"key": "Einheit", "title": "Einheit", "width": 90},
        {"key": "Einkaufspreis EUR netto", "title": "Einkaufspreis EUR netto", "width": 150},
        {"key": "Verkaufspreis €, BE", "title": "Verkaufspreis €, BE", "width": 140},
    ]
    data = {
        "Prosema Artikelnummer": "010.020.0010",
        "PROSEMA Kurztext": "Testname",
        "PROSEMA Langtext": "Lang",
        "Einheit": "Stk.",
        "Einkaufspreis EUR netto": "1.00",
        "Verkaufspreis €, BE": "2.50",
    }
    snapshot = ArticleSnapshot(
        status="complete",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        row_count=1,
        columns=stored_columns,
    )
    db_session.add(snapshot)
    db_session.flush()
    row = ArticleSnapshotRow(
        snapshot_id=snapshot.id,
        position=0,
        data=data,
        article_number="010.020.0010",
        article_name="Testname",
        hauptgruppe_code="010",
        untergruppe_code="020",
        active=True,
        weclapp_id="1",
    )
    db_session.add(row)
    db_session.flush()

    expected_titles = [
        "Prosema-Artikelnummer",
        "Prosema-Artikelname",
        "Prosema-Langtext",
        "Einheit",
        "Einkaufspreis EUR netto",
        "Nettoverkaufspreis CHF",
    ]
    config = build_grid_config(snapshot, [row])
    assert config["fields"] == [col["key"] for col in stored_columns]
    assert [col["title"] for col in config["columns"]] == expected_titles
    assert config["data"] == [["010.020.0010", "Testname", "Lang", "Stk.", "1.00", "2.50"]]

    wb = build_excel_workbook(snapshot, [row], SnapshotFilters())
    ws = wb["Artikel"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert headers == expected_titles
    assert ws.cell(2, 1).value == "010.020.0010"
    assert ws.cell(2, 1).number_format == "@"


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_snapshot_routes_never_write_to_weclapp(db_session):
    snapshot = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(snapshot)
    db_session.commit()

    client = MagicMock()
    client.iter_pages.return_value = []
    client.get.return_value = {}
    client.post = MagicMock()
    client.put = MagicMock()

    with patch("app.weclapp.weclapp_client_for", return_value=client):
        pull_snapshot_rows(db_session, snapshot, oid=PLAIN_USER["oid"])

    client.post.assert_not_called()
    client.put.assert_not_called()


def _months_ago(months: int, *, day: int = 15) -> datetime:
    now = datetime.now(UTC)
    year = now.year
    month = now.month - months
    while month <= 0:
        month += 12
        year -= 1
    return datetime(year, month, day, 12, 0, tzinfo=UTC)


def _add_complete_snapshot(
    db_session: Session,
    *,
    created_at: datetime,
    number: str,
) -> ArticleSnapshot:
    snap = ArticleSnapshot(
        status="complete",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        row_count=1,
        created_at=created_at,
        columns=[
            {"key": "Prosema Artikelnummer", "title": "Prosema Artikelnummer", "width": 160}
        ],
    )
    db_session.add(snap)
    db_session.flush()
    db_session.add(
        ArticleSnapshotRow(
            snapshot_id=snap.id,
            position=0,
            data={"Prosema Artikelnummer": number},
            article_number=number,
            article_name="x",
            hauptgruppe_code="010",
            untergruppe_code="020",
            active=True,
            weclapp_id=number,
        )
    )
    return snap


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_deletes_21st_oldest_snapshot(db_session, caplog):
    ids: list[uuid.UUID] = []
    # All older than RETENTION_KEEP_DAYS so the count floor alone decides.
    base = (datetime.now(UTC) - timedelta(days=30)).replace(
        hour=12, minute=0, second=0, microsecond=0
    )
    for index in range(21):
        snap = _add_complete_snapshot(
            db_session,
            created_at=base + timedelta(minutes=index),
            number=f"n-{index}",
        )
        ids.append(snap.id)
    db_session.commit()

    from app.snapshots import apply_retention

    with caplog.at_level("INFO", logger="app.snapshots"):
        deleted = apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    assert len(deleted) == 1
    assert deleted[0] == ids[0]
    remaining = list(
        db_session.scalars(
            select(ArticleSnapshot).where(ArticleSnapshot.weclapp_tenant == TENANT)
        )
    )
    assert len(remaining) == 20
    assert db_session.get(ArticleSnapshot, ids[0]) is None
    kept_ids = [item.id for item in remaining]
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id.in_(kept_ids))
        )
        == 20
    )
    assert "snapshot retention removed 1 snapshots, 1 rows" in caplog.text


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_keeps_monthly_archive_beyond_recent_20(db_session):
    recent_ids: list[uuid.UUID] = []
    # Older than RETENTION_KEEP_DAYS so eviction is by count, not the day floor.
    base = (datetime.now(UTC) - timedelta(days=30)).replace(
        hour=12, minute=0, second=0, microsecond=0
    )
    for index in range(25):
        snap = _add_complete_snapshot(
            db_session,
            created_at=base + timedelta(minutes=index),
            number=f"recent-{index}",
        )
        recent_ids.append(snap.id)
    monthly = _add_complete_snapshot(
        db_session,
        created_at=_months_ago(3),
        number="monthly-3",
    )
    older_same_month = _add_complete_snapshot(
        db_session,
        created_at=_months_ago(3, day=5),
        number="monthly-3-older",
    )
    db_session.commit()

    from app.snapshots import apply_retention

    deleted = apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    remaining_ids = {
        item.id
        for item in db_session.scalars(
            select(ArticleSnapshot).where(ArticleSnapshot.weclapp_tenant == TENANT)
        )
    }
    assert monthly.id in remaining_ids
    assert older_same_month.id not in remaining_ids
    assert set(recent_ids[:5]).issubset(set(deleted))
    assert set(recent_ids[5:]).issubset(remaining_ids)
    assert older_same_month.id in set(deleted)
    assert (
        db_session.scalar(
            select(func.count()).where(
                ArticleSnapshotRow.snapshot_id == older_same_month.id
            )
        )
        == 0
    )
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id == monthly.id)
        )
        == 1
    )


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_drops_snapshots_older_than_12_months(db_session):
    now = datetime.now(UTC)
    base = datetime(now.year, now.month, 15, 12, 0, tzinfo=UTC)
    recent = [
        _add_complete_snapshot(
            db_session,
            created_at=base + timedelta(minutes=index),
            number=f"r-{index}",
        )
        for index in range(20)
    ]
    boundary_keep = _add_complete_snapshot(
        db_session,
        created_at=_months_ago(11),
        number="keep-11",
    )
    too_old = _add_complete_snapshot(
        db_session,
        created_at=_months_ago(12),
        number="drop-12",
    )
    db_session.commit()

    from app.snapshots import apply_retention

    deleted = apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    assert too_old.id in deleted
    assert boundary_keep.id not in deleted
    assert {item.id for item in recent}.isdisjoint(deleted)
    assert db_session.get(ArticleSnapshot, too_old.id) is None
    assert db_session.get(ArticleSnapshot, boundary_keep.id) is not None


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_keeps_snapshot_within_day_floor_outside_newest_20(db_session):
    """Keep a 10-day-old complete snapshot that sits outside the newest 20.

    (Prompt said 15-day-old; with RETENTION_KEEP_DAYS=14 that would be deleted.)
    """
    base = datetime.now(UTC).replace(hour=12, minute=0, second=0, microsecond=0)
    for index in range(20):
        _add_complete_snapshot(
            db_session,
            created_at=base + timedelta(minutes=index),
            number=f"recent-{index}",
        )
    within_floor = _add_complete_snapshot(
        db_session,
        created_at=base - timedelta(days=10),
        number="day-floor-10",
    )
    outside_floor = _add_complete_snapshot(
        db_session,
        created_at=base - timedelta(days=16),
        number="outside-floor-16",
    )
    db_session.commit()

    from app.snapshots import apply_retention

    deleted = apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    assert within_floor.id not in deleted
    assert db_session.get(ArticleSnapshot, within_floor.id) is not None
    assert outside_floor.id in deleted
    assert db_session.get(ArticleSnapshot, outside_floor.id) is None


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_deletes_incomplete_orphan_older_than_7_days(db_session, caplog):
    base = datetime.now(UTC)
    orphan = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        created_at=base - timedelta(days=8),
    )
    db_session.add(orphan)
    db_session.flush()
    db_session.add(
        ArticleSnapshotRow(
            snapshot_id=orphan.id,
            position=0,
            data={"Prosema Artikelnummer": "orphan"},
            article_number="orphan",
            article_name="x",
            hauptgruppe_code="010",
            untergruppe_code="020",
            active=True,
            weclapp_id="orphan",
        )
    )
    db_session.commit()

    from app.snapshots import apply_retention

    with caplog.at_level("INFO", logger="app.snapshots"):
        apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    assert db_session.get(ArticleSnapshot, orphan.id) is None
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id == orphan.id)
        )
        == 0
    )
    assert "snapshot retention removed 1 incomplete snapshots, 1 rows" in caplog.text


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_keeps_incomplete_snapshot_younger_than_orphan_days(db_session):
    base = datetime.now(UTC)
    in_flight = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
        created_at=base - timedelta(hours=1),
    )
    db_session.add(in_flight)
    db_session.commit()

    from app.snapshots import apply_retention

    apply_retention(db_session, tenant=TENANT)
    db_session.commit()

    assert db_session.get(ArticleSnapshot, in_flight.id) is not None


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_retention_failure_does_not_roll_back_snapshot(db_session, caplog):
    snapshot = ArticleSnapshot(
        status="running",
        created_by_oid=PLAIN_USER["oid"],
        created_by_name=PLAIN_USER["name"],
        weclapp_tenant=TENANT,
    )
    db_session.add(snapshot)
    db_session.commit()

    data_rows = [master_row_to_snapshot_data(_sample_master("010.020.0010"))]
    indexed = [extract_indexed_fields(row) for row in data_rows]
    columns = build_snapshot_columns(data_rows)
    client = MagicMock()

    with (
        patch("app.weclapp.weclapp_client_for", return_value=client),
        patch("app.snapshots.flatten_articles", return_value=(data_rows, indexed, columns)),
        patch("app.snapshots.apply_retention", side_effect=RuntimeError("disk")),
        caplog.at_level("ERROR", logger="app.snapshots"),
    ):
        result = pull_snapshot_rows(db_session, snapshot, oid=PLAIN_USER["oid"])

    db_session.refresh(snapshot)
    assert result["row_count"] == 1
    assert result["deleted_snapshots"] == []
    assert snapshot.status == "complete"
    assert snapshot.row_count == 1
    assert (
        db_session.scalar(
            select(func.count()).where(ArticleSnapshotRow.snapshot_id == snapshot.id)
        )
        == 1
    )
    assert "snapshot retention failed" in caplog.text


def test_job_handler_registered():
    assert "weclapp_article_snapshot" in HANDLERS


def _grid_article_numbers(html: str) -> list[str]:
    match = re.search(
        r'<script type="application/json" id="snapshot-grid-config">(.*?)</script>',
        html,
        re.DOTALL,
    )
    assert match is not None
    config = json.loads(match.group(1))
    fields = config["fields"]
    idx = next(
        i
        for i, key in enumerate(fields)
        if "Artikelnummer" in key or key == "article_number"
    )
    return [row[idx] for row in config["data"]]


def _request(frage: str | None = None, **params: str) -> MagicMock:
    query: dict[str, str] = dict(params)
    if frage is not None:
        query["frage"] = frage
    request = MagicMock()
    request.query_params = query
    return request


def _make_assistant_query(
    db_session,
    snapshot: ArticleSnapshot,
    *,
    numbers: list[str] | None,
    truncated: bool = False,
    user_oid: str = PLAIN_USER["oid"],
    question: str = "Welche Artikel?",
    outcome: str = "answered",
    answer_de: str | None = "Drei Artikel.",
) -> AssistantQuery:
    query = AssistantQuery(
        user_oid=user_oid,
        user_name=PLAIN_USER["name"],
        question_de=question,
        snapshot_id=snapshot.id,
        tool_calls=[],
        answer_de=answer_de,
        outcome=outcome,
        applied_article_numbers=numbers,
        applied_filter={"conditions": []},
        selection_truncated=truncated,
    )
    db_session.add(query)
    db_session.flush()
    return query


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_assistant_selection_renders_exactly_those_rows(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["010.020.0010", "020.010.0005"],
    )
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert "Gefiltert: 2 von 3 Zeilen" in response.text
    assert _grid_article_numbers(response.text) == ["010.020.0010", "020.010.0005"]


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_assistant_selection_ands_with_q(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["010.020.0010", "020.010.0005"],
    )
    db_session.commit()

    response = user_client.get(
        f"/artikel-uebersicht/{snapshot.id}?frage={query.id}&q=010.020"
    )
    assert "Gefiltert: 1 von 3 Zeilen" in response.text
    assert _grid_article_numbers(response.text) == ["010.020.0010"]


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_assistant_selection_paging(db_session, user_client):
    rows = [
        master_row_to_snapshot_data(
            _sample_master(f"010.020.{i:04d}", name=f"Artikel {i}")
        )
        for i in range(300)
    ]
    snapshot = _make_complete_snapshot(db_session, rows=rows)
    numbers = [f"010.020.{i:04d}" for i in range(300)]
    query = _make_assistant_query(db_session, snapshot, numbers=numbers)
    db_session.commit()

    page2 = user_client.get(
        f"/artikel-uebersicht/{snapshot.id}?frage={query.id}&seite=2"
    )
    assert page2.status_code == 200
    assert "Gefiltert: 300 von 300 Zeilen" in page2.text
    shown = _grid_article_numbers(page2.text)
    assert len(shown) == 50
    assert shown == [f"010.020.{i:04d}" for i in range(GRID_PAGE_SIZE, 300)]


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_truncated_selection_renders_unfiltered_with_hinweis(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(db_session, snapshot, numbers=None, truncated=True)
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert "Gefiltert: 2 von 3 Zeilen" in response.text
    ctx = _viewer_context(
        db_session, snapshot, PLAIN_USER, _request(frage=str(query.id))
    )
    assert ctx["assistant_truncated"] is True
    assert ctx["assistant_selection_count"] is None
    assert ctx["assistant_hinweis"] == MSG_SELECTION_TRUNCATED
    assert ctx["assistant_asked_at"] is not None
    assert ctx["assistant_question"] == "Welche Artikel?"


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_frage_from_other_user_is_ignored(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["020.010.0005"],
        user_oid="someone-else",
    )
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert "Gefiltert: 2 von 3 Zeilen" in response.text
    ctx = _viewer_context(
        db_session, snapshot, PLAIN_USER, _request(frage=str(query.id))
    )
    assert ctx["assistant_query_id"] is None
    assert ctx["assistant_hinweis"] == MSG_FRAGE_OTHER_USER


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_frage_from_other_snapshot_is_ignored(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    other = _make_complete_snapshot(db_session)
    query = _make_assistant_query(db_session, other, numbers=["020.010.0005"])
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert "Gefiltert: 2 von 3 Zeilen" in response.text
    ctx = _viewer_context(
        db_session, snapshot, PLAIN_USER, _request(frage=str(query.id))
    )
    assert ctx["assistant_query_id"] is None
    assert ctx["assistant_hinweis"] == MSG_FRAGE_OTHER_SNAPSHOT


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_nonexistent_frage_is_ignored(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()
    missing = uuid.uuid4()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={missing}")
    assert response.status_code == 200
    assert "Gefiltert: 2 von 3 Zeilen" in response.text
    ctx = _viewer_context(
        db_session, snapshot, PLAIN_USER, _request(frage=str(missing))
    )
    assert ctx["assistant_hinweis"] == MSG_FRAGE_NOT_FOUND


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_selection_forces_nur_aktive_false_unless_explicit(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["010.020.0010", "010.030.0001", "020.010.0005"],
    )
    db_session.commit()

    implicit = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert "Gefiltert: 3 von 3 Zeilen" in implicit.text
    assert "010.030.0001" in _grid_article_numbers(implicit.text)

    explicit = user_client.get(
        f"/artikel-uebersicht/{snapshot.id}?frage={query.id}&nur_aktive=1"
    )
    assert "Gefiltert: 2 von 3 Zeilen" in explicit.text
    assert "010.030.0001" not in _grid_article_numbers(explicit.text)


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_selection_number_absent_from_snapshot_matches_nothing(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(db_session, snapshot, numbers=["999.999.9999"])
    db_session.commit()

    response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert "Gefiltert: 0 von 3 Zeilen" in response.text


@patch("app.config.settings.weclapp_tenant", TENANT)
def test_excel_with_frage_exports_selection_and_records_question(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["020.010.0005"],
        question="Nur Metall?",
    )
    db_session.commit()

    excel = user_client.get(f"/artikel-uebersicht/{snapshot.id}/excel?frage={query.id}")
    assert excel.status_code == 200
    wb = load_workbook(io.BytesIO(excel.content))
    ws = wb["Artikel"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    artikel_col = headers.index("Prosema-Artikelnummer") + 1
    numbers = {ws.cell(row, artikel_col).value for row in range(2, ws.max_row + 1)}
    assert numbers == {"020.010.0005"}
    abfrage = wb["Abfrage"]
    pairs = {
        abfrage.cell(row, 1).value: abfrage.cell(row, 2).value
        for row in range(2, abfrage.max_row + 1)
    }
    assert pairs["Frage"] == "Nur Metall?"
    assert pairs["Datenstand"]


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_post_frage_on_non_current_snapshot_does_not_call_model(db_session, user_client):
    older = _make_complete_snapshot(db_session)
    older.created_at = datetime(2026, 1, 1, tzinfo=UTC)
    newer = _make_complete_snapshot(db_session)
    newer.created_at = datetime(2026, 8, 1, tzinfo=UTC)
    db_session.commit()
    assert newer.id != older.id

    with (
        patch("app.config.settings.assistant_enabled", True),
        patch("app.routes.snapshots.ask") as ask_mock,
    ):
        response = user_client.post(
            f"/artikel-uebersicht/{older.id}/frage",
            data={"frage": "Wie viele Artikel?"},
        )
    ask_mock.assert_not_called()
    assert response.status_code == 200


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_post_frage_disabled_does_not_call_model(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    with (
        patch("app.config.settings.assistant_enabled", False),
        patch("app.routes.snapshots.ask") as ask_mock,
    ):
        response = user_client.post(
            f"/artikel-uebersicht/{snapshot.id}/frage",
            data={"frage": "Wie viele Artikel?"},
        )
    ask_mock.assert_not_called()
    assert response.status_code == 200


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_post_frage_redirects_with_audit_id(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()
    audit_id = uuid.uuid4()
    result = MagicMock()
    result.audit_id = audit_id

    with (
        patch("app.config.settings.assistant_enabled", True),
        patch("app.routes.snapshots.ask", return_value=result) as ask_mock,
    ):
        response = user_client.post(
            f"/artikel-uebersicht/{snapshot.id}/frage?q=holz",
            data={"frage": "Welche Artikel?"},
            follow_redirects=False,
        )
    ask_mock.assert_called_once()
    assert response.status_code == 303
    location = response.headers["location"]
    assert f"frage={audit_id}" in location
    assert "q=holz" in location
    assert str(snapshot.id) in location


def _href_before_label(html: str, label: str) -> str:
    match = re.search(
        rf'href="([^"]*)"[^>]*>\s*{re.escape(label)}',
        html,
    )
    assert match is not None, f"no href for {label!r}"
    return match.group(1)


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_card_renders_on_current_snapshot_when_enabled(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}")
    assert response.status_code == 200
    assert 'id="snapshot-frage-card"' in response.text
    assert 'name="frage"' in response.text
    assert (
        f"Stell {settings.assistant_name} eine Frage zur Artikelliste."
        in response.text
    )
    assert 'id="snapshot-frage-examples"' in response.text
    examples = json.loads(
        re.search(
            r'<script type="application/json" id="snapshot-frage-examples">(.*?)</script>',
            response.text,
            re.DOTALL,
        ).group(1)
    )
    assert examples == list(EXAMPLE_QUESTIONS)
    assert EXAMPLE_QUESTIONS[0] in response.text
    assert 'hx-boost="false"' in response.text
    assert 'id="snapshot-frage-status"' in response.text
    assert 'id="snapshot-frage-running-banner"' in response.text
    assert "Suchanfrage läuft..." in response.text
    assert "is-frage-busy" in response.text
    assert "snapshot_grid.js?v=" in response.text
    assert 'id="snapshot-frage"' in response.text
    assert re.search(r'id="snapshot-frage"[^>]*value=""', response.text)


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_card_hidden_on_archived_snapshot(db_session, user_client):
    older = _make_complete_snapshot(db_session)
    older.created_at = datetime(2026, 1, 1, tzinfo=UTC)
    newer = _make_complete_snapshot(db_session)
    newer.created_at = datetime(2026, 8, 1, tzinfo=UTC)
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{older.id}")
    assert response.status_code == 200
    assert 'id="snapshot-frage-card"' not in response.text
    assert 'name="frage"' not in response.text

    with patch("app.config.settings.assistant_enabled", True):
        current = user_client.get(f"/artikel-uebersicht/{newer.id}")
    assert 'id="snapshot-frage-card"' in current.text


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_card_hidden_when_assistant_disabled(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", False):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}")
    assert response.status_code == 200
    assert 'id="snapshot-frage-card"' not in response.text
    assert 'name="frage"' not in response.text


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_banner_shows_question_timestamp_and_count(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["010.020.0010", "020.010.0005"],
        question="Welche Artikel von Dural?",
    )
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert 'id="snapshot-frage-banner"' in response.text
    assert "alert-info" in response.text
    assert "Auswahl aus der Frage vom" in response.text
    assert "«Welche Artikel von Dural?»" in response.text
    assert f"{settings.assistant_name} hat 2 Artikel gefunden." in response.text
    assert "Auswahl durch die Frage" in response.text
    match = re.search(r'id="snapshot-frage"[^>]*value="([^"]*)"', response.text)
    assert match is not None
    assert match.group(1) == "Welche Artikel von Dural?"


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_banner_unverified_is_warning_without_empty_prose(
    db_session, user_client
):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session,
        snapshot,
        numbers=["020.010.0005"],
        outcome="answered_unverified",
        answer_de=None,
    )
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert "alert-warning" in response.text
    assert MSG_UNVERIFIED in response.text
    assert f"{settings.assistant_name} hat 1 Artikel gefunden." in response.text
    assert "«Welche Artikel?»" in response.text
    banner = response.text.split('id="snapshot-frage-banner"', 1)[1]
    banner = banner.split("</div>", 2)[0]
    assert "<p class=\"mb-1\"></p>" not in banner
    assert not re.search(r"<p class=\"mb-1\">\s*</p>", banner)


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_frage_banner_truncated_explains_no_selection(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session, snapshot, numbers=None, truncated=True
    )
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    assert response.status_code == 200
    assert "Das Ergebnis ist zu gross, um es als Auswahl zu setzen" in response.text
    assert "Es ist keine Auswahl aktiv." in response.text
    assert "Stell eine engere Frage." in response.text
    assert "Artikel ausgewählt" not in response.text
    assert "Auswahl durch die Frage" not in response.text


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_auswahl_aufheben_drops_frage_keeps_q(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session, snapshot, numbers=["020.010.0005"]
    )
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(
            f"/artikel-uebersicht/{snapshot.id}?frage={query.id}&q=holz"
        )
    href = _href_before_label(response.text, "Auswahl aufheben")
    assert "frage=" not in href
    assert "q=holz" in href
    assert str(snapshot.id) in href


@patch("app.config.settings.weclapp_tenant", TENANT)
@patch("app.assistant.tools.settings.weclapp_tenant", TENANT)
def test_excel_link_includes_frage_when_selection_active(db_session, user_client):
    snapshot = _make_complete_snapshot(db_session)
    query = _make_assistant_query(
        db_session, snapshot, numbers=["020.010.0005"]
    )
    db_session.commit()

    with patch("app.config.settings.assistant_enabled", True):
        response = user_client.get(f"/artikel-uebersicht/{snapshot.id}?frage={query.id}")
    match = re.search(
        r'href="(/artikel-uebersicht/[^"]+/excel[^"]*)"',
        response.text,
    )
    assert match is not None
    assert f"frage={query.id}" in match.group(1)
