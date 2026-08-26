"""Gruppenschlüssel laden und parsen."""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook

from core.numbering import GroupDictionary, normalize_group_code, normalize_group_name


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _find_column(headers: Sequence[object | None], header_name: str) -> int:
    """Return 0-based index of ``header_name`` in a header row."""
    target = _normalize_header(header_name)
    for idx, cell in enumerate(headers):
        if _normalize_header(cell) == target:
            return idx
    present = [
        str(cell).strip()
        for cell in headers
        if cell not in (None, "")
    ]
    raise ValueError(
        f"Spalte {header_name!r} nicht gefunden. "
        f"Vorhandene Spaltenüberschriften: {', '.join(present) or '(keine)'}"
    )


def build_main_name_to_code(
    entries: Sequence[tuple[object | None, object | None]],
) -> dict[str, str]:
    """Build Hauptgruppe name→code map from (code, name) pairs.

    Empty code or name rows are skipped. Duplicate names raise.
    """
    main_name_to_code: dict[str, str] = {}
    for code_raw, name_raw in entries:
        if code_raw in (None, "") or name_raw in (None, ""):
            continue
        code = normalize_group_code(code_raw)
        key = normalize_group_name(name_raw)
        if key in main_name_to_code:
            raise ValueError(
                f"Doppelte Hauptgruppen-Bezeichnung im Gruppenschlüssel: {name_raw!r} "
                f"(Codes {main_name_to_code[key]!r} und {code!r})."
            )
        main_name_to_code[key] = code
    return main_name_to_code


def build_sub_name_to_code(
    entries: Sequence[tuple[object | None, object | None, object | None]],
) -> dict[tuple[str, str], str]:
    """Build Untergruppe (main_code, name)→code map from (main, sub, name) triples.

    Empty rows are skipped. Duplicate (Hauptgruppe, Bezeichnung) raise.
    """
    sub_name_to_code: dict[tuple[str, str], str] = {}
    for main_code_raw, sub_code_raw, sub_name_raw in entries:
        if main_code_raw in (None, "") or sub_code_raw in (None, "") or sub_name_raw in (None, ""):
            continue
        main_code = normalize_group_code(main_code_raw)
        sub_code = normalize_group_code(sub_code_raw)
        key = (main_code, normalize_group_name(sub_name_raw))
        if key in sub_name_to_code:
            raise ValueError(
                f"Doppelte Untergruppen-Bezeichnung im Gruppenschlüssel: "
                f"Hauptgruppe {main_code!r}, Bezeichnung {sub_name_raw!r} "
                f"(Codes {sub_name_to_code[key]!r} und {sub_code!r})."
            )
        sub_name_to_code[key] = sub_code
    return sub_name_to_code


def parse_group_dictionary(
    main_entries: Sequence[tuple[object | None, object | None]],
    sub_entries: Sequence[tuple[object | None, object | None, object | None]],
) -> GroupDictionary:
    """Parse already-extracted dictionary rows (no workbook needed)."""
    return GroupDictionary(
        main_name_to_code=build_main_name_to_code(main_entries),
        sub_name_to_code=build_sub_name_to_code(sub_entries),
    )


def load_group_dictionary(path: Path) -> GroupDictionary:
    """Load Gruppenschlüssel from an Excel workbook at an absolute path."""
    if not path.exists():
        raise FileNotFoundError(
            f"Gruppenschlüssel nicht gefunden: {path}\n"
            "Die Gruppencodes können ohne diese Datei nicht aufgelöst werden."
        )

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Hauptgruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Hauptgruppen".')
        if "Untergruppen" not in wb.sheetnames:
            raise ValueError('Gruppenschlüssel enthält kein Tabellenblatt "Untergruppen".')

        ws_main = wb["Hauptgruppen"]
        main_rows = list(ws_main.iter_rows(values_only=True))
        if not main_rows:
            raise ValueError('Tabellenblatt "Hauptgruppen" ist leer.')
        main_headers = main_rows[0]
        code_idx = _find_column(main_headers, "Code")
        name_idx = _find_column(main_headers, "Bezeichnung")
        main_entries = [
            (row[code_idx] if code_idx < len(row) else None,
             row[name_idx] if name_idx < len(row) else None)
            for row in main_rows[1:]
        ]

        ws_sub = wb["Untergruppen"]
        sub_rows = list(ws_sub.iter_rows(values_only=True))
        if not sub_rows:
            raise ValueError('Tabellenblatt "Untergruppen" ist leer.')
        sub_headers = sub_rows[0]
        main_code_idx = _find_column(sub_headers, "Hauptgruppe")
        sub_code_idx = _find_column(sub_headers, "Untergruppe")
        sub_name_idx = _find_column(sub_headers, "Bezeichnung")
        sub_entries = [
            (
                row[main_code_idx] if main_code_idx < len(row) else None,
                row[sub_code_idx] if sub_code_idx < len(row) else None,
                row[sub_name_idx] if sub_name_idx < len(row) else None,
            )
            for row in sub_rows[1:]
        ]

        return parse_group_dictionary(main_entries, sub_entries)
    finally:
        wb.close()
